"""Full-feature Korean voice classifier: MFCC-DTW + duration/RMS/ZCR/spectral/pitch features + STT alias.
Generated for testing short Korean consonant/vowel responses.
"""

import sys
import time
import difflib
import unicodedata
import queue
import wave
import pickle
import re
import json
import random
from pathlib import Path
from collections import Counter, defaultdict

import numpy as np
import sounddevice as sd
import webrtcvad

from scipy.spatial.distance import cdist
from scipy.signal import get_window
from python_speech_features import mfcc, delta
from faster_whisper import WhisperModel

try:
    from sklearn.ensemble import RandomForestClassifier, ExtraTreesClassifier
    import joblib
    HAS_SKLEARN = True
except Exception:
    HAS_SKLEARN = False
    RandomForestClassifier = None
    ExtraTreesClassifier = None
    joblib = None

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False
    load_workbook = None

from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QLabel,
    QVBoxLayout,
    QHBoxLayout,
    QTextEdit,
    QLineEdit,
    QSpinBox,
    QDoubleSpinBox,
    QComboBox,
    QCheckBox,
    QMessageBox,
    QInputDialog,
    QGroupBox,
    QGridLayout,
    QProgressBar,
)

# ============================================================
# Label Sets
# ============================================================


# ============================================================
# Audio Settings
# ============================================================

SAMPLE_RATE = 16000
FRAME_MS = 30
FRAME_SIZE = int(SAMPLE_RATE * FRAME_MS / 1000)
CHANNELS = 1
BASE_DIR = Path("voice_profiles")
BASE_DIR.mkdir(exist_ok=True)

# ============================================================
# Text Normalization / Alias Utility
# ============================================================

def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = text.strip().lower()
    text = re.sub(r"[\s\.,!?~·…\-_'\"`]+", "", text)
    text = text.replace("쌍 기억", "쌍기역")
    text = text.replace("쌍 디귿", "쌍디귿")
    text = text.replace("쌍 비읍", "쌍비읍")
    text = text.replace("쌍 시옷", "쌍시옷")
    text = text.replace("쌍 지읒", "쌍지읒")
    return text


def read_wav_int16(path):
    """Read a WAV file and return mono int16 audio for feedback learning."""
    with wave.open(str(path), "rb") as wf:
        n_channels = wf.getnchannels()
        sampwidth = wf.getsampwidth()
        n_frames = wf.getnframes()
        raw = wf.readframes(n_frames)
    if sampwidth != 2:
        raise ValueError("Only 16-bit PCM WAV files are supported for feedback learning.")
    audio = np.frombuffer(raw, dtype=np.int16)
    if n_channels > 1:
        audio = audio.reshape(-1, n_channels).mean(axis=1).astype(np.int16)
    return audio

MANUAL_ALIASES = {
    "기역": ["기역", "기억", "기어", "기역이"],
    "쌍기역": ["쌍기역", "쌍기억", "쌍 기억", "상기역"],
    "니은": ["니은", "니언", "니은이"],
    "디귿": ["디귿", "디귿이", "디긋", "디귿"],
    "쌍디귿": ["쌍디귿", "쌍디긋", "쌍 디귿", "상디귿"],
    "리을": ["리을", "리울", "리을이"],
    "미음": ["미음", "미움", "미음이"],
    "비읍": ["비읍", "비업", "비음", "비읍이"],
    "쌍비읍": ["쌍비읍", "쌍비업", "쌍 비읍", "상비읍"],
    "시옷": ["시옷", "시옷이", "시옷"],
    "쌍시옷": ["쌍시옷", "쌍 시옷", "상시옷"],
    "이응": ["이응", "이응이"],
    "지읒": ["지읒", "지읏", "지읒이"],
    "쌍지읒": ["쌍지읒", "쌍지읏", "쌍 지읒", "상지읒"],
    "치읓": ["치읓", "치읏", "치읓이"],
    "키읔": ["키읔", "키윽", "키읔이"],
    "티읕": ["티읕", "티읏", "티읕이"],
    "피읖": ["피읖", "피읍", "피읖이"],
    "히읗": ["히읗", "히읏", "이읗", "히읗이"],
}

def text_similarity(a: str, b: str) -> float:
    """Simple character bigram similarity in [0, 1]."""
    a = normalize_text(a)
    b = normalize_text(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    def grams(s):
        if len(s) == 1:
            return {s}
        return {s[i:i+2] for i in range(len(s) - 1)}
    ga, gb = grams(a), grams(b)
    return len(ga & gb) / max(1, len(ga | gb))

# ============================================================
# Tunable Audio Recorder with VAD + RMS Gate
# ============================================================

class AudioRecorder:
    def __init__(
        self,
        sample_rate=16000,
        frame_ms=30,
        vad_level=3,
        speech_frames_required=4,
        pre_buffer_frames=15,
        max_wait_sec=8.0,
        max_record_sec=4.0,
        end_silence_sec=0.8,
        noise_calibration_sec=0.45,
        energy_multiplier=2.5,
        min_energy_rms=120.0,
        min_record_sec=0.25,
    ):
        self.sample_rate = sample_rate
        self.frame_ms = frame_ms
        self.frame_size = int(sample_rate * frame_ms / 1000)
        self.vad_level = int(vad_level)
        self.vad = webrtcvad.Vad(self.vad_level)

        self.speech_frames_required = int(speech_frames_required)
        self.pre_buffer_frames = int(pre_buffer_frames)
        self.max_wait_sec = float(max_wait_sec)
        self.max_record_sec = float(max_record_sec)
        self.end_silence_frames = int(float(end_silence_sec) * 1000 / frame_ms)
        self.noise_calibration_sec = float(noise_calibration_sec)
        self.energy_multiplier = float(energy_multiplier)
        self.min_energy_rms = float(min_energy_rms)
        self.min_record_sec = float(min_record_sec)

        self.last_noise_rms = None
        self.last_energy_threshold = None
        self.last_duration_sec = None

    @staticmethod
    def _rms(frame):
        x = frame.astype(np.float32).flatten()
        return float(np.sqrt(np.mean(x * x) + 1e-8))

    def record_until_silence(self):
        q_audio = queue.Queue()

        def callback(indata, frames, time_info, status):
            q_audio.put(indata.copy())

        pre_speech_frames = []
        speech_frames = []
        noise_rms_values = []

        triggered = False
        speech_count = 0
        silence_count = 0

        record_start_time = time.perf_counter()
        voice_start_time = None
        noise_end_time = record_start_time + self.noise_calibration_sec

        # Noise gate is initialized after the short background-noise measurement.
        energy_threshold = None

        with sd.InputStream(
            samplerate=self.sample_rate,
            channels=1,
            dtype="int16",
            blocksize=self.frame_size,
            callback=callback,
        ):
            while True:
                now = time.perf_counter()

                if not triggered and now - record_start_time > self.max_wait_sec:
                    return None, None

                if triggered and now - voice_start_time > self.max_record_sec:
                    break

                try:
                    frame = q_audio.get(timeout=0.1)
                except queue.Empty:
                    continue

                frame_rms = self._rms(frame)

                # 1) Measure background noise first.
                if not triggered and now < noise_end_time:
                    noise_rms_values.append(frame_rms)
                    pre_speech_frames.append(frame)
                    if len(pre_speech_frames) > self.pre_buffer_frames:
                        pre_speech_frames.pop(0)
                    continue

                if energy_threshold is None:
                    noise_rms = float(np.median(noise_rms_values)) if noise_rms_values else 0.0
                    energy_threshold = max(self.min_energy_rms, noise_rms * self.energy_multiplier)
                    self.last_noise_rms = noise_rms
                    self.last_energy_threshold = energy_threshold

                vad_speech = self.vad.is_speech(frame.tobytes(), self.sample_rate)
                energy_speech = frame_rms >= energy_threshold
                is_speech = vad_speech and energy_speech

                if not triggered:
                    pre_speech_frames.append(frame)
                    if len(pre_speech_frames) > self.pre_buffer_frames:
                        pre_speech_frames.pop(0)

                    if is_speech:
                        speech_count += 1
                    else:
                        speech_count = 0

                    if speech_count >= self.speech_frames_required:
                        triggered = True
                        voice_start_time = time.perf_counter()
                        speech_frames.extend(pre_speech_frames)
                        speech_frames.append(frame)
                else:
                    speech_frames.append(frame)
                    # Ending can be a bit more permissive: VAD OR energy keeps recording.
                    keep_speech = vad_speech or energy_speech
                    if keep_speech:
                        silence_count = 0
                    else:
                        silence_count += 1

                    current_duration = (len(speech_frames) * self.frame_ms) / 1000.0
                    if current_duration >= self.min_record_sec and silence_count >= self.end_silence_frames:
                        break

        if not speech_frames or voice_start_time is None:
            return None, None

        audio = np.concatenate(speech_frames, axis=0).flatten()
        self.last_duration_sec = len(audio) / self.sample_rate
        voice_rt = voice_start_time - record_start_time
        return audio, voice_rt

    @staticmethod
    def save_wav(path, audio, sample_rate=16000):
        path.parent.mkdir(parents=True, exist_ok=True)
        with wave.open(str(path), "wb") as wf:
            wf.setnchannels(1)
            wf.setsampwidth(2)
            wf.setframerate(sample_rate)
            wf.writeframes(audio.tobytes())

# ============================================================
# MFCC Feature Extraction
# ============================================================

def extract_mfcc_features(audio_int16, sample_rate=16000):
    """
    Output: frames x 39
      - 13 MFCC: spectral/timbre envelope of the voice
      - 13 delta: temporal change of MFCC
      - 13 delta-delta: acceleration of the MFCC change
    """
    audio = audio_int16.astype(np.float32)
    if np.max(np.abs(audio)) > 0:
        audio = audio / np.max(np.abs(audio))

    mfcc_feat = mfcc(
        audio,
        samplerate=sample_rate,
        numcep=13,
        nfilt=26,
        nfft=512,
        winlen=0.025,
        winstep=0.010,
        preemph=0.97,
        appendEnergy=True,
    )
    delta_feat = delta(mfcc_feat, 2)
    delta_delta_feat = delta(delta_feat, 2)
    feat = np.concatenate([mfcc_feat, delta_feat, delta_delta_feat], axis=1)
    return cmvn(feat)



# ============================================================
# Additional Sound Feature Extraction
# ============================================================

GLOBAL_FEATURE_KEYS = [
    "duration", "rms_mean", "rms_std", "rms_max", "rms_range",
    "onset_sharpness", "zcr_mean", "zcr_std",
    "spectral_centroid_mean", "spectral_centroid_std",
    "spectral_bandwidth_mean", "spectral_rolloff_mean",
    "f0_mean", "f0_std", "voiced_ratio",
]

def _normalize_audio(audio_int16):
    x = audio_int16.astype(np.float32)
    m = np.max(np.abs(x)) if len(x) else 0
    if m > 0:
        x = x / m
    return x

def _frame_signal(x, frame_length, hop_length):
    if len(x) < frame_length:
        x = np.pad(x, (0, frame_length-len(x)))
    frames = []
    for st in range(0, len(x)-frame_length+1, hop_length):
        frames.append(x[st:st+frame_length])
    if not frames:
        frames = [np.pad(x, (0, max(0, frame_length-len(x))))[:frame_length]]
    return np.stack(frames, axis=0)

def _estimate_pitch_autocorr(x, sample_rate=SAMPLE_RATE, fmin=70, fmax=400):
    frame_length = int(0.04 * sample_rate)
    hop_length = int(0.01 * sample_rate)
    frames = _frame_signal(x, frame_length, hop_length)
    min_lag = int(sample_rate / fmax)
    max_lag = int(sample_rate / fmin)
    f0s = []
    for fr in frames:
        fr = fr - np.mean(fr)
        if np.sqrt(np.mean(fr*fr)+1e-8) < 0.02:
            continue
        corr = np.correlate(fr, fr, mode='full')[len(fr)-1:]
        if len(corr) <= max_lag:
            continue
        corr[:min_lag] = 0
        lag = int(np.argmax(corr[min_lag:max_lag]) + min_lag)
        peak = corr[lag] / (corr[0] + 1e-8)
        if peak > 0.25:
            f0s.append(sample_rate / lag)
    if not f0s:
        return 0.0, 0.0, 0.0
    f0s = np.asarray(f0s, dtype=np.float32)
    return float(np.mean(f0s)), float(np.std(f0s)), float(len(f0s)/max(1, len(frames)))

def extract_global_sound_features(audio_int16, sample_rate=SAMPLE_RATE):
    """Extract scalar sound features useful for short consonant/vowel recognition.

    Features include duration, RMS envelope, onset sharpness, zero crossing rate,
    spectral centroid/bandwidth/rolloff, and simple pitch statistics.
    """
    x = _normalize_audio(audio_int16)
    duration = len(x) / sample_rate if len(x) else 0.0
    frame_length = int(0.025 * sample_rate)
    hop_length = int(0.010 * sample_rate)
    frames = _frame_signal(x, frame_length, hop_length)
    rms = np.sqrt(np.mean(frames*frames, axis=1) + 1e-8)
    rms_mean = float(np.mean(rms)); rms_std=float(np.std(rms)); rms_max=float(np.max(rms))
    rms_range = float(np.max(rms)-np.min(rms))
    drms = np.diff(rms)
    onset_sharpness = float(np.max(drms)) if len(drms) else 0.0
    signs = np.sign(frames)
    zcr = np.mean(np.abs(np.diff(signs, axis=1)) > 0, axis=1)
    zcr_mean=float(np.mean(zcr)); zcr_std=float(np.std(zcr))
    fft_size=512
    win = get_window('hann', frame_length)
    freqs = np.fft.rfftfreq(fft_size, d=1/sample_rate)
    mags = np.abs(np.fft.rfft(frames*win, n=fft_size, axis=1)) + 1e-8
    power = mags*mags
    psum = np.sum(power, axis=1, keepdims=True) + 1e-8
    centroid = np.sum(power*freqs[None,:], axis=1)/psum.flatten()
    bandwidth = np.sqrt(np.sum(power*(freqs[None,:]-centroid[:,None])**2, axis=1)/psum.flatten())
    cumulative = np.cumsum(power, axis=1)
    roll=[]
    for i in range(power.shape[0]):
        idx = np.searchsorted(cumulative[i], 0.85*cumulative[i,-1])
        roll.append(freqs[min(idx, len(freqs)-1)])
    roll=np.asarray(roll)
    f0_mean, f0_std, voiced_ratio = _estimate_pitch_autocorr(x, sample_rate)
    d = {
        'duration': float(duration),
        'rms_mean': rms_mean,
        'rms_std': rms_std,
        'rms_max': rms_max,
        'rms_range': rms_range,
        'onset_sharpness': onset_sharpness,
        'zcr_mean': zcr_mean,
        'zcr_std': zcr_std,
        'spectral_centroid_mean': float(np.mean(centroid)),
        'spectral_centroid_std': float(np.std(centroid)),
        'spectral_bandwidth_mean': float(np.mean(bandwidth)),
        'spectral_rolloff_mean': float(np.mean(roll)),
        'f0_mean': f0_mean,
        'f0_std': f0_std,
        'voiced_ratio': voiced_ratio,
    }
    v = np.array([d[k] for k in GLOBAL_FEATURE_KEYS], dtype=np.float32)
    return d, v

def _global_stats_from_templates(templates, candidate_labels):
    vecs=[]
    for lab in candidate_labels:
        for t in templates.get(lab, []):
            if isinstance(t, dict) and 'global_vec' in t:
                vecs.append(t['global_vec'])
    if not vecs:
        return np.zeros(len(GLOBAL_FEATURE_KEYS), dtype=np.float32), np.ones(len(GLOBAL_FEATURE_KEYS), dtype=np.float32)
    arr=np.stack(vecs, axis=0)
    return np.mean(arr, axis=0), np.std(arr, axis=0)+1e-6

def cmvn(feat):
    return (feat - np.mean(feat, axis=0, keepdims=True)) / (np.std(feat, axis=0, keepdims=True) + 1e-8)

# ============================================================
# DTW Distance
# ============================================================

def dtw_distance(x, y):
    if x is None or y is None or len(x) == 0 or len(y) == 0:
        return np.inf
    dist = cdist(x, y, metric="euclidean")
    n, m = dist.shape
    dp = np.full((n + 1, m + 1), np.inf)
    dp[0, 0] = 0.0
    for i in range(1, n + 1):
        # simple DP; enough for short utterances
        for j in range(1, m + 1):
            dp[i, j] = dist[i - 1, j - 1] + min(dp[i - 1, j], dp[i, j - 1], dp[i - 1, j - 1])
    return float(dp[n, m] / (n + m))

# ============================================================
# Whisper STT helper
# ============================================================

class STTEngine:
    def __init__(self, device="cuda", compute_type="float16", fallback_to_cpu=True):
        self.model = None
        self.model_name = "tiny"
        self.device = device
        self.compute_type = compute_type
        self.fallback_to_cpu = fallback_to_cpu
        self.active_device = device
        self.active_compute_type = compute_type

    def load(self, model_name="tiny"):
        reload_needed = (
            self.model is None
            or self.model_name != model_name
            or self.active_device != self.device
            or self.active_compute_type != self.compute_type
        )
        if reload_needed:
            self.model_name = model_name
            try:
                self.model = WhisperModel(
                    model_name,
                    device=self.device,
                    compute_type=self.compute_type,
                )
                self.active_device = self.device
                self.active_compute_type = self.compute_type
            except Exception as e:
                if not self.fallback_to_cpu:
                    raise
                print(f"[STT] CUDA load failed: {e}")
                print("[STT] Falling back to CPU/int8.")
                self.model = WhisperModel(
                    model_name,
                    device="cpu",
                    compute_type="int8",
                )
                self.active_device = "cpu"
                self.active_compute_type = "int8"
        return self.model

    def transcribe(self, wav_path, model_name="tiny"):
        model = self.load(model_name)
        t0 = time.perf_counter()
        segments, info = model.transcribe(
            str(wav_path),
            language="ko",
            beam_size=5,
            vad_filter=False,
            condition_on_previous_text=False,
        )
        text = "".join(seg.text for seg in segments).strip()
        elapsed = time.perf_counter() - t0
        # Return elapsed only; device info is available as self.active_device.
        return text, normalize_text(text), elapsed

# ============================================================
# Voice Profile
# ============================================================

class VoiceProfile:
    def __init__(self, subject_id, gender_group="Unspecified"):
        self.subject_id = subject_id
        self.gender_group = gender_group or "Unspecified"
        self.subject_dir = BASE_DIR / subject_id
        self.wav_dir = self.subject_dir / "calibration_wav"
        self.rejected_dir = self.subject_dir / "rejected_wav"
        self.trial_dir = self.subject_dir / "trial_wav"
        self.profile_path = self.subject_dir / "profile_alias_hybrid_sets_tunable.pkl"
        self.meta_path = self.subject_dir / "profile_metadata.json"
        self.templates = defaultdict(list)
        self.alias_counts = defaultdict(Counter)
        self.subject_dir.mkdir(parents=True, exist_ok=True)
        self.wav_dir.mkdir(parents=True, exist_ok=True)
        self.rejected_dir.mkdir(parents=True, exist_ok=True)
        self.trial_dir.mkdir(parents=True, exist_ok=True)

    def add_template(self, label, audio, stt_text=""):
        mfcc_feat = extract_mfcc_features(audio, SAMPLE_RATE)
        global_dict, global_vec = extract_global_sound_features(audio, SAMPLE_RATE)
        self.templates[label].append({
            'mfcc': mfcc_feat,
            'global_dict': global_dict,
            'global_vec': global_vec,
        })
        norm = normalize_text(stt_text)
        if norm:
            self.alias_counts[label][norm] += 1
        # Manual aliases are weak priors.
        for alias in MANUAL_ALIASES.get(label, []):
            self.alias_counts[label][normalize_text(alias)] += 1

    def save_metadata(self):
        self.subject_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "subject_id": self.subject_id,
            "gender_group": self.gender_group,
            "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        with open(self.meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)

    def load_metadata(self):
        if self.meta_path.exists():
            try:
                with open(self.meta_path, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                self.gender_group = meta.get("gender_group", "Unspecified")
            except Exception:
                self.gender_group = "Unspecified"
        return self.gender_group

    def save(self):
        with open(self.profile_path, "wb") as f:
            pickle.dump({
                "templates": dict(self.templates),
                "alias_counts": {k: dict(v) for k, v in self.alias_counts.items()},
                "gender_group": self.gender_group,
            }, f)
        self.save_metadata()

    def load(self):
        if not self.profile_path.exists():
            return False
        with open(self.profile_path, "rb") as f:
            data = pickle.load(f)
        self.gender_group = data.get("gender_group", self.load_metadata())
        self.templates = defaultdict(list, data.get("templates", {}))
        self.alias_counts = defaultdict(Counter)
        for k, v in data.get("alias_counts", {}).items():
            self.alias_counts[k] = Counter(v)
        return True

    def merge_pooled_profile(self, other_profile, candidate_labels, source_name="pooled", distance_penalty=0.40, alias_boost=False):
        """Merge another subject profile as a weak pooled reference.

        Pooled templates receive a positive distance_penalty, so they can help when
        the personal profile is sparse but do not dominate the personal profile.
        """
        for label in candidate_labels:
            for item in other_profile.templates.get(label, []):
                if not isinstance(item, dict):
                    item = {"mfcc": item, "global_vec": None, "global_dict": {}}
                copied = dict(item)
                copied["source"] = source_name
                copied["distance_penalty"] = float(distance_penalty)
                self.templates[label].append(copied)
            if alias_boost:
                for alias, count in other_profile.alias_counts.get(label, Counter()).items():
                    self.alias_counts[label][alias] += max(1, int(count * 0.25))

    def acoustic_scores(self, audio, candidate_labels):
        input_feat = extract_mfcc_features(audio, SAMPLE_RATE)
        input_global_dict, input_global_vec = extract_global_sound_features(audio, SAMPLE_RATE)
        global_mean, global_std = _global_stats_from_templates(self.templates, candidate_labels)
        input_global_z = (input_global_vec - global_mean) / global_std
        scores = []
        for label in candidate_labels:
            if label not in self.templates:
                continue
            mfcc_distances = []
            global_distances = []
            for t in self.templates[label]:
                if isinstance(t, dict):
                    template_mfcc = t.get('mfcc')
                    template_global_vec = t.get('global_vec')
                else:
                    # Backward compatibility with old profiles
                    template_mfcc = t
                    template_global_vec = None
                penalty = float(t.get("distance_penalty", 0.0)) if isinstance(t, dict) else 0.0
                mfcc_distances.append(dtw_distance(input_feat, template_mfcc) + penalty)
                if template_global_vec is not None:
                    template_global_z = (template_global_vec - global_mean) / global_std
                    global_distances.append(float(np.linalg.norm(input_global_z-template_global_z)/np.sqrt(len(GLOBAL_FEATURE_KEYS))) + penalty)
            if mfcc_distances:
                mfcc_best = float(min(mfcc_distances))
                global_best = float(min(global_distances)) if global_distances else 0.0
                # Combined acoustic score: MFCC-DTW remains dominant; global features help with duration, energy, ZCR, pitch, spectral cues.
                combined = mfcc_best + 0.25 * global_best
                scores.append({
                    "label": label,
                    "best_distance": combined,
                    "mfcc_distance": mfcc_best,
                    "global_distance": global_best,
                    "avg_distance": float(np.mean(sorted(mfcc_distances)[:min(2, len(mfcc_distances))])),
                    "template_count": len(mfcc_distances),
                })
        return sorted(scores, key=lambda x: x["best_distance"])

    def alias_bonus(self, label, stt_norm):
        if not stt_norm:
            return 0.0
        aliases = self.alias_counts.get(label, Counter())
        if not aliases:
            return 0.0

        best = 0.0
        total = max(1, sum(aliases.values()))
        for alias, count in aliases.items():
            sim = text_similarity(stt_norm, alias)
            freq_weight = min(1.0, count / total + 0.25)
            best = max(best, sim * freq_weight)
        return best

    def predict_hybrid(self, audio, candidate_labels, wav_path, stt_engine=None, stt_model="tiny", force_stt=False,
                       uncertain_ratio=0.90, uncertain_margin=0.45, alias_weight=0.65):
        t0 = time.perf_counter()
        acoustic = self.acoustic_scores(audio, candidate_labels)
        acoustic_time = time.perf_counter() - t0
        if not acoustic:
            return {"error": "No acoustic templates for selected set."}

        best = acoustic[0]
        second = acoustic[1] if len(acoustic) > 1 else None
        margin = (second["best_distance"] - best["best_distance"]) if second else 999.0
        ratio = best["best_distance"] / (second["best_distance"] + 1e-8) if second else 0.0
        uncertain = ratio >= uncertain_ratio or margin <= uncertain_margin

        stt_raw = ""
        stt_norm = ""
        stt_time = 0.0
        used_stt = False
        if stt_engine is not None and (force_stt or uncertain):
            stt_raw, stt_norm, stt_time = stt_engine.transcribe(wav_path, stt_model)
            used_stt = True

        hybrid_scores = []
        for item in acoustic:
            bonus = self.alias_bonus(item["label"], stt_norm)
            # Lower score is better. Alias bonus lowers distance.
            adjusted = item["best_distance"] - alias_weight * bonus
            hybrid_scores.append({**item, "alias_bonus": bonus, "adjusted_distance": adjusted})
        hybrid_scores = sorted(hybrid_scores, key=lambda x: x["adjusted_distance"])

        final = hybrid_scores[0]
        final_second = hybrid_scores[1] if len(hybrid_scores) > 1 else None
        final_margin = final_second["adjusted_distance"] - final["adjusted_distance"] if final_second else 999.0

        warning = ""
        if final_second:
            for group in CONFUSION_GROUPS:
                if final["label"] in group and final_second["label"] in group:
                    warning = "Warning: top candidates are in a known vowel-confusion group."
                    break

        return {
            "predicted_label": final["label"],
            "second_label": final_second["label"] if final_second else "",
            "best_distance": best["best_distance"],
            "second_distance": second["best_distance"] if second else 0.0,
            "margin": margin,
            "ratio": ratio,
            "uncertain": uncertain,
            "used_stt": used_stt,
            "stt_raw": stt_raw,
            "stt_norm": stt_norm,
            "stt_time": stt_time,
            "acoustic_time": acoustic_time,
            "final_adjusted_distance": final["adjusted_distance"],
            "final_margin": final_margin,
            "scores": hybrid_scores,
            "warning": warning,
        }


# ============================================================
# Machine-Learning Classifier Utilities
# ============================================================

def _mfcc_stats_vector(mfcc_seq):
    """
    Convert frame-wise MFCC sequence into a fixed-length vector.
    This is much faster at recognition time than DTW over all templates.
    """
    if mfcc_seq is None or len(mfcc_seq) == 0:
        return np.zeros(39 * 6, dtype=np.float32)

    arr = np.asarray(mfcc_seq, dtype=np.float32)
    stats = [
        np.mean(arr, axis=0),
        np.std(arr, axis=0),
        np.min(arr, axis=0),
        np.max(arr, axis=0),
        np.percentile(arr, 10, axis=0),
        np.percentile(arr, 90, axis=0),
    ]
    return np.concatenate(stats).astype(np.float32)


def make_ml_feature_from_audio(audio_int16):
    """
    Feature vector for the learned model.

    Components:
    - MFCC/Delta/Delta-Delta summary statistics
    - Duration, RMS, ZCR, spectral, pitch/formant-like global sound features
    """
    mfcc_seq = extract_mfcc_features(audio_int16, SAMPLE_RATE)
    _, global_vec = extract_global_sound_features(audio_int16, SAMPLE_RATE)
    return np.concatenate([
        _mfcc_stats_vector(mfcc_seq),
        np.asarray(global_vec, dtype=np.float32),
    ]).astype(np.float32)


def make_ml_feature_from_template(template_item):
    """
    Build a model feature vector from stored calibration/feedback templates.
    Compatible with older profile formats.
    """
    if isinstance(template_item, dict):
        mfcc_seq = template_item.get("mfcc")
        global_vec = template_item.get("global_vec")
    else:
        mfcc_seq = template_item
        global_vec = None

    mfcc_vec = _mfcc_stats_vector(mfcc_seq)

    if global_vec is None:
        global_vec = np.zeros(len(GLOBAL_FEATURE_KEYS), dtype=np.float32)
    else:
        global_vec = np.asarray(global_vec, dtype=np.float32)

    return np.concatenate([mfcc_vec, global_vec]).astype(np.float32)


def ml_model_path_for_subject(subject_id):
    return BASE_DIR / subject_id / "learned_voice_model.joblib"


def collect_ml_training_data(
    subject_id,
    candidate_labels,
    gender_group="Unspecified",
    include_pooled=True,
    same_gender_only=True,
):
    """
    Collect feature vectors from the current subject and, optionally,
    other subjects as pooled training data.

    Current subject samples are weighted more strongly. Pooled samples are
    used as population-level reference data, not as a replacement for the
    subject-specific profile.
    """
    X = []
    y = []
    sample_weight = []

    subject_id_lower = subject_id.lower()
    target_group = gender_group or "Unspecified"

    for subject_dir in BASE_DIR.iterdir():
        if not subject_dir.is_dir():
            continue

        is_current = subject_dir.name.lower() == subject_id_lower

        if not is_current and not include_pooled:
            continue

        profile = VoiceProfile(subject_dir.name)
        if not profile.load():
            continue

        other_group = profile.gender_group or profile.load_metadata()
        if not is_current and same_gender_only and other_group != target_group:
            continue

        weight = 1.0 if is_current else 0.35

        for label in candidate_labels:
            for item in profile.templates.get(label, []):
                X.append(make_ml_feature_from_template(item))
                y.append(label)
                sample_weight.append(weight)

    if not X:
        return None, None, None

    return (
        np.stack(X, axis=0).astype(np.float32),
        np.asarray(y),
        np.asarray(sample_weight, dtype=np.float32),
    )


def train_learned_voice_model(
    subject_id,
    candidate_labels,
    gender_group="Unspecified",
    include_pooled=True,
    same_gender_only=True,
):
    """
    Train a small ensemble model from calibration/feedback data.

    We use RandomForest + ExtraTrees because they:
    - train quickly on small-to-medium calibration datasets
    - support nonlinear combinations of sound features
    - output class probabilities for Top-K candidate display
    - are extremely fast at prediction time
    """
    if not HAS_SKLEARN:
        raise RuntimeError("scikit-learn/joblib not installed. Run: pip install scikit-learn joblib")

    X, y, w = collect_ml_training_data(
        subject_id=subject_id,
        candidate_labels=candidate_labels,
        gender_group=gender_group,
        include_pooled=include_pooled,
        same_gender_only=same_gender_only,
    )

    if X is None or len(np.unique(y)) < 2:
        raise RuntimeError("Not enough training data. Need at least 2 labels with samples.")

    rf = RandomForestClassifier(
        n_estimators=260,
        max_features="sqrt",
        class_weight="balanced_subsample",
        random_state=42,
        n_jobs=-1,
    )
    et = ExtraTreesClassifier(
        n_estimators=260,
        max_features="sqrt",
        class_weight="balanced",
        random_state=43,
        n_jobs=-1,
    )

    rf.fit(X, y, sample_weight=w)
    et.fit(X, y, sample_weight=w)

    model_bundle = {
        "rf": rf,
        "et": et,
        "labels": sorted(list(set(y))),
        "candidate_labels": list(candidate_labels),
        "feature_dim": int(X.shape[1]),
        "trained_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "subject_id": subject_id,
        "gender_group": gender_group,
        "include_pooled": include_pooled,
        "same_gender_only": same_gender_only,
        "n_samples": int(X.shape[0]),
        "n_classes": int(len(np.unique(y))),
    }

    path = ml_model_path_for_subject(subject_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump(model_bundle, path)

    return model_bundle, path


def load_learned_voice_model(subject_id):
    if not HAS_SKLEARN:
        return None
    path = ml_model_path_for_subject(subject_id)
    if not path.exists():
        return None
    return joblib.load(path)


def predict_with_learned_voice_model(model_bundle, audio_int16, candidate_labels=None):
    """
    Return probability-ranked candidate labels.
    """
    t0 = time.perf_counter()
    x = make_ml_feature_from_audio(audio_int16).reshape(1, -1)

    rf = model_bundle["rf"]
    et = model_bundle["et"]

    rf_classes = list(rf.classes_)
    et_classes = list(et.classes_)

    prob_map = defaultdict(float)

    rf_prob = rf.predict_proba(x)[0]
    et_prob = et.predict_proba(x)[0]

    for label, p in zip(rf_classes, rf_prob):
        prob_map[label] += 0.5 * float(p)

    for label, p in zip(et_classes, et_prob):
        prob_map[label] += 0.5 * float(p)

    if candidate_labels is not None:
        allowed = set(candidate_labels)
        prob_map = {k: v for k, v in prob_map.items() if k in allowed}

    if not prob_map:
        return None

    total = sum(prob_map.values()) + 1e-12
    ranked = sorted(
        [{"label": k, "probability": v / total} for k, v in prob_map.items()],
        key=lambda d: d["probability"],
        reverse=True,
    )

    elapsed = time.perf_counter() - t0
    return ranked, elapsed



def rerank_ml_with_stt_alias(profile, ranked, stt_norm, alias_weight=0.35):
    """
    Adjust learned-model probabilities using subject-specific STT alias statistics.
    This replaces the old DTW fallback: no template DTW is performed.
    """
    if not stt_norm:
        return ranked

    adjusted = []
    for item in ranked:
        label = item["label"]
        p = float(item["probability"])
        alias = float(profile.alias_bonus(label, stt_norm))
        # Probability-space bonus. Keep it conservative so acoustic ML remains primary.
        p2 = p + alias_weight * alias
        adjusted.append({
            "label": label,
            "probability": p2,
            "raw_ml_probability": p,
            "stt_alias_score": alias,
        })

    total = sum(x["probability"] for x in adjusted) + 1e-12
    for x in adjusted:
        x["probability"] = x["probability"] / total

    adjusted = sorted(adjusted, key=lambda d: d["probability"], reverse=True)
    return adjusted


def learned_model_result_dict(
    ranked,
    ml_time,
    voice_rt,
    wav_path,
    noise_rms,
    energy_threshold,
    recorded_duration,
    gender_group,
    use_pooled,
    same_gender_only,
    pooled_penalty,
):
    best = ranked[0]
    second = ranked[1] if len(ranked) > 1 else {"label": "", "probability": 0.0}

    top_prob = float(best["probability"])
    second_prob = float(second["probability"])

    scores = []
    for item in ranked:
        p = float(item["probability"])
        scores.append({
            "label": item["label"],
            "best_distance": 1.0 - p,
            "mfcc_distance": 0.0,
            "global_distance": 0.0,
            "avg_distance": 1.0 - p,
            "template_count": 0,
            "alias_bonus": 0.0,
            "adjusted_distance": 1.0 - p,
            "ml_probability": p,
        })

    return {
        "predicted_label": best["label"],
        "second_label": second["label"],
        "best_distance": 1.0 - top_prob,
        "second_distance": 1.0 - second_prob,
        "margin": top_prob - second_prob,
        "ratio": (1.0 - top_prob) / ((1.0 - second_prob) + 1e-8),
        "uncertain": False,
        "used_stt": False,
        "stt_raw": "",
        "stt_norm": "",
        "stt_time": 0.0,
        "used_dtw": False,
        "acoustic_time": ml_time,
        "final_adjusted_distance": 1.0 - top_prob,
        "final_margin": top_prob - second_prob,
        "scores": scores,
        "warning": "",
        "used_ml_model": True,
        "ml_confidence": top_prob,
        "ml_time": ml_time,
        "voice_rt": voice_rt,
        "wav_path": str(wav_path),
        "noise_rms": noise_rms,
        "energy_threshold": energy_threshold,
        "recorded_duration": recorded_duration,
        "gender_group": gender_group,
        "use_pooled": use_pooled,
        "same_gender_only": same_gender_only,
        "pooled_penalty": pooled_penalty,
        "pooled_subjects": [],
    }



# ============================================================
# STT-based Syllable/Word Candidate Matching
# ============================================================

DEFAULT_SYLLABLE_XLSX = "syllable_top200.xlsx"

HANGUL_SYLLABLE_START = 0xAC00
HANGUL_SYLLABLE_END = 0xD7A3
CHOSUNG_LIST = list("ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ")
JUNGSUNG_LIST = list("ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ")
JONGSUNG_LIST = [""] + list("ㄱㄲㄳㄴㄵㄶㄷㄹㄺㄻㄼㄽㄾㄿㅀㅁㅂㅄㅅㅆㅇㅈㅊㅋㅌㅍㅎ")


def normalize_text_for_match(text):
    if text is None:
        return ""
    text = unicodedata.normalize("NFC", str(text))
    text = re.sub(r"[\s\.\,\?\!\:\;\-_\(\)\[\]\{\}\"\'`~]", "", text)
    return text.strip()


def decompose_hangul_syllable(ch):
    if not ch:
        return ("", "", "")
    code = ord(ch)
    if code < HANGUL_SYLLABLE_START or code > HANGUL_SYLLABLE_END:
        return (ch, "", "")
    sindex = code - HANGUL_SYLLABLE_START
    cho = sindex // 588
    jung = (sindex % 588) // 28
    jong = sindex % 28
    return CHOSUNG_LIST[cho], JUNGSUNG_LIST[jung], JONGSUNG_LIST[jong]


def hangul_phonetic_similarity(a, b):
    a = normalize_text_for_match(a)
    b = normalize_text_for_match(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0

    seq_sim = difflib.SequenceMatcher(None, a, b).ratio()
    max_len = max(len(a), len(b))
    char_scores = []

    for i in range(max_len):
        ca = a[i] if i < len(a) else ""
        cb = b[i] if i < len(b) else ""

        if ca == cb and ca:
            char_scores.append(1.0)
            continue

        aa = decompose_hangul_syllable(ca)
        bb = decompose_hangul_syllable(cb)

        score = 0.0
        if aa[0] and aa[0] == bb[0]:
            score += 0.35
        if aa[1] and aa[1] == bb[1]:
            score += 0.45
        if aa[2] == bb[2]:
            score += 0.20
        char_scores.append(score)

    jamo_sim = float(sum(char_scores) / max(len(char_scores), 1))

    contain_bonus = 0.0
    if a in b or b in a:
        contain_bonus = 0.15

    return min(1.0, 0.45 * seq_sim + 0.55 * jamo_sim + contain_bonus)


def load_syllable_candidates_from_xlsx(xlsx_path=DEFAULT_SYLLABLE_XLSX):
    path = Path(xlsx_path)
    if not path.exists():
        path = Path.cwd() / xlsx_path

    if not path.exists():
        return []

    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required. Run: python -m pip install openpyxl")

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    candidates = []
    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)
    if first is None:
        return []

    header = [str(x).strip().lower() if x is not None else "" for x in first]
    has_header = any(h in header for h in ["syl", "syll", "syllable", "count"])

    if has_header:
        if "syl" in header:
            syl_col = header.index("syl")
        elif "syll" in header:
            syl_col = header.index("syll")
        elif "syllable" in header:
            syl_col = header.index("syllable")
        else:
            syl_col = 0
        count_col = header.index("count") if "count" in header else None
    else:
        syl_col = 0
        count_col = 1 if len(first) > 1 else None
        rows = iter([first] + list(rows))

    for row in rows:
        if row is None or len(row) <= syl_col:
            continue
        syl = normalize_text_for_match(row[syl_col])
        if not syl:
            continue
        count = 0
        if count_col is not None and len(row) > count_col and row[count_col] is not None:
            try:
                count = int(row[count_col])
            except Exception:
                count = 0
        candidates.append({"label": syl, "count": count})

    best = {}
    for item in candidates:
        label = item["label"]
        if label not in best or item["count"] > best[label]["count"]:
            best[label] = item

    candidates = list(best.values())
    candidates.sort(key=lambda x: x.get("count", 0), reverse=True)
    return candidates


def rank_candidates_by_stt_text(stt_text, candidates, top_k=10):
    stt_norm = normalize_text_for_match(stt_text)
    ranked = []

    for item in candidates:
        label = item["label"] if isinstance(item, dict) else str(item)
        count = item.get("count", 0) if isinstance(item, dict) else 0

        sim = hangul_phonetic_similarity(stt_norm, label)

        freq_bonus = 0.0
        if count > 0:
            freq_bonus = min(0.05, np.log10(count + 1) / 200.0)

        score = sim + freq_bonus

        ranked.append({
            "label": label,
            "score": float(score),
            "similarity": float(sim),
            "count": int(count),
        })

    ranked.sort(key=lambda x: x["score"], reverse=True)
    return ranked[:top_k]



# ============================================================
# Hangul Tactile Learning v20 - Voice Response Integration
# ============================================================
# This app is intentionally self-contained and uses the voice core above.
# It replaces large visual-choice grids with voice recognition followed by
# Top-5 candidate buttons + manual correction.

import csv
from dataclasses import dataclass
from datetime import date

try:
    import serial
    import serial.tools.list_ports
except Exception:
    serial = None

from PySide6.QtCore import QTimer
from PySide6.QtWidgets import QMainWindow, QScrollArea

OUTPUT_DIR = Path("hangul_learning_results")
OUTPUT_DIR.mkdir(exist_ok=True)
BAUDRATE = 115200
STIM_DELAY_MS = 500
DEFAULT_ISI_MS = 150
SYLLABLE_ISI_MS = 250
SYLLABLE_FILE = Path("syllable_top200.xlsx")
MAX_TRIALS_PER_ITEM = 30
REPS_REQUIRED_IN_RECENT_WINDOW = 28
RECENT_WINDOW = 28
ALT_ACC_THRESHOLD = 0.90

@dataclass
class StimItem:
    label: str
    command_body: str
    session_name: str
    description: str = ""


def single_motor(motor, duration_ms):
    return f"{motor}/{duration_ms}"


def sequence(*parts):
    return ".".join([str(p) for p in parts if str(p).strip()])


def repeat_pattern(pattern, isi_ms=DEFAULT_ISI_MS):
    return sequence(pattern, f"0/{isi_ms}", pattern)

# ------------------------------------------------------------------
# Jamo stimulus definitions
# ------------------------------------------------------------------
SHORT_CONSONANTS = [
    StimItem("ㄱ", single_motor(3, 150), "basic_consonants", "ㄱ = motor 3, 150 ms"),
    StimItem("ㄴ", single_motor(2, 150), "basic_consonants", "ㄴ = motor 2, 150 ms"),
    StimItem("ㄷ", single_motor(1, 150), "basic_consonants", "ㄷ = motor 1, 150 ms"),
    StimItem("ㅂ", single_motor(6, 150), "basic_consonants", "ㅂ = motor 6, 150 ms"),
    StimItem("ㅅ", single_motor(4, 150), "basic_consonants", "ㅅ = motor 4, 150 ms"),
    StimItem("ㅇ", single_motor(9, 150), "basic_consonants", "ㅇ = motor 9, 150 ms"),
    StimItem("ㅈ", single_motor(7, 150), "basic_consonants", "ㅈ = motor 7, 150 ms"),
]

LONG_CONSONANTS = [
    StimItem("ㅋ", single_motor(3, 350), "long_consonants", "ㅋ = ㄱ position, 350 ms"),
    StimItem("ㄹ", single_motor(2, 350), "long_consonants", "ㄹ = ㄴ position, 350 ms"),
    StimItem("ㅌ", single_motor(1, 350), "long_consonants", "ㅌ = ㄷ position, 350 ms"),
    StimItem("ㅍ", single_motor(6, 350), "long_consonants", "ㅍ = ㅂ position, 350 ms"),
    StimItem("ㅎ", single_motor(4, 350), "long_consonants", "ㅎ = ㅅ position, 350 ms"),
    StimItem("ㅁ", single_motor(9, 350), "long_consonants", "ㅁ = ㅇ position, 350 ms"),
    StimItem("ㅊ", single_motor(7, 350), "long_consonants", "ㅊ = ㅈ position, 350 ms"),
]

BASE_CONSONANT_COMMAND = {item.label: item.command_body for item in SHORT_CONSONANTS + LONG_CONSONANTS}
DOUBLE_CONSONANT_BASE = {"ㄲ": "ㄱ", "ㄸ": "ㄷ", "ㅃ": "ㅂ", "ㅆ": "ㅅ", "ㅉ": "ㅈ"}
DOUBLE_CONSONANTS = [
    StimItem(label, repeat_pattern(BASE_CONSONANT_COMMAND[base], DEFAULT_ISI_MS), "double_consonants", f"{label} = {base} + {DEFAULT_ISI_MS} ms + {base}")
    for label, base in DOUBLE_CONSONANT_BASE.items()
]
ALL_CONSONANTS = SHORT_CONSONANTS + LONG_CONSONANTS + DOUBLE_CONSONANTS

V_I = single_motor(5, 150)
V_EU = single_motor(5, 350)
V_O = sequence("5/150", "5/d,2/i/100", "2/150")
V_U = sequence("5/150", "5/d,8/i/100", "8/150")
V_A = sequence("5/150", "5/d,4/i/100", "4/150")
V_EO = sequence("5/150", "5/d,6/i/100", "6/150")

BASIC_VOWELS = [
    StimItem("ㅣ", V_I, "basic_vowels", "ㅣ = motor 5, 150 ms"),
    StimItem("ㅡ", V_EU, "basic_vowels", "ㅡ = motor 5, 350 ms"),
    StimItem("ㅏ", V_A, "basic_vowels", "ㅏ = 5 to 4 line"),
    StimItem("ㅓ", V_EO, "basic_vowels", "ㅓ = 5 to 6 line"),
    StimItem("ㅗ", V_O, "basic_vowels", "ㅗ = 5 to 2 line"),
    StimItem("ㅜ", V_U, "basic_vowels", "ㅜ = 5 to 8 line"),
]
BASE_VOWEL_COMMAND = {item.label: item.command_body for item in BASIC_VOWELS}

DOUBLE_VOWELS = [
    StimItem("ㅢ", sequence(V_EU, f"0/{DEFAULT_ISI_MS}", V_I), "double_vowels", "ㅢ = ㅡ + ㅣ"),
    StimItem("ㅑ", repeat_pattern(V_A, DEFAULT_ISI_MS), "double_vowels", "ㅑ = ㅏ repeated"),
    StimItem("ㅕ", repeat_pattern(V_EO, DEFAULT_ISI_MS), "double_vowels", "ㅕ = ㅓ repeated"),
    StimItem("ㅛ", repeat_pattern(V_O, DEFAULT_ISI_MS), "double_vowels", "ㅛ = ㅗ repeated"),
    StimItem("ㅠ", repeat_pattern(V_U, DEFAULT_ISI_MS), "double_vowels", "ㅠ = ㅜ repeated"),
]
DOUBLE_VOWEL_COMMAND = {item.label: item.command_body for item in DOUBLE_VOWELS}

VOWEL_COMMAND_LOOKUP = {**BASE_VOWEL_COMMAND, **DOUBLE_VOWEL_COMMAND}

def combine_vowels(label, *components):
    parts = []
    for i, comp in enumerate(components):
        if comp not in VOWEL_COMMAND_LOOKUP:
            return None
        if i > 0:
            parts.append(f"0/{DEFAULT_ISI_MS}")
        parts.append(VOWEL_COMMAND_LOOKUP[comp])
    return StimItem(label, sequence(*parts), "complex_vowels", f"{label} = {' + '.join(components)}")

COMPLEX_VOWELS = [
    combine_vowels("ㅐ", "ㅏ", "ㅣ"),
    combine_vowels("ㅔ", "ㅓ", "ㅣ"),
    combine_vowels("ㅒ", "ㅑ", "ㅣ"),
    combine_vowels("ㅖ", "ㅕ", "ㅣ"),
    combine_vowels("ㅘ", "ㅗ", "ㅏ"),

    # ㅙ and ㅞ are generated directly from base/double components.
    # Do not use ("ㅗ", "ㅐ") or ("ㅜ", "ㅔ") here because ㅐ/ㅔ are
    # created in this same list and are not yet in VOWEL_COMMAND_LOOKUP.
    combine_vowels("ㅙ", "ㅗ", "ㅏ", "ㅣ"),
    combine_vowels("ㅚ", "ㅗ", "ㅣ"),
    combine_vowels("ㅝ", "ㅜ", "ㅓ"),
    combine_vowels("ㅞ", "ㅜ", "ㅓ", "ㅣ"),
    combine_vowels("ㅟ", "ㅜ", "ㅣ"),
]
COMPLEX_VOWELS = [x for x in COMPLEX_VOWELS if x is not None]
ALL_VOWELS = BASIC_VOWELS + DOUBLE_VOWELS + COMPLEX_VOWELS
assert any(item.label == "ㅙ" for item in COMPLEX_VOWELS), "ㅙ missing from complex vowels"
assert any(item.label == "ㅞ" for item in COMPLEX_VOWELS), "ㅞ missing from complex vowels"

CONSONANT_COMMANDS = {item.label: item.command_body for item in ALL_CONSONANTS}
VOWEL_COMMANDS = {item.label: item.command_body for item in ALL_VOWELS}

# Hangul decomposition for syllable stimulus generation
CHOSUNG = ["ㄱ", "ㄲ", "ㄴ", "ㄷ", "ㄸ", "ㄹ", "ㅁ", "ㅂ", "ㅃ", "ㅅ", "ㅆ", "ㅇ", "ㅈ", "ㅉ", "ㅊ", "ㅋ", "ㅌ", "ㅍ", "ㅎ"]
JUNGSUNG = ["ㅏ", "ㅐ", "ㅑ", "ㅒ", "ㅓ", "ㅔ", "ㅕ", "ㅖ", "ㅗ", "ㅘ", "ㅙ", "ㅚ", "ㅛ", "ㅜ", "ㅝ", "ㅞ", "ㅟ", "ㅠ", "ㅡ", "ㅢ", "ㅣ"]
JONGSUNG = ["", "ㄱ", "ㄲ", "ㄳ", "ㄴ", "ㄵ", "ㄶ", "ㄷ", "ㄹ", "ㄺ", "ㄻ", "ㄼ", "ㄽ", "ㄾ", "ㄿ", "ㅀ", "ㅁ", "ㅂ", "ㅄ", "ㅅ", "ㅆ", "ㅇ", "ㅈ", "ㅊ", "ㅋ", "ㅌ", "ㅍ", "ㅎ"]

def decompose_learning_syllable(ch):
    code = ord(str(ch)[0])
    base = 0xAC00
    if code < base or code > 0xD7A3:
        return None
    sidx = code - base
    return CHOSUNG[sidx // 588], JUNGSUNG[(sidx % 588) // 28], JONGSUNG[sidx % 28]

def make_syllable_command_for_learning(syllable):
    dec = decompose_learning_syllable(str(syllable)[0])
    if dec is None:
        return None
    cho, jung, jong = dec
    parts = []
    if cho not in CONSONANT_COMMANDS or jung not in VOWEL_COMMANDS:
        return None
    parts.append(CONSONANT_COMMANDS[cho])
    parts.append(VOWEL_COMMANDS[jung])
    if jong:
        if jong in CONSONANT_COMMANDS:
            parts.append(CONSONANT_COMMANDS[jong])
        else:
            return None
    seq = []
    for i, part in enumerate(parts):
        if i > 0:
            seq.append(f"0/{SYLLABLE_ISI_MS}")
        seq.append(part)
    return sequence(*seq)

def load_syllable_items(limit, session_name):
    if not SYLLABLE_FILE.exists():
        print(f"[WARN] {SYLLABLE_FILE} not found; {session_name} will be empty.", flush=True)
        return []
    try:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(str(SYLLABLE_FILE), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"[WARN] Failed to read {SYLLABLE_FILE}: {e}", flush=True)
        return []
    if not rows:
        return []
    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
    start_idx = 1 if any(h in header for h in ["syl", "syll", "syllable", "count"]) else 0
    syl_col = header.index("syl") if "syl" in header else 0
    labels = []
    for row in rows[start_idx:]:
        if row and len(row) > syl_col and row[syl_col] is not None:
            s = str(row[syl_col]).strip()
            if s:
                labels.append(s[0])
        if len(labels) >= limit:
            break
    items = []
    for label in labels:
        cmd = make_syllable_command_for_learning(label)
        if cmd:
            items.append(StimItem(label, cmd, session_name, f"{label} syllable"))
    return items

SYLLABLE_TOP50 = load_syllable_items(50, "syllable_top50")
SYLLABLE_TOP100 = load_syllable_items(100, "syllable_top100")
SYLLABLE_TOP200 = load_syllable_items(200, "syllable_top200")

SESSIONS = [
    {"name": "basic_consonants", "title": "Session 1: Basic consonants", "items": SHORT_CONSONANTS, "type": "jamo"},
    {"name": "long_consonants", "title": "Session 2: Long consonants", "items": LONG_CONSONANTS, "type": "jamo"},
    {"name": "double_consonants", "title": "Session 3: Double consonants", "items": DOUBLE_CONSONANTS, "type": "jamo"},
    {"name": "all_consonants", "title": "Session 4: All consonants", "items": ALL_CONSONANTS, "type": "jamo"},
    {"name": "basic_vowels", "title": "Session 5: Basic vowels", "items": BASIC_VOWELS, "type": "jamo"},
    {"name": "double_vowels", "title": "Session 6: Double vowels", "items": DOUBLE_VOWELS, "type": "jamo"},
    {"name": "complex_vowels", "title": "Session 7: Complex vowels", "items": COMPLEX_VOWELS, "type": "jamo"},
    {"name": "all_vowels", "title": "Session 8: All vowels", "items": ALL_VOWELS, "type": "jamo"},
    {"name": "syllable_top50", "title": "Session 9: Top 50 syllables", "items": SYLLABLE_TOP50, "type": "syllable"},
    {"name": "syllable_top100", "title": "Session 10: Top 100 syllables", "items": SYLLABLE_TOP100, "type": "syllable"},
    {"name": "syllable_top200", "title": "Session 11: Top 200 syllables", "items": SYLLABLE_TOP200, "type": "syllable"},
]

# ------------------------------------------------------------------
# Voice label naming and model training for Hangul learning
# ------------------------------------------------------------------
JAMO_SPOKEN_NAME = {
    "ㄱ": "기역", "ㄲ": "쌍기역", "ㄴ": "니은", "ㄷ": "디귿", "ㄸ": "쌍디귿",
    "ㄹ": "리을", "ㅁ": "미음", "ㅂ": "비읍", "ㅃ": "쌍비읍", "ㅅ": "시옷", "ㅆ": "쌍시옷",
    "ㅇ": "이응", "ㅈ": "지읒", "ㅉ": "쌍지읒", "ㅊ": "치읓", "ㅋ": "키읔", "ㅌ": "티읕", "ㅍ": "피읖", "ㅎ": "히읗",
    "ㅏ": "아", "ㅐ": "애", "ㅑ": "야", "ㅒ": "얘", "ㅓ": "어", "ㅔ": "에", "ㅕ": "여", "ㅖ": "예",
    "ㅗ": "오", "ㅘ": "와", "ㅙ": "왜", "ㅚ": "외", "ㅛ": "요", "ㅜ": "우", "ㅝ": "워", "ㅞ": "웨", "ㅟ": "위", "ㅠ": "유", "ㅡ": "으", "ㅢ": "의", "ㅣ": "이",
}
ALL_JAMO_LABELS = [item.label for item in ALL_CONSONANTS + ALL_VOWELS]

def hangul_model_path(subject):
    return BASE_DIR / str(subject) / "hangul_learning_jamo_model.joblib"

def _template_keys_for_jamo(label):
    keys = [label]
    if label in JAMO_SPOKEN_NAME:
        keys.append(JAMO_SPOKEN_NAME[label])
    return keys

def train_hangul_jamo_model(subject_id, labels=None, include_pooled=True):
    if not HAS_SKLEARN:
        raise RuntimeError("scikit-learn/joblib not installed. Run: python -m pip install scikit-learn joblib")
    labels = labels or ALL_JAMO_LABELS
    X, y, weights = [], [], []
    current = str(subject_id).lower()
    for sdir in BASE_DIR.iterdir():
        if not sdir.is_dir():
            continue
        is_current = sdir.name.lower() == current
        if not is_current and not include_pooled:
            continue
        vp = VoiceProfile(sdir.name)
        if not vp.load():
            continue
        w = 1.0 if is_current else 0.25
        for label in labels:
            for key in _template_keys_for_jamo(label):
                for item in vp.templates.get(key, []):
                    X.append(make_ml_feature_from_template(item))
                    y.append(label)
                    weights.append(w)
    if not X or len(set(y)) < 2:
        raise RuntimeError("Not enough voice calibration data to train jamo model.")
    X = np.stack(X).astype(np.float32)
    y = np.asarray(y)
    weights = np.asarray(weights, dtype=np.float32)
    rf = RandomForestClassifier(n_estimators=240, max_features="sqrt", class_weight="balanced_subsample", random_state=101, n_jobs=-1)
    et = ExtraTreesClassifier(n_estimators=240, max_features="sqrt", class_weight="balanced", random_state=102, n_jobs=-1)
    rf.fit(X, y, sample_weight=weights)
    et.fit(X, y, sample_weight=weights)
    bundle = {
        "rf": rf,
        "et": et,
        "labels": sorted(set(y)),
        "subject_id": subject_id,
        "trained_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "n_samples": len(X),
        "pooled_only": not any(str(sdir.name).lower() == current for sdir in BASE_DIR.iterdir() if sdir.is_dir()),
        "include_pooled": include_pooled,
    }
    path = hangul_model_path(subject_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump(bundle, path)
    return bundle, path

def load_hangul_jamo_model(subject_id):
    """Load jamo ML model. If the new jamo-specific model is missing,
    fall back to the legacy learned_voice_model.joblib saved by the voice classifier."""
    subject_id = str(subject_id).strip()
    possible_paths = [
        Path("voice_profiles") / subject_id / "jamo_voice_model.joblib",
        Path("voice_profiles") / subject_id / "hangul_jamo_model.joblib",
        Path("voice_profiles") / subject_id / "learned_voice_model.joblib",
    ]

    if not HAS_SKLEARN or joblib is None:
        return None

    for path in possible_paths:
        if path.exists():
            try:
                return joblib.load(path)
            except Exception:
                continue
    return None


def predict_hangul_jamo(model_bundle, audio, labels, top_k=5):
    """Predict jamo label using either new jamo model or legacy spoken-name voice model."""
    x = make_ml_feature_from_audio(audio).reshape(1, -1)
    prob = defaultdict(float)

    # Support both bundle formats from this app and legacy voice classifier.
    models = []
    if isinstance(model_bundle, dict):
        if "rf" in model_bundle:
            models.append(model_bundle["rf"])
        if "et" in model_bundle:
            models.append(model_bundle["et"])
    else:
        models.append(model_bundle)

    for model in models:
        try:
            p = model.predict_proba(x)[0]
            classes = list(model.classes_)
        except Exception:
            continue

        weight = 1.0 / max(len(models), 1)
        for c, v in zip(classes, p):
            app_label = model_class_to_app_label(c)
            prob[app_label] += weight * float(v)

    allowed = set([normalize_response_label(x) for x in labels]) if labels else None

    if allowed:
        ranked = [
            {"label": k, "score": v, "probability": v}
            for k, v in prob.items()
            if k in allowed
        ]
    else:
        ranked = [{"label": k, "score": v, "probability": v} for k, v in prob.items()]

    # If allowed filtering removed everything, do NOT return empty.
    # Fall back to closest available model classes, then current labels.
    if not ranked:
        ranked = [{"label": k, "score": v, "probability": v} for k, v in prob.items()]

    ranked.sort(key=lambda d: d["score"], reverse=True)

    if allowed:
        # Prioritize allowed labels, but if no model probability exists for some allowed labels,
        # append them with zero score so UI still shows valid choices.
        seen = {r["label"] for r in ranked}
        for lab in labels:
            lab = normalize_response_label(lab)
            if lab not in seen:
                ranked.append({"label": lab, "score": 0.0, "probability": 0.0})
                seen.add(lab)

    return ranked[:top_k]


class TrialVoiceWorker(QThread):
    status = Signal(str)
    result_ready = Signal(dict)

    def __init__(self, subject_id, session_type, candidate_labels, recorder_settings, syllable_xlsx="syllable_top200.xlsx", stt_model="small", wav_dir=None):
        super().__init__()
        self.subject_id = subject_id
        self.session_type = session_type
        self.candidate_labels = list(candidate_labels)
        self.recorder_settings = recorder_settings
        self.syllable_xlsx = syllable_xlsx
        self.stt_model = stt_model
        self.wav_dir = Path(wav_dir) if wav_dir else BASE_DIR / str(subject_id) / "learning_trial_wav"
        self.wav_dir.mkdir(parents=True, exist_ok=True)


    def is_syllable_trial(self):
        """Return True for syllable/word sessions.

        Syllable sessions must use STT + candidate-list matching, not jamo ML.
        """
        if str(self.session_type).lower() == "syllable":
            return True

        # Fallback: if candidates contain Hangul syllables such as 가/나/이,
        # treat as syllable session. Jamo labels are usually one of ㄱ/ㅏ/etc.
        jamo_chars = set("ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ")
        for lab in self.candidate_labels:
            s = str(lab).strip()
            if not s:
                continue
            if any(("가" <= ch <= "힣") for ch in s):
                return True
            if s not in jamo_chars and len(s) >= 1:
                # Non-jamo candidate such as a word/syllable.
                return True
        return False


    def run(self):
        try:
            rec = AudioRecorder(**self.recorder_settings)
            self.status.emit("Voice detecting...")
            audio, onset_rt = rec.record_until_silence()
            if audio is None:
                self.result_ready.emit({"error": "no_voice", "message": "No voice detected."})
                return
            wav_path = self.wav_dir / f"trial_voice_{int(time.time()*1000)}.wav"
            rec.save_wav(wav_path, audio, SAMPLE_RATE)
            t0 = time.perf_counter()
            if self.is_syllable_trial():
                stt = STTEngine()
                self.status.emit(f"Loading STT model: {self.stt_model}")
                stt.load(self.stt_model)
                raw, norm, stt_time = stt.transcribe(str(wav_path), self.stt_model)
                self.status.emit(f"STT result: raw={raw} / norm={norm}")
                candidates = load_syllable_candidates_from_xlsx(self.syllable_xlsx)

                # Syllable sessions must be restricted to the current session candidate list.
                # Do not allow jamo candidates or outside-list candidates here.
                if self.candidate_labels:
                    allowed = set(str(x).strip() for x in self.candidate_labels)
                    candidates = [c for c in candidates if str(c.get("label", "")).strip() in allowed]
                    if not candidates:
                        candidates = [{"label": lab, "count": 0} for lab in self.candidate_labels]

                ranked = rank_candidates_by_stt_text(norm, candidates, top_k=5)

                # If STT produced no usable text, do not pretend the first list items are predictions.
                # Still show choices so the participant can answer, but mark confidence as 0.
                if (not norm) and self.candidate_labels:
                    ranked = [{"label": lab, "score": 0.0, "similarity": 0.0, "count": 0, "fallback_no_stt": True}
                              for lab in self.candidate_labels[:5]]

                elapsed = time.perf_counter() - t0
                self.result_ready.emit({
                    "engine": "stt_syllable", "stt_model": self.stt_model, "top": ranked, "stt_raw": raw, "stt_norm": norm,
                    "stt_time": stt_time, "process_time": elapsed, "voice_onset_rt_sec": onset_rt,
                    "wav_path": str(wav_path), "noise_rms": rec.last_noise_rms, "energy_threshold": rec.last_energy_threshold,
                    "recorded_duration": rec.last_duration_sec,
                })
            else:
                model = load_hangul_jamo_model(self.subject_id)
                if model is None:
                    self.result_ready.emit({"error": "no_model", "message": "No jamo ML model. Run calibration first."})
                    return
                ranked = predict_hangul_jamo(model, audio, self.candidate_labels, top_k=5)
                elapsed = time.perf_counter() - t0
                self.result_ready.emit({
                    "engine": "ml_jamo",
                    "top": ranked,
                    "top_candidates": ranked,
                    "scores": ranked,
                    "predicted_label": ranked[0]["label"] if ranked else "",
                    "second_label": ranked[1]["label"] if len(ranked) > 1 else "",
                    "stt_raw": "",
                    "stt_norm": "",
                    "stt_time": 0.0,
                    "process_time": elapsed,
                    "voice_onset_rt_sec": onset_rt,
                    "wav_path": str(wav_path),
                    "noise_rms": rec.last_noise_rms,
                    "energy_threshold": rec.last_energy_threshold,
                    "recorded_duration": rec.last_duration_sec,
                })
        except Exception as e:
            self.result_ready.emit({"error": "exception", "message": str(e)})


class VoiceCalibrationWorker(QThread):
    status = Signal(str)
    prompt = Signal(str)
    progress = Signal(int, int)
    finished_calibration = Signal(bool, str)

    def __init__(self, subject, recorder_settings, reps=3, include_pooled=True):
        super().__init__()
        self.subject = subject
        self.recorder_settings = recorder_settings
        self.reps = reps
        self.include_pooled = include_pooled
        self._stop_requested = False

    def stop(self):
        self._stop_requested = True

    def run(self):
        try:
            profile = VoiceProfile(self.subject)
            labels = ALL_JAMO_LABELS
            total = len(labels) * self.reps
            done = 0
            rec = AudioRecorder(**self.recorder_settings)
            self.status.emit(f"Calibration started: {len(labels)} labels × {self.reps} reps")
            for label in labels:
                if self._stop_requested:
                    self.finished_calibration.emit(False, "Calibration cancelled by user.")
                    return
                spoken = JAMO_SPOKEN_NAME.get(label, label)
                for rep in range(1, self.reps + 1):
                    if self._stop_requested:
                        self.finished_calibration.emit(False, "Calibration cancelled by user.")
                        return
                    self.prompt.emit(f"{label}  ({spoken})  —  {rep}/{self.reps}회 읽어주세요")
                    self.status.emit(f"Listening for calibration: {label} ({spoken}) rep {rep}/{self.reps}")
                    audio, rt = rec.record_until_silence()
                    if audio is None:
                        self.status.emit(f"No voice detected for {label} rep {rep}. Retrying once...")
                        self.prompt.emit(f"음성이 감지되지 않았습니다. 다시 읽어주세요: {label} ({spoken})")
                        audio, rt = rec.record_until_silence()
                    if audio is None:
                        self.status.emit(f"Skipped calibration: {label} rep {rep}")
                        done += 1
                        self.progress.emit(done, total)
                        continue
                    wav_path = profile.wav_dir / f"calib_{label}_{rep}_{int(time.time()*1000)}.wav"
                    rec.save_wav(wav_path, audio, SAMPLE_RATE)
                    profile.add_template(label, audio, stt_text=spoken)
                    profile.save()
                    done += 1
                    self.progress.emit(done, total)
                    self.status.emit(f"Saved calibration {label} rep {rep}: {wav_path}")
            profile.save_metadata()
            profile.save()
            self.prompt.emit("Calibration recordings completed. Training ML model...")
            self.status.emit("Training jamo ML model from current + pooled voice profiles...")
            bundle, path = train_hangul_jamo_model(self.subject, ALL_JAMO_LABELS, include_pooled=self.include_pooled)
            self.finished_calibration.emit(True, f"Voice calibration + ML training completed.\n{path}\nSamples: {bundle['n_samples']}")
        except Exception as e:
            self.finished_calibration.emit(False, f"Calibration failed: {e}")



def existing_voice_model_path(subject_id):
    return Path("voice_profiles") / str(subject_id).strip() / "learned_voice_model.joblib"


def existing_voice_model_available(subject_id):
    subject_id = str(subject_id).strip()
    paths = [
        Path("voice_profiles") / subject_id / "jamo_voice_model.joblib",
        Path("voice_profiles") / subject_id / "hangul_jamo_model.joblib",
        Path("voice_profiles") / subject_id / "learned_voice_model.joblib",
    ]
    return any(p.exists() for p in paths)










# Participant-facing jamo display mapping.
# In this learning app, the internal label is the compact jamo itself (e.g., ㄱ, ㅣ, ㅘ).
# Some legacy voice models use spoken names (e.g., 기역, 이, 와), so we convert those to jamo.
JAMO_TO_SPOKEN_MAP = {
    "ㄱ": "기역", "ㄲ": "쌍기역", "ㄴ": "니은", "ㄷ": "디귿", "ㄸ": "쌍디귿",
    "ㄹ": "리을", "ㅁ": "미음", "ㅂ": "비읍", "ㅃ": "쌍비읍",
    "ㅅ": "시옷", "ㅆ": "쌍시옷", "ㅇ": "이응", "ㅈ": "지읒", "ㅉ": "쌍지읒",
    "ㅊ": "치읓", "ㅋ": "키읔", "ㅌ": "티읕", "ㅍ": "피읖", "ㅎ": "히읗",
    "ㅏ": "아", "ㅐ": "애", "ㅑ": "야", "ㅒ": "얘", "ㅓ": "어", "ㅔ": "에",
    "ㅕ": "여", "ㅖ": "예", "ㅗ": "오", "ㅘ": "와", "ㅙ": "왜", "ㅚ": "외",
    "ㅛ": "요", "ㅜ": "우", "ㅝ": "워", "ㅞ": "웨", "ㅟ": "위", "ㅠ": "유",
    "ㅡ": "으", "ㅢ": "의", "ㅣ": "이",
}
SPOKEN_TO_JAMO_MAP = {v: k for k, v in JAMO_TO_SPOKEN_MAP.items()}

def display_label(label):
    """Participant-facing display. Jamo remains jamo; spoken names are converted to jamo."""
    label = str(label).strip()
    return SPOKEN_TO_JAMO_MAP.get(label, label)

def normalize_response_label(label):
    """Convert manual input/model labels into app-internal jamo if possible."""
    label = str(label).strip()
    return SPOKEN_TO_JAMO_MAP.get(label, label)

def model_class_to_app_label(label):
    """Convert model class label to current app label convention."""
    return normalize_response_label(label)




def normalize_candidate_for_session(label, session_type="jamo"):
    """Normalize a candidate according to session type.

    Jamo sessions:
        spoken names such as 기역/이 are converted to ㄱ/ㅣ.

    Syllable sessions:
        keep the actual syllable text such as 이/지/하.
        Do NOT convert 이 -> ㅣ, because syllable candidates must come from the syllable list.
    """
    label = str(label).strip()
    if str(session_type).lower() == "syllable":
        return label
    return normalize_response_label(label)


def display_candidate_for_session(label, session_type="jamo"):
    label = str(label).strip()
    if str(session_type).lower() == "syllable":
        return label
    return display_label(label)


class HangulLearningVoiceApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hangul Tactile Learning v20 - Voice Response")
        self.resize(1180, 820)
        self.setStyleSheet("""
            QLabel { font-size: 18px; }
            QGroupBox { font-size: 18px; font-weight: bold; margin-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
            QPushButton { font-size: 18px; padding: 10px 14px; min-height: 38px; }
            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox { font-size: 17px; min-height: 34px; }
            QTextEdit { font-size: 13px; }
        """)
        self.serial_port = None
        self.subject = ""
        self.day = None
        self.current_session_idx = 0
        self.current_trial_index = 0
        self.current_item = None
        self.current_options = []
        self.current_voice_result = None
        self.trial_start_time = None
        self.stimulus_sent_time = None
        self.answer_time = None
        self.can_answer = False
        self.waiting_for_next_trial = False
        self.all_rows = []
        self.session_trial_rows = []
        self.current_csv_path = None
        self.current_state_path = None
        self.rng = random.Random()
        self.calib_worker = None
        self.build_ui()
        self.refresh_ports()
        self.pending_after_feedback_action = None
        self.pending_after_feedback_message = ''
        self.awaiting_feedback_next = False

    def set_status_text(self, text):
        """Set user-facing status text safely across UI versions."""
        text = str(text)
        if hasattr(self, "status_label"):
            self.set_status_text(text)
        elif hasattr(self, "feedback_label"):
            self.feedback_label.setText(text)
        elif hasattr(self, "calib_prompt_label"):
            self.calib_prompt_label.setText(text)
        if hasattr(self, "log_box"):
            self.log_box.append(text)


    def build_ui(self):
        root = QWidget(); self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        top = QHBoxLayout()
        self.subject_edit = QLineEdit(); self.subject_edit.setPlaceholderText("Subject ID, e.g., s01")
        self.port_combo = QComboBox(); self.refresh_button = QPushButton("Refresh ports")
        self.connect_button = QPushButton("Connect")
        self.calibrate_button = QPushButton("Calibrate + Train")
        self.cancel_calib_button = QPushButton("Stop Calibration")
        self.cancel_calib_button.setEnabled(False)
        self.start_button = QPushButton("Start / Resume")
        self.stt_model_combo = QComboBox()
        self.stt_model_combo.addItems(["tiny", "base", "small"])
        self.stt_model_combo.setCurrentText("small")
        self.use_existing_model_check = QCheckBox("Use model")
        self.use_existing_model_check.setChecked(True)
        self.force_calibration_check = QCheckBox("Force calib")
        self.force_calibration_check.setChecked(False)
        self.check_model_button = QPushButton("Check Model")
        for w in [QLabel("Subject"), self.subject_edit, QLabel("COM"), self.port_combo, self.refresh_button, self.connect_button, self.calibrate_button, self.cancel_calib_button, self.use_existing_model_check, self.force_calibration_check, self.check_model_button, QLabel("STT"), self.stt_model_combo, self.start_button]:
            top.addWidget(w)
        layout.addLayout(top)

        self.voice_params_visible = True
        self.toggle_voice_params_button = QPushButton("Hide voice params")
        self.toggle_voice_params_button.setMaximumWidth(150)
        onset_group = QGroupBox("Voice detection parameters")
        self.onset_group = onset_group
        grid = QGridLayout(onset_group)
        grid.addWidget(self.toggle_voice_params_button, 0, 6)
        self.vad_level = QSpinBox(); self.vad_level.setRange(0,3); self.vad_level.setValue(2)
        self.speech_frames = QSpinBox(); self.speech_frames.setRange(1,10); self.speech_frames.setValue(2)
        self.pre_buffer = QSpinBox(); self.pre_buffer.setRange(0,50); self.pre_buffer.setValue(20)
        self.end_silence = QDoubleSpinBox(); self.end_silence.setRange(0.1,2.0); self.end_silence.setSingleStep(0.05); self.end_silence.setValue(0.45)
        self.noise_calib = QDoubleSpinBox(); self.noise_calib.setRange(0,2.0); self.noise_calib.setSingleStep(0.05); self.noise_calib.setValue(0.25)
        self.energy_mult = QDoubleSpinBox(); self.energy_mult.setRange(1.0,10.0); self.energy_mult.setSingleStep(0.1); self.energy_mult.setValue(2.0)
        self.min_rms = QDoubleSpinBox(); self.min_rms.setRange(10,5000); self.min_rms.setValue(100)
        self.min_record = QDoubleSpinBox(); self.min_record.setRange(0.05,1.0); self.min_record.setSingleStep(0.05); self.min_record.setValue(0.15)
        self.max_record = QDoubleSpinBox(); self.max_record.setRange(0.5,10.0); self.max_record.setSingleStep(0.5); self.max_record.setValue(3.0)
        vals = [("VAD level",self.vad_level),("Speech frames",self.speech_frames),("Pre-buffer",self.pre_buffer),("End silence",self.end_silence),("Noise calib",self.noise_calib),("Energy mult",self.energy_mult),("Min RMS",self.min_rms),("Min record",self.min_record),("Max record",self.max_record)]
        for spin in [self.vad_level,self.speech_frames,self.pre_buffer,self.end_silence,self.noise_calib,self.energy_mult,self.min_rms,self.min_record,self.max_record]:
            spin.setMaximumWidth(150)
        for i,(lab,w) in enumerate(vals):
            grid.addWidget(QLabel(lab), i//3, (i%3)*2); grid.addWidget(w, i//3, (i%3)*2+1)
        layout.addWidget(onset_group)

        self.interval_params_visible = True
        self.toggle_interval_params_button = QPushButton("Hide interval params")
        self.toggle_interval_params_button.setMaximumWidth(160)
        interval_group = QGroupBox("Stimulus interval parameters")
        self.interval_group = interval_group
        interval_grid = QGridLayout(interval_group)
        interval_grid.addWidget(self.toggle_interval_params_button, 0, 2)
        self.within_unit_interval_ms = QSpinBox()
        self.within_unit_interval_ms.setRange(0, 1000)
        self.within_unit_interval_ms.setSingleStep(10)
        self.within_unit_interval_ms.setValue(150)
        self.within_unit_interval_ms.setMaximumWidth(150)
        self.cv_interval_ms = QSpinBox()
        self.cv_interval_ms.setRange(0, 1500)
        self.cv_interval_ms.setSingleStep(10)
        self.cv_interval_ms.setValue(250)
        self.cv_interval_ms.setMaximumWidth(150)
        interval_grid.addWidget(QLabel("Within consonant/vowel unit interval (ms)"), 0, 0)
        interval_grid.addWidget(self.within_unit_interval_ms, 0, 1)
        interval_grid.addWidget(QLabel("Consonant–vowel interval (ms)"), 1, 0)
        interval_grid.addWidget(self.cv_interval_ms, 1, 1)
        layout.addWidget(interval_group)


        calib_group = QGroupBox("Voice Calibration Status")
        calib_layout = QGridLayout(calib_group)
        self.calib_prompt_label = QLabel("Calibration: not started")
        self.calib_prompt_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #2F2A7A; padding: 10px;")
        self.calib_progress = QProgressBar()
        self.calib_progress.setRange(0, 100)
        self.calib_progress.setValue(0)
        self.calib_progress.setFormat("%v / %m")
        calib_layout.addWidget(self.calib_prompt_label, 0, 0, 1, 3)
        calib_layout.addWidget(self.calib_progress, 1, 0, 1, 3)
        layout.addWidget(calib_group)

        self.day_label = QLabel("Day: -")
        self.session_label = QLabel("Session: -")
        self.progress_label = QLabel("Progress: -")
        self.target_label = QLabel("Target: -")
        self.feedback_label = QLabel("Press Stimulus. Voice detection starts exactly when stimulus is sent.")
        for lab in [self.day_label, self.session_label, self.progress_label, self.target_label, self.feedback_label]:
            lab.setStyleSheet("font-size: 20px; font-weight: bold; padding: 6px;")
        self.stim_button = QPushButton("STIMULUS")
        self.stim_button.setMinimumHeight(64)
        self.stim_button.setStyleSheet("font-size: 20px; font-weight: bold; background-color: #3E3A7A; color: white; border-radius: 14px; padding: 18px;")
        self.stim_button.setEnabled(False)
        info = QHBoxLayout()
        for w in [self.day_label, self.session_label, self.progress_label, self.target_label, self.stim_button]:
            info.addWidget(w)
        layout.addLayout(info)
        layout.addWidget(self.feedback_label)

        self.option_area = QScrollArea(); self.option_widget = QWidget(); self.option_layout = QGridLayout(self.option_widget)
        self.option_area.setWidgetResizable(True); self.option_area.setWidget(self.option_widget)
        self.option_area.setMinimumHeight(130)
        layout.addWidget(self.option_area, stretch=2)
        manual_row = QHBoxLayout()
        self.no_answer_button = QPushButton("답이 없음")
        self.no_answer_button.setMinimumHeight(48)
        self.manual_input = QLineEdit(); self.manual_input.setPlaceholderText("Top-5에 답이 없으면 직접 입력: ㄱ, ㅒ, 기역, 얘 모두 가능")
        self.manual_input.setMinimumHeight(46)
        self.manual_submit = QPushButton("Submit manual")
        self.continue_button = QPushButton("Next Trial")
        self.continue_button.setMinimumHeight(48)
        self.continue_button.setEnabled(False)
        self.continue_button.hide()
        self.continue_button.setStyleSheet("font-size: 18px; font-weight: bold; background-color: #3E3A7A; color: white; border-radius: 8px; padding: 8px;")
        self.manual_submit.setMinimumHeight(48)
        for w in [self.no_answer_button, self.manual_input, self.manual_submit, self.continue_button]: manual_row.addWidget(w)
        layout.addLayout(manual_row)
        log_row = QHBoxLayout()
        self.log_title = QLabel("Debug log: calibration 저장 경로 / 음성 인식 결과 / 오류 메시지 확인용")
        self.toggle_log_button = QPushButton("Hide log")
        self.toggle_log_button.setMaximumWidth(130)
        log_row.addWidget(self.log_title)
        log_row.addWidget(self.toggle_log_button)
        layout.addLayout(log_row)
        self.log_box = QTextEdit(); self.log_box.setReadOnly(True); self.log_box.setMaximumHeight(70); layout.addWidget(self.log_box, stretch=0)

        self.refresh_button.clicked.connect(self.refresh_ports)
        self.connect_button.clicked.connect(self.toggle_serial)
        self.calibrate_button.clicked.connect(self.run_voice_calibration)
        self.cancel_calib_button.clicked.connect(self.cancel_voice_calibration)
        self.start_button.clicked.connect(self.start_learning)
        self.check_model_button.clicked.connect(self.check_existing_voice_model)
        self.stim_button.clicked.connect(self.on_stimulus_clicked)
        self.no_answer_button.clicked.connect(lambda: self.manual_input.setFocus())
        self.manual_submit.clicked.connect(self.submit_manual_answer)
        self.continue_button.clicked.connect(self.continue_after_feedback)
        self.manual_input.returnPressed.connect(self.submit_manual_answer)
        self.toggle_log_button.clicked.connect(self.toggle_log_visibility)
        self.toggle_voice_params_button.clicked.connect(self.toggle_voice_params)
        self.toggle_interval_params_button.clicked.connect(self.toggle_interval_params)


    def toggle_voice_params(self):
        self.voice_params_visible = not getattr(self, "voice_params_visible", True)
        widgets = [
            getattr(self, "vad_level", None),
            getattr(self, "speech_frames", None),
            getattr(self, "pre_buffer", None),
            getattr(self, "end_silence", None),
            getattr(self, "noise_calib", None),
            getattr(self, "energy_mult", None),
            getattr(self, "min_rms", None),
            getattr(self, "min_record", None),
            getattr(self, "max_record", None),
        ]
        # Hide child widgets except the group title and toggle button.
        if hasattr(self, "onset_group"):
            for child in self.onset_group.findChildren(QWidget):
                if child is not getattr(self, "toggle_voice_params_button", None):
                    child.setVisible(self.voice_params_visible)
        self.toggle_voice_params_button.setText("Hide voice params" if self.voice_params_visible else "Show voice params")
        if hasattr(self, "onset_group"):
            self.onset_group.setMaximumHeight(16777215 if self.voice_params_visible else 55)

    def toggle_interval_params(self):
        self.interval_params_visible = not getattr(self, "interval_params_visible", True)
        if hasattr(self, "interval_group"):
            for child in self.interval_group.findChildren(QWidget):
                if child is not getattr(self, "toggle_interval_params_button", None):
                    child.setVisible(self.interval_params_visible)
        self.toggle_interval_params_button.setText("Hide interval params" if self.interval_params_visible else "Show interval params")
        if hasattr(self, "interval_group"):
            self.interval_group.setMaximumHeight(16777215 if self.interval_params_visible else 55)


    def toggle_log_visibility(self):
        visible = self.log_box.isVisible()
        self.log_box.setVisible(not visible)
        self.toggle_log_button.setText("Show log" if visible else "Hide log")

    def log_msg(self, msg):
        self.log_box.append(str(msg)); print(msg, flush=True)

    def recorder_settings(self):
        return dict(vad_level=self.vad_level.value(), speech_frames_required=self.speech_frames.value(), pre_buffer_frames=self.pre_buffer.value(), max_wait_sec=8.0, max_record_sec=self.max_record.value(), end_silence_sec=self.end_silence.value(), noise_calibration_sec=self.noise_calib.value(), energy_multiplier=self.energy_mult.value(), min_energy_rms=self.min_rms.value(), min_record_sec=self.min_record.value())

    def refresh_ports(self):
        self.port_combo.clear()
        if serial is None:
            self.port_combo.addItem("pyserial not installed")
            return
        ports = list(serial.tools.list_ports.comports())
        for p in ports:
            self.port_combo.addItem(f"{p.device} - {p.description}", p.device)
        if not ports:
            self.port_combo.addItem("No ports", None)

    def toggle_serial(self):
        if serial is None:
            QMessageBox.warning(self, "No serial", "Install pyserial first.")
            return
        if self.serial_port and self.serial_port.is_open:
            self.serial_port.close(); self.serial_port = None; self.connect_button.setText("Connect"); return
        port = self.port_combo.currentData()
        if not port:
            QMessageBox.warning(self, "No port", "Select COM port."); return
        try:
            self.serial_port = serial.Serial(port, BAUDRATE, timeout=1)
            self.connect_button.setText("Disconnect")
        except Exception as e:
            QMessageBox.critical(self, "Serial error", str(e))


    def current_stimulus_intervals(self):
        within_ms = self.within_unit_interval_ms.value() if hasattr(self, "within_unit_interval_ms") else 150
        cv_ms = self.cv_interval_ms.value() if hasattr(self, "cv_interval_ms") else 250
        return int(within_ms), int(cv_ms)

    def interval_command_gap(self, gap_ms):
        gap_ms = int(gap_ms)
        if gap_ms <= 0:
            return ""
        # Serial command sequence uses comma-separated temporal gaps in the existing protocol.
        # This string is used only if the current stimulus command builder supports explicit gaps.
        return f"wait/{gap_ms}"


    def send_serial_command(self, body):
        cmd = f"@{body}\n"
        if self.serial_port and self.serial_port.is_open:
            try: self.serial_port.write(cmd.encode("utf-8"))
            except Exception as e:
                QMessageBox.critical(self, "Serial write failed", str(e)); return False
        else:
            print(f"[SIM SERIAL] {cmd.strip()}", flush=True)
        return True

    def subject_dir(self):
        p = OUTPUT_DIR / f"subject_{self.subject}"; p.mkdir(parents=True, exist_ok=True); return p
    def voice_trial_dir(self):
        p = self.subject_dir() / "voice_trials"; p.mkdir(exist_ok=True); return p
    def today_str(self): return date.today().isoformat()
    def today_csv_path(self): return self.subject_dir() / f"subject_{self.subject}_{self.today_str()}.csv"
    def today_state_path(self): return self.subject_dir() / f"subject_{self.subject}_{self.today_str()}.json"


    def existing_model_available_for_current_subject(self):
        subject = self.subject_edit.text().strip()
        if not subject:
            return False
        return load_hangul_jamo_model(subject) is not None or existing_voice_model_available(subject)

    def check_existing_voice_model(self):
        subject = self.subject_edit.text().strip()
        if not subject:
            QMessageBox.warning(self, "Missing subject", "Enter subject ID first.")
            return False

        jamo_model_ok = load_hangul_jamo_model(subject) is not None
        legacy_model_path = existing_voice_model_path(subject)
        legacy_model_ok = legacy_model_path.exists()

        lines = [
            f"Subject: {subject}",
            f"Jamo ML model available: {jamo_model_ok}",
            f"Legacy learned_voice_model.joblib available: {legacy_model_ok}",
            f"Legacy path: {legacy_model_path}",
        ]

        if jamo_model_ok or legacy_model_ok:
            lines.append("")
            lines.append("Calibration can be skipped.")
            self.set_status_text("Existing voice model found.")
        else:
            lines.append("")
            lines.append("No usable model found. Calibration is recommended.")
            self.set_status_text("No voice model found.")

        msg = "\n".join(lines)
        self.log_msg("[Voice Model Check]\n" + msg)
        QMessageBox.information(self, "Voice Model Check", msg)
        return jamo_model_ok or legacy_model_ok

    def should_skip_calibration_from_existing_model(self):
        if hasattr(self, "force_calibration_check") and self.force_calibration_check.isChecked():
            return False
        if hasattr(self, "use_existing_model_check") and not self.use_existing_model_check.isChecked():
            return False
        return self.existing_model_available_for_current_subject()


    def run_voice_calibration(self):
        subject = self.subject_edit.text().strip()
        if not subject:
            QMessageBox.warning(self, "Missing subject", "Enter subject ID first."); return
        if self.calib_worker is not None and self.calib_worker.isRunning():
            QMessageBox.information(self, "Calibration running", "Calibration is already running."); return
        self.subject = subject
        if self.should_skip_calibration_from_existing_model():
            reply = QMessageBox.question(
                self,
                "Existing model found",
                "Existing model is available. Re-run calibration anyway?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply == QMessageBox.No:
                self.calib_prompt_label.setText("Existing model loaded. Calibration skipped.")
                self.log_msg("[Voice Calibration] Skipped because existing model is available.")
                return

        self.calib_prompt_label.setText("Calibration 준비 중...")
        self.calib_progress.setValue(0)
        self.calibrate_button.setEnabled(False)
        self.cancel_calib_button.setEnabled(True)
        self.start_button.setEnabled(False)
        self.log_msg("Voice calibration started in the main window. You can stop it with Stop Calibration.")
        self.calib_worker = VoiceCalibrationWorker(subject, self.recorder_settings(), reps=3, include_pooled=True)
        self.calib_worker.prompt.connect(self.calib_prompt_label.setText)
        self.calib_worker.status.connect(self.log_msg)
        self.calib_worker.progress.connect(self.on_calibration_progress)
        self.calib_worker.finished_calibration.connect(self.on_calibration_finished)
        self.calib_worker.start()

    def cancel_voice_calibration(self):
        if self.calib_worker is not None and self.calib_worker.isRunning():
            self.calib_prompt_label.setText("Calibration 중단 요청됨. 현재 녹음이 끝나면 멈춥니다...")
            self.calib_worker.stop()
            self.cancel_calib_button.setEnabled(False)

    def on_calibration_progress(self, done, total):
        self.calib_progress.setRange(0, max(1, total))
        self.calib_progress.setValue(done)
        self.calib_progress.setFormat(f"{done} / {total}")

    def on_calibration_finished(self, ok, message):
        self.calibrate_button.setEnabled(True)
        self.cancel_calib_button.setEnabled(False)
        self.start_button.setEnabled(True)
        self.log_msg(message)
        self.calib_prompt_label.setText("Calibration 완료" if ok else "Calibration 중단/실패")
        if ok:
            QMessageBox.information(self, "Calibration done", message)
        else:
            QMessageBox.warning(self, "Calibration stopped", message)


    def rows_for_session_idx(self, session_idx):
        """Rows from today's CSV belonging to one session index."""
        return [
            r for r in self.all_rows
            if str(r.get("session_idx", "")) == str(session_idx + 1)
        ]

    def first_incomplete_session_idx_from_rows(self):
        """Find the first session not completed in today's saved rows."""
        if not self.all_rows:
            return 0

        old_idx = self.current_session_idx
        old_rows = self.session_trial_rows

        try:
            for idx in range(len(SESSIONS)):
                self.current_session_idx = idx
                self.session_trial_rows = self.rows_for_session_idx(idx)
                if not self.check_session_completed():
                    return idx
            return len(SESSIONS) - 1
        finally:
            self.current_session_idx = old_idx
            self.session_trial_rows = old_rows

    def restore_session_state_from_rows(self, session_idx):
        """Restore global and session-local trial indices from today's CSV."""
        session_idx = max(0, min(int(session_idx), len(SESSIONS) - 1))
        self.current_session_idx = session_idx
        self.session_trial_rows = self.rows_for_session_idx(session_idx)
        self.current_trial_index = len(self.all_rows)

        self.log_msg(
            f"[Resume] Global next index={self.current_trial_index + 1}, "
            f"Session={self.current_session_idx + 1}, "
            f"Session next trial={len(self.session_trial_rows) + 1}"
        )

    def ask_resume_start_mode(self):
        """Ask how to continue when today's data already exist."""
        if not self.all_rows:
            return 0

        first_incomplete = self.first_incomplete_session_idx_from_rows()

        choices = [
            f"Continue from unfinished session: {first_incomplete + 1}. {SESSIONS[first_incomplete]['title']}",
            "Start from Session 1",
            "Choose session manually",
        ]

        choice, ok = QInputDialog.getItem(
            self,
            "Resume learning",
            "Today's data already exist for this subject. Choose start position:",
            choices,
            0,
            False,
        )
        if not ok:
            return None

        if choice.startswith("Continue"):
            return first_incomplete

        if choice.startswith("Start from Session 1"):
            return 0

        session_choices = [f"{i+1}. {s['title']}" for i, s in enumerate(SESSIONS)]
        selected, ok = QInputDialog.getItem(
            self,
            "Choose session",
            "Start from which session?",
            session_choices,
            first_incomplete,
            False,
        )
        if not ok:
            return None

        try:
            return int(selected.split(".")[0]) - 1
        except Exception:
            return first_incomplete



    def build_pooled_jamo_model_if_needed(self):
        """Create a jamo ML model for the current subject using other subjects' voice profiles.

        This is used when the current participant has no calibration/model yet.
        It allows the experiment to start with a population-level pooled model.
        """
        if load_hangul_jamo_model(self.subject) is not None:
            return True

        if not HAS_SKLEARN or joblib is None:
            self.log_msg("[Voice Model] scikit-learn/joblib not available; cannot build pooled model.")
            return False

        try:
            bundle, path = train_hangul_jamo_model(
                self.subject,
                labels=ALL_JAMO_LABELS,
                include_pooled=True,
            )
            self.log_msg(
                f"[Voice Model] Built pooled jamo model for {self.subject}: {path} "
                f"(samples={bundle.get('n_samples', '')}, pooled_only={bundle.get('pooled_only', False)})"
            )
            return True
        except Exception as e:
            self.log_msg(f"[Voice Model] Failed to build pooled jamo model: {e}")
            return False


    def start_learning(self):
        subject = self.subject_edit.text().strip()
        if not subject:
            QMessageBox.warning(self, "Missing subject", "Enter subject ID.")
            return

        self.subject = subject
        self.day = self.today_str()
        self.current_csv_path = self.today_csv_path()
        self.current_state_path = self.today_state_path()

        # Load today's data only. If there is no file for today, start from Session 1.
        if self.current_csv_path.exists():
            try:
                with open(self.current_csv_path, "r", encoding="utf-8-sig", newline="") as f:
                    self.all_rows = [dict(r) for r in csv.DictReader(f)]
            except Exception:
                self.all_rows = []
        else:
            self.all_rows = []

        if self.all_rows:
            requested_idx = self.ask_resume_start_mode()
            if requested_idx is None:
                return
            self.restore_session_state_from_rows(requested_idx)
        else:
            self.current_trial_index = 0
            self.current_session_idx = 0
            self.session_trial_rows = []
            self.log_msg("[Resume] No data for today. Starting from Session 1.")

        if load_hangul_jamo_model(self.subject) is None:
            if self.should_skip_calibration_from_existing_model():
                self.log_msg("[Voice Calibration] Existing voice model selected. Skipping calibration.")
            else:
                built = self.build_pooled_jamo_model_if_needed()
                if not built:
                    reply = QMessageBox.question(
                        self,
                        "No voice model",
                        "No current or pooled jamo ML model could be built.\n\n"
                        "Run calibration now?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes,
                    )
                    if reply == QMessageBox.Yes:
                        self.run_voice_calibration()
                        return
                else:
                    QMessageBox.information(
                        self,
                        "Pooled voice model",
                        "No current-subject voice model was found.\n"
                        "A pooled jamo ML model was built from other participants' voice_profiles.\n\n"
                        "You can start the experiment now, but subject-specific calibration is still recommended when possible."
                    )

        self.update_ui_for_session()
        self.next_trial()


    def current_session(self): return SESSIONS[self.current_session_idx]
    def current_items(self): return self.current_session()["items"]
    def is_syllable_session(self): return self.current_session()["type"] == "syllable"

    def update_ui_for_session(self):
        sess = self.current_session()
        self.day_label.setText(f"Day: {self.day}")
        self.session_label.setText(sess["title"])
        self.update_progress_label()

    def update_progress_label(self):
        self.progress_label.setText(f"Session {self.current_session_idx+1}/{len(SESSIONS)} | Total trials: {self.current_trial_index}")

    def next_trial(self):
        self.session_trial_rows = self.rows_for_session_idx(self.current_session_idx)
        items = self.current_items()
        if not items:
            QMessageBox.warning(self, "No items", f"No items in {self.current_session()['name']}."); return
        counts = Counter([r.get("target_label", "") for r in self.session_trial_rows])
        min_count = min([counts.get(item.label,0) for item in items])
        pool = [item for item in items if counts.get(item.label,0) == min_count]
        self.current_item = self.rng.choice(pool)
        self.current_options = [item.label for item in items]
        self.current_voice_result = None
        self.awaiting_feedback_next = False
        self.can_answer = False
        self.trial_start_time = time.time(); self.stimulus_sent_time = None; self.answer_time = None
        self.target_label.setText("Target: hidden")
        self.feedback_label.setText("Press Stimulus. Voice detection starts when stimulus is sent.")
        self.clear_options()
        self.manual_input.clear()
        self.stim_button.setText("Stimulus"); self.stim_button.setEnabled(True)
        if hasattr(self, "continue_button"):
            self.continue_button.setEnabled(False)
            self.continue_button.setText("Next Trial")
        self.update_progress_label()

    def on_stimulus_clicked(self):
        if getattr(self, "awaiting_feedback_next", False):
            self.continue_after_feedback()
            return
        if self.current_item is None: return
        self.stim_button.setEnabled(False)
        self.feedback_label.setText("Stimulus will start after 500 ms...")
        QTimer.singleShot(STIM_DELAY_MS, self.send_stimulus_and_start_voice)
    def send_stimulus_and_start_voice(self):
        within_interval_ms, cv_interval_ms = self.current_stimulus_intervals()
        self.stimulus_sent_time = time.time()
        ok = self.send_serial_command(self.current_item.command_body)
        if not ok:
            self.stim_button.setEnabled(True); return
        self.feedback_label.setText("Stimulus sent. Listening for voice...")
        sess_type = self.current_session()["type"]
        candidate_labels = [item.label for item in self.current_items()]
        self.voice_worker = TrialVoiceWorker(self.subject, sess_type, candidate_labels, self.recorder_settings(), syllable_xlsx=str(SYLLABLE_FILE), stt_model=self.stt_model_combo.currentText() if hasattr(self, "stt_model_combo") else "small", wav_dir=self.voice_trial_dir())
        self.voice_worker.status.connect(self.feedback_label.setText)
        self.voice_worker.result_ready.connect(self.handle_voice_result)
        self.log_msg(f"[Voice Route] session_type={self.current_session()['type']} candidates={len(self.current_options)}")
        self.voice_worker.start()




    def extract_candidate_labels(self, result):
        """Return up to 5 labels from any recognition result format.

        Important:
        - jamo sessions normalize spoken names, e.g., 기역 -> ㄱ
        - syllable sessions keep syllables, e.g., 이 remains 이, not ㅣ
        """
        session_type = self.current_session().get("type", "jamo")
        candidates = []

        def add_label(x):
            if x is None:
                return
            lab = None
            if isinstance(x, dict):
                lab = (
                    x.get("label")
                    or x.get("predicted_label")
                    or x.get("text")
                    or x.get("candidate")
                    or x.get("class")
                )
            elif isinstance(x, (list, tuple)) and len(x) > 0:
                lab = x[0]
            else:
                lab = x
            if lab is None:
                return

            lab = normalize_candidate_for_session(str(lab).strip(), session_type)

            # In syllable sessions, restrict candidates to current session list only.
            if session_type == "syllable":
                allowed = set(str(it.label).strip() for it in self.current_items())
                if lab not in allowed:
                    return

            if lab and lab not in candidates:
                candidates.append(lab)

        for key in ["predicted_label", "prediction", "pred", "best_label", "top_label"]:
            add_label(result.get(key))

        for key in ["second_label", "second_best", "runner_up"]:
            add_label(result.get(key))

        for key in ["top_candidates", "candidates", "ranked", "top5", "top_labels", "candidate_labels", "top"]:
            vals = result.get(key)
            if vals:
                for v in vals:
                    add_label(v)

        scores = result.get("scores") or result.get("score_list") or result.get("probs") or []
        if scores:
            try:
                if isinstance(scores[0], dict):
                    if any("ml_probability" in s for s in scores):
                        scores_sorted = sorted(scores, key=lambda s: s.get("ml_probability", 0), reverse=True)
                    elif any("probability" in s for s in scores):
                        scores_sorted = sorted(scores, key=lambda s: s.get("probability", 0), reverse=True)
                    elif any("prob" in s for s in scores):
                        scores_sorted = sorted(scores, key=lambda s: s.get("prob", 0), reverse=True)
                    elif any("score" in s for s in scores):
                        scores_sorted = sorted(scores, key=lambda s: s.get("score", 0), reverse=True)
                    else:
                        scores_sorted = scores
                else:
                    scores_sorted = scores
            except Exception:
                scores_sorted = scores

            for s in scores_sorted:
                add_label(s)

        # If still empty, fallback to current session labels only.
        if not candidates:
            try:
                for it in self.current_items():
                    add_label(it.label)
                    if len(candidates) >= 5:
                        break
            except Exception:
                pass

        # Last-resort fallback only for jamo sessions.
        if not candidates and session_type != "syllable":
            for lab in ["ㄱ", "ㄴ", "ㄷ", "ㄹ", "ㅁ"]:
                add_label(lab)

        return candidates[:5]


    def show_candidate_buttons(self, candidate_labels):
        """Render response buttons. Always clears old option area first."""
        self.clear_options()

        if not candidate_labels:
            msg = QLabel("인식 후보가 없습니다. 아래에 직접 입력해주세요.")
            msg.setStyleSheet("font-size: 24px; font-weight: bold; padding: 18px;")
            self.option_layout.addWidget(msg, 0, 0, 1, 5)
            return

        for i, lab in enumerate(candidate_labels):
            disp = display_candidate_for_session(lab, self.current_session().get("type", "jamo"))
            btn = QPushButton(f"{i+1}. {disp}")
            btn.setMinimumHeight(86)
            btn.setMinimumWidth(150)
            btn.setStyleSheet(
                "font-size: 34px; font-weight: bold; padding: 14px; "
                "border: 2px solid #3E3A7A; border-radius: 12px; background-color: white;"
            )
            btn.setEnabled(True)
            btn.setEnabled(True)
            btn.clicked.connect(lambda _checked=False, ans=lab: self.finalize_answer(ans))
            self.option_layout.addWidget(btn, 0, i)


    def clear_options(self):
        while self.option_layout.count():
            item = self.option_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()



    def reset_for_voice_retry(self, message="Voice was not detected. Press Stimulus again."):
        """Return UI to a retryable state after voice detection/recognition failure."""
        self.can_answer = False
        self.awaiting_feedback_next = False
        self.current_voice_result = None

        if hasattr(self, "stim_button"):
            self.stim_button.setText("Stimulus")
            self.stim_button.setEnabled(True)

        if hasattr(self, "feedback_label"):
            self.feedback_label.setText(message)

        if hasattr(self, "manual_input"):
            self.manual_input.clear()

        try:
            self.clear_options()
        except Exception:
            pass

        self.log_msg("[Voice Retry] " + str(message))


    def handle_voice_result(self, result):
        self.current_voice_result = result

        if "error" in result:
            msg = f"Voice recognition failed: {result['error']}. Press Stimulus again."
            self.reset_for_voice_retry(msg)
            return


        if result.get("no_speech") or result.get("voice_detected") is False:
            self.reset_for_voice_retry("Voice was not detected. Press Stimulus again.")
            return

        engine = result.get("engine", result.get("voice_engine", result.get("mode", "voice")))
        rt = result.get("voice_onset_rt_sec", result.get("voice_rt", result.get("rt_from_stimulus_sec", None)))

        if rt is not None:
            self.feedback_label.setText(f"Voice recognized by {engine} | onset RT={float(rt):.3f}s. Select answer or type manually.")
        else:
            self.feedback_label.setText(f"Voice recognized by {engine}. Select answer or type manually.")

        candidate_labels = self.extract_candidate_labels(result)
        if not candidate_labels:
            self.log_msg(f"raw result keys: {list(result.keys())}")
        self.can_answer = True
        self.stim_button.setEnabled(False)
        self.show_candidate_buttons(candidate_labels)

        self.log_msg("\\n[Voice Result]")
        self.log_msg(f"engine: {engine}")
        if result.get("stt_model"):
            self.log_msg(f"stt_model: {result.get('stt_model')}")
        if result.get("stt_raw") or result.get("stt_norm"):
            self.log_msg(f"stt_raw: {result.get('stt_raw', '')}")
            self.log_msg(f"stt_norm: {result.get('stt_norm', '')}")
        if rt is not None:
            self.log_msg(f"voice onset RT: {float(rt):.3f}s")
        self.log_msg(f"predicted: {result.get('predicted_label', '')}")
        self.log_msg(f"candidates: {', '.join([display_label(c) for c in candidate_labels])}")


    def submit_manual_answer(self):
        # Manual jamo input such as ㄱ/ㅣ is converted to internal label if needed.
        ans = normalize_candidate_for_session(self.manual_input.text().strip(), self.current_session().get("type", "jamo"))
        if not ans:
            QMessageBox.warning(self, "No manual answer", "Type the answer first."); return
        self.finalize_answer(ans, manual=True)

    def show_answer_feedback(self, selected, correct):
        """Show explicit colored learning feedback and wait for user to continue."""
        self.clear_options()

        target = self.current_item.label if self.current_item is not None else ""
        selected_disp = display_candidate_for_session(selected, self.current_session().get("type", "jamo"))
        target_disp = display_candidate_for_session(target, self.current_session().get("type", "jamo"))

        if correct:
            title = QLabel(f"정답입니다!  {target_disp}")
            title.setStyleSheet(
                "font-size: 34px; font-weight: bold; color: #0B7A0B; padding: 14px;"
            )
            self.option_layout.addWidget(title, 0, 0, 1, 5)

            correct_btn = QPushButton(f"정답: {target_disp}")
            correct_btn.setMinimumHeight(90)
            correct_btn.setStyleSheet(
                "font-size: 36px; font-weight: bold; color: white; "
                "background-color: #1F9D45; border-radius: 14px; padding: 12px;"
            )
            self.option_layout.addWidget(correct_btn, 1, 0, 1, 5)
        else:
            title = QLabel("오답입니다. 정답을 확인하세요.")
            title.setStyleSheet(
                "font-size: 32px; font-weight: bold; color: #B00020; padding: 14px;"
            )
            self.option_layout.addWidget(title, 0, 0, 1, 5)

            wrong_btn = QPushButton(f"내 응답: {selected_disp}")
            wrong_btn.setMinimumHeight(85)
            wrong_btn.setStyleSheet(
                "font-size: 32px; font-weight: bold; color: white; "
                "background-color: #D64545; border-radius: 14px; padding: 12px;"
            )
            self.option_layout.addWidget(wrong_btn, 1, 0, 1, 2)

            arrow = QLabel("→")
            arrow.setAlignment(Qt.AlignCenter)
            arrow.setStyleSheet("font-size: 42px; font-weight: bold; padding: 10px;")
            self.option_layout.addWidget(arrow, 1, 2)

            correct_btn = QPushButton(f"정답: {target_disp}")
            correct_btn.setMinimumHeight(85)
            correct_btn.setStyleSheet(
                "font-size: 32px; font-weight: bold; color: white; "
                "background-color: #1F9D45; border-radius: 14px; padding: 12px;"
            )
            self.option_layout.addWidget(correct_btn, 1, 3, 1, 2)

        self.feedback_label.setText(
            f"선택: {selected_disp}  |  정답: {target_disp}  |  {'Correct' if correct else 'Wrong'}"
        )

        self.feedback_label.setText(self.feedback_label.text() + "  |  확인 후 Stimulus/Next 버튼을 누르세요.")

    def continue_after_feedback(self):
        """Move to next trial/session after participant confirms feedback with Stimulus/Next button."""
        self.awaiting_feedback_next = False

        if hasattr(self, "continue_button"):
            self.continue_button.setEnabled(False)
            self.continue_button.hide()

        action = getattr(self, "pending_after_feedback_action", None)
        msg = getattr(self, "pending_after_feedback_message", "")

        self.pending_after_feedback_action = None
        self.pending_after_feedback_message = ""

        if action == "advance":
            self.advance_session()
        elif action == "failed":
            QMessageBox.information(self, "Session ended", msg or "Max trials reached. Moving to next session.")
            self.advance_session()
        else:
            self.next_trial()


    def estimate_current_stimulus_duration_sec(self):
        """Estimate tactile stimulus duration from the current command string.

        Command conventions in this app commonly include motor commands such as
        "1/300" and temporal gaps such as "wait/150". This function sums the
        millisecond values to estimate when the tactile stimulus sequence ends.
        """
        if self.current_item is None:
            return 0.0

        cmd = str(getattr(self.current_item, "command_body", "") or "")
        nums = []

        # Sum durations from token patterns like "1/300", "wait/150", "5/d,2/i/100".
        for m in re.finditer(r"/(\d+(?:\.\d+)?)", cmd):
            try:
                nums.append(float(m.group(1)))
            except Exception:
                pass

        # If parsing fails, fall back to 0 so rt_from_stimulus equals rt_sec.
        if not nums:
            return 0.0

        return sum(nums) / 1000.0


    def finalize_answer(self, selected, manual=False):
        selected = normalize_candidate_for_session(selected, self.current_session().get("type", "jamo"))
        if not self.can_answer or self.current_item is None:
            return

        self.answer_time = time.time()
        correct = int(str(selected).strip() == str(self.current_item.label).strip())

        # rt_sec: stimulus onset to voice onset.
        # rt_from_stimulus_sec: voice onset after the stimulus sequence ends.
        rt_sec = None
        if self.current_voice_result:
            rt_sec = self.current_voice_result.get("voice_onset_rt_sec")
        try:
            rt_sec = float(rt_sec)
        except Exception:
            rt_sec = ""

        stim_duration_sec = self.estimate_current_stimulus_duration_sec()
        if rt_sec == "":
            rt_from_stimulus = ""
        else:
            rt_from_stimulus = max(0.0, float(rt_sec) - float(stim_duration_sec))

        within_interval_ms, cv_interval_ms = self.current_stimulus_intervals()

        row = {
            "subject": self.subject,
            "day": self.day,
            "session_idx": self.current_session_idx + 1,
            "session_name": self.current_session()["name"],
            "session_title": self.current_session()["title"],
            "trial_index_global": self.current_trial_index + 1,
            "trial_index_session": len(self.session_trial_rows) + 1,
            "target_label": self.current_item.label,
            "target_command": self.current_item.command_body,
            "stimulus_duration_sec": stim_duration_sec,
            "within_unit_interval_ms": within_interval_ms,
            "cv_interval_ms": cv_interval_ms,
            "response_mode": "voice_ml" if not self.is_syllable_session() else "voice_stt_list",
            "option_count": len(self.current_options),
            "options": "|".join(self.current_options),
            "selected_label": selected,
            "manual_response": int(manual),
            "correct": correct,
            "rt_sec": rt_sec,
            "rt_from_stimulus_sec": rt_from_stimulus,
            "trial_elapsed_sec": self.answer_time - self.trial_start_time if self.trial_start_time else "",
            "session_elapsed_sec": "",
            "day_elapsed_sec": "",
            "trial_start_time": self.trial_start_time,
            "stimulus_sent_time": self.stimulus_sent_time,
            "answer_time": self.answer_time,
            "voice_engine": self.current_voice_result.get("engine", "") if self.current_voice_result else "",
            "voice_wav_path": self.current_voice_result.get("wav_path", "") if self.current_voice_result else "",
            "voice_stt_raw": self.current_voice_result.get("stt_raw", "") if self.current_voice_result else "",
            "voice_stt_norm": self.current_voice_result.get("stt_norm", "") if self.current_voice_result else "",
            "voice_process_time_sec": self.current_voice_result.get("process_time", "") if self.current_voice_result else "",
            "voice_recorded_duration_sec": self.current_voice_result.get("recorded_duration", "") if self.current_voice_result else "",
            "top_candidates": json.dumps(self.current_voice_result.get("top", []), ensure_ascii=False) if self.current_voice_result else "",
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        }

        self.all_rows.append(row)
        self.session_trial_rows.append(row)
        self.current_trial_index += 1
        self.save_csv()
        self.save_state()

        self.can_answer = False
        self.show_answer_feedback(selected, bool(correct))
        self.awaiting_feedback_next = True
        self.stim_button.setText("Next")
        self.stim_button.setEnabled(True)

        if self.check_session_completed():
            self.pending_after_feedback_action = "advance"
            self.pending_after_feedback_message = "Session completed."
        elif self.check_session_failed():
            self.pending_after_feedback_action = "failed"
            self.pending_after_feedback_message = "Max trials reached for at least one item. Moving to next session."
        else:
            self.pending_after_feedback_action = "next"
            self.pending_after_feedback_message = ""


    def check_session_completed(self):
        """Return True when the current learning session should advance.

        Syllable sessions:
        - one exposure per syllable is enough.

        Jamo sessions:
        A) Perfect streak criterion:
           - recent window = number of labels × 2
           - if every trial in that recent window is correct, session advances.
           - e.g., 7 labels -> 14 consecutive correct trials.

        OR

        B) Accuracy criterion:
           - recent window = number of labels × 4
           - if accuracy in that window >= 90%, session advances.
           - e.g., 7 labels -> recent 28 trials at >= 90%.
        """
        items = self.current_items()
        if not items:
            return False

        # Syllable sessions: each syllable only needs to be experienced once.
        if self.is_syllable_session():
            seen = set(r.get("target_label", "") for r in self.session_trial_rows)
            item_labels = set(item.label for item in items)
            return item_labels.issubset(seen)

        n_labels = len(items)

        # Criterion A: last label_count * 2 trials are all correct.
        perfect_window = max(n_labels * 2, 1)
        if len(self.session_trial_rows) >= perfect_window:
            recent = self.session_trial_rows[-perfect_window:]
            if all(str(r.get("correct", "0")) in ["1", "True", "true"] for r in recent):
                return True

        # Criterion B: last label_count * 4 trials have accuracy >= 90%.
        acc_window = max(n_labels * 4, 1)
        if len(self.session_trial_rows) >= acc_window:
            recent_acc = self.session_trial_rows[-acc_window:]
            correct_n = sum(
                1 for r in recent_acc
                if str(r.get("correct", "0")) in ["1", "True", "true"]
            )
            acc = correct_n / max(len(recent_acc), 1)
            if acc >= ALT_ACC_THRESHOLD:
                return True

        return False


    def check_session_failed(self):
        counts = Counter([r.get("target_label", "") for r in self.session_trial_rows])
        return any(counts.get(item.label, 0) >= MAX_TRIALS_PER_ITEM for item in self.current_items())

    def advance_session(self):
        """Ask before moving to the next session."""
        if self.current_session_idx + 1 >= len(SESSIONS):
            QMessageBox.information(self, "Finished", "All sessions completed.")
            self.stim_button.setEnabled(False)
            self.feedback_label.setText("All sessions completed.")
            return

        next_idx = self.current_session_idx + 1
        reply = QMessageBox.question(
            self,
            "Next session?",
            f"Current session completed.\n\nProceed to next session?\n\n{next_idx + 1}. {SESSIONS[next_idx]['title']}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )

        if reply != QMessageBox.Yes:
            self.feedback_label.setText("Session completed. Press Start / Resume when ready to continue.")
            self.stim_button.setEnabled(False)
            return

        self.current_session_idx = next_idx
        self.session_trial_rows = self.rows_for_session_idx(self.current_session_idx)
        self.update_ui_for_session()
        self.next_trial()


    def save_state(self, completed_all=False):
        if not self.current_state_path: return
        state = {"subject": self.subject, "day": self.day, "current_session_idx": self.current_session_idx, "current_trial_index": self.current_trial_index, "current_csv_path": str(self.current_csv_path), "all_rows": self.all_rows, "completed_all": completed_all}
        with open(self.current_state_path, "w", encoding="utf-8") as f: json.dump(state, f, ensure_ascii=False, indent=2)

    def save_csv(self):
        if not self.current_csv_path: return
        fields = []
        preferred = ["subject","day","session_idx","session_name","session_title","trial_index_global","trial_index_session","target_label","target_command","response_mode","option_count","options","selected_label","manual_response","correct","rt_sec","rt_from_stimulus_sec","voice_onset_rt_sec","trial_elapsed_sec","session_elapsed_sec","day_elapsed_sec","trial_start_time","stimulus_sent_time","answer_time","voice_engine","voice_wav_path","voice_stt_raw","voice_stt_norm","voice_process_time_sec","voice_recorded_duration_sec","top_candidates","timestamp"]
        for k in preferred: fields.append(k)
        for r in self.all_rows:
            for k in r.keys():
                if k not in fields: fields.append(k)
        with open(self.current_csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore"); w.writeheader(); w.writerows(self.all_rows)

    def save_xlsx(self):
        try:
            import pandas as pd
            if self.current_csv_path and self.all_rows:
                pd.DataFrame(self.all_rows).to_excel(self.current_csv_path.with_suffix(".xlsx"), index=False)
        except Exception as e: self.log_msg(f"XLSX save failed: {e}")

def main():
    app = QApplication(sys.argv)
    win = HangulLearningVoiceApp(); win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
