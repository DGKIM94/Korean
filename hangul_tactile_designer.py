# -*- coding: utf-8 -*-
"""
Hangul Tactile Designer
==========================
A highly configurable Hangul tactile-code designer and tester.

Core functions
- Map each consonant/vowel to an atomic tactile stimulus, a composite of other
  jamo, or a raw serial script.
- Configure every jamo/syllable transition as end-to-start ISI.
- Configure syllable and word boundaries with the same ISI meaning.
- Preserve the current Arduino dot syntax and raw /i, /d motion commands.
- Save/load a complete setup as JSON.
- Type a Hangul string such as "각" and immediately feel the compiled stimulus.
- Run voice-response quizzes using the legacy v20 voice backend and a syllable
  XLSX file.
- Light Apple-inspired PySide6 UI with rounded frosted cards, gray controls, and clear selection states.

Hardware defaults for the current dual-arm rig
- LEFT arm  = device 2 = '#'
- RIGHT arm = device 1 = '@'
- Logical UI positions 1..9 are mapped to physical motors:
    1->3, 2->2, 3->1, 4->6, 5->5, 6->4, 7->9, 8->8, 9->7
- Serial baud: 115200

The app can run in DRY RUN mode without Arduino connection.
"""

from __future__ import annotations

import copy
import csv
import hashlib
import importlib.util
import json
import os
import random
import re
import subprocess
import sys
import time
import traceback
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    from PySide6.QtCore import Qt, QTimer, QSize
    from PySide6.QtGui import QColor, QFont, QIcon
    from PySide6.QtWidgets import (
        QApplication,
        QAbstractItemView,
        QCheckBox,
        QComboBox,
        QDialog,
        QFileDialog,
        QFrame,
        QGraphicsDropShadowEffect,
        QGridLayout,
        QHBoxLayout,
        QHeaderView,
        QLabel,
        QLineEdit,
        QListWidget,
        QListWidgetItem,
        QMainWindow,
        QMessageBox,
        QPlainTextEdit,
        QPushButton,
        QScrollArea,
        QSizePolicy,
        QSpinBox,
        QStackedWidget,
        QTableWidget,
        QTableWidgetItem,
        QTabWidget,
        QVBoxLayout,
        QWidget,
    )
except Exception as exc:  # pragma: no cover
    print("PySide6가 필요합니다.  python -m pip install PySide6", file=sys.stderr)
    raise

try:
    import serial  # type: ignore
    import serial.tools.list_ports  # type: ignore
except Exception:
    serial = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

APP_VERSION = "Hangul Tactile Designer portable v17"
# APP_DIR is always the user-visible program folder. RESOURCE_DIR is where
# bundled read-only files live when built with PyInstaller.
if getattr(sys, "frozen", False):
    APP_DIR = Path(sys.executable).resolve().parent
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", APP_DIR))
else:
    APP_DIR = Path(__file__).resolve().parent
    RESOURCE_DIR = APP_DIR
DATA_DIR = APP_DIR / "hangul_designer_results"
SETUP_DIR = APP_DIR / "hangul_tactile_setups"
DATA_DIR.mkdir(exist_ok=True)
SETUP_DIR.mkdir(exist_ok=True)

UI_ASSET_DIR = APP_DIR / "hangul_ui_assets"
UI_ASSET_DIR.mkdir(exist_ok=True)

def _write_ui_asset(name: str, content: str) -> Path:
    path = UI_ASSET_DIR / name
    try:
        if not path.exists() or path.read_text(encoding="utf-8") != content:
            path.write_text(content, encoding="utf-8")
    except Exception:
        pass
    return path

ARROW_UP_SVG = _write_ui_asset(
    "chevron_up.svg",
    """<svg xmlns='http://www.w3.org/2000/svg' width='20' height='20' viewBox='0 0 20 20'>
    <path d='M4.5 12.25L10 6.75l5.5 5.5' fill='none' stroke='#6E6E73' stroke-width='2.15' stroke-linecap='round' stroke-linejoin='round'/>
    </svg>""",
)
ARROW_DOWN_SVG = _write_ui_asset(
    "chevron_down.svg",
    """<svg xmlns='http://www.w3.org/2000/svg' width='20' height='20' viewBox='0 0 20 20'>
    <path d='M4.5 7.75L10 13.25l5.5-5.5' fill='none' stroke='#6E6E73' stroke-width='2.15' stroke-linecap='round' stroke-linejoin='round'/>
    </svg>""",
)
TOGGLE_ON_SVG = _write_ui_asset(
    "toggle_on.svg",
    """<svg xmlns='http://www.w3.org/2000/svg' width='56' height='34' viewBox='0 0 56 34'>
    <rect x='1' y='1' width='54' height='32' rx='16' fill='#34C759'/>
    <circle cx='39' cy='17' r='13' fill='#FFFFFF'/>
    <circle cx='39' cy='17' r='12.2' fill='#FFFFFF' stroke='#E5E5EA' stroke-width='.7'/>
    </svg>""",
)
TOGGLE_OFF_SVG = _write_ui_asset(
    "toggle_off.svg",
    """<svg xmlns='http://www.w3.org/2000/svg' width='56' height='34' viewBox='0 0 56 34'>
    <rect x='1' y='1' width='54' height='32' rx='16' fill='#E5E5EA'/>
    <circle cx='17' cy='17' r='13' fill='#FFFFFF'/>
    <circle cx='17' cy='17' r='12.2' fill='#FFFFFF' stroke='#D1D1D6' stroke-width='.7'/>
    </svg>""",
)

CONSONANTS = list("ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ")
VOWELS = list("ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ")

# General text-answer quiz groups.  The vowel groups cover the complete
# VOWELS list without overlap.  CVC uses only single final consonants because
# the requested structure is consonant + vowel + consonant, not a compound
# final such as ㄳ or ㄺ.
BASIC_CONSONANTS = list("ㄱㄴㄷㄹㅁㅂㅅㅇㅈㅊㅋㅌㅍㅎ")
DOUBLE_CONSONANTS = list("ㄲㄸㅃㅆㅉ")
BASIC_VOWELS = list("ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅛㅜㅠㅡㅣ")
DIPHTHONG_VOWELS = list("ㅘㅙㅚㅝㅞㅟㅢ")
SINGLE_FINAL_CONSONANTS = list("ㄱㄲㄴㄷㄹㅁㅂㅅㅆㅇㅈㅊㅋㅌㅍㅎ")
COMPOUND_FINALS: Dict[str, List[str]] = {
    "ㄳ": ["ㄱ", "ㅅ"], "ㄵ": ["ㄴ", "ㅈ"], "ㄶ": ["ㄴ", "ㅎ"],
    "ㄺ": ["ㄹ", "ㄱ"], "ㄻ": ["ㄹ", "ㅁ"], "ㄼ": ["ㄹ", "ㅂ"],
    "ㄽ": ["ㄹ", "ㅅ"], "ㄾ": ["ㄹ", "ㅌ"], "ㄿ": ["ㄹ", "ㅍ"],
    "ㅀ": ["ㄹ", "ㅎ"], "ㅄ": ["ㅂ", "ㅅ"],
}

CHOSUNG = CONSONANTS
JUNGSUNG = VOWELS
JONGSUNG = [""] + list("ㄱㄲㄳㄴㄵㄶㄷㄹㄺㄻㄼㄽㄾㄿㅀㅁㅂㅄㅅㅆㅇㅈㅊㅋㅌㅍㅎ")


def compose_hangul_syllable(cho: str, jung: str, jong: str = "") -> str:
    """Compose one modern Hangul syllable from compatibility jamo."""
    try:
        cho_index = CHOSUNG.index(str(cho))
        jung_index = JUNGSUNG.index(str(jung))
        jong_index = JONGSUNG.index(str(jong))
    except ValueError as exc:
        raise ValueError(f"조합할 수 없는 자모입니다: {cho}{jung}{jong}") from exc
    return chr(0xAC00 + (cho_index * 21 + jung_index) * 28 + jong_index)


TEMPORAL_PRESETS = {
    "S200": {"label": "짧음 · 200 ms", "on1_ms": 200, "gap_ms": 0, "on2_ms": 0},
    "L400": {"label": "김 · 400 ms", "on1_ms": 400, "gap_ms": 0, "on2_ms": 0},
    "R200_100_200": {"label": "연속 · 200–100–200", "on1_ms": 200, "gap_ms": 100, "on2_ms": 200},
    "CUSTOM": {"label": "사용자 설정", "on1_ms": 200, "gap_ms": 100, "on2_ms": 200},
}

DEFAULT_LOGICAL_TO_MOTOR = {
    "1": 3, "2": 2, "3": 1,
    "4": 6, "5": 5, "6": 4,
    "7": 9, "8": 8, "9": 7,
}

# ----------------------------- data model -----------------------------

# Matrix classes. Motion is split automatically using its actual compiled
# duration so short and long motion can use different ISI values.
TIMING_CLASS_LABELS: Dict[str, str] = {
    "short": "짧음",
    "long": "김",
    "repeat": "연속/반복",
    "motion_short": "짧은 모션",
    "motion_long": "긴 모션",
    "composite": "Composite",
    "custom": "사용자/기타",
}

# The mapping editor also offers a generic motion override. Its short/long
# matrix class is still selected automatically from the compiled duration.
TIMING_CLASS_EDITOR_LABELS: Dict[str, str] = {
    "short": "짧음",
    "long": "김",
    "repeat": "연속/반복",
    "motion": "모션 · 길이 자동",
    "motion_short": "짧은 모션 · 고정",
    "motion_long": "긴 모션 · 고정",
    "composite": "Composite",
    "custom": "사용자/기타",
}

TIMING_CONTEXT_LABELS: Dict[str, str] = {
    "composite": "Composite 내부",
    "cv": "CV · 자음→모음",
    "cvc_cv": "CVC · 첫 자음→모음",
    "cvc_vc": "CVC · 모음→끝 자음",
    "compound_final": "복합 종성 내부",
}

BOUNDARY_ISI_DEFAULTS_MS: Dict[str, int] = {
    "inter_syllable": 350,
    "inter_word": 650,
}

# Used only while migrating old onset-to-onset SOA setup files to the new
# end-to-start ISI model. Runtime timing always uses the actual compiled duration.
NOMINAL_TIMING_CLASS_DURATION_MS: Dict[str, int] = {
    "short": 200,
    "long": 400,
    "repeat": 500,
    "motion": 400,          # legacy alias
    "motion_short": 300,
    "motion_long": 550,
    "composite": 500,
    "custom": 300,
}

LEGACY_ISI_DEFAULTS_MS: Dict[str, int] = {
    "composite": 150,
    "cv": 250,
    "cvc_cv": 250,
    "cvc_vc": 250,
    "compound_final": 150,
}


def _default_timing_defaults_ms() -> Dict[str, int]:
    # Every value is an end-to-start ISI: wait this long after the complete
    # previous tactile stimulus has ended before starting the next stimulus.
    return {ctx: int(isi) for ctx, isi in LEGACY_ISI_DEFAULTS_MS.items()}


def _deep_copy_timing_matrix(src: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, int]]]:
    out: Dict[str, Dict[str, Dict[str, int]]] = {}
    for ctx, rows in (src or {}).items():
        if not isinstance(rows, dict):
            continue
        out[str(ctx)] = {}
        for prev_cls, cols in rows.items():
            if not isinstance(cols, dict):
                continue
            out[str(ctx)][str(prev_cls)] = {
                str(next_cls): int(value) for next_cls, value in cols.items()
            }
    return out


def _expand_motion_duration_matrix(
    src: Dict[str, Dict[str, Dict[str, int]]],
    defaults: Dict[str, int],
) -> Dict[str, Dict[str, Dict[str, int]]]:
    """Upgrade the old single `motion` row/column to short/long motion.

    All matrix values are end-to-start ISIs. Old setup files remain usable:
    both new motion classes inherit the old motion cell until edited.
    """
    copied = _deep_copy_timing_matrix(src)
    out: Dict[str, Dict[str, Dict[str, int]]] = {}
    for context in TIMING_CONTEXT_LABELS:
        old_rows = copied.get(context, {})
        default = int(defaults.get(context, 0))
        rows: Dict[str, Dict[str, int]] = {}
        for prev_cls in TIMING_CLASS_LABELS:
            prev_candidates = [prev_cls]
            if prev_cls in ("motion_short", "motion_long"):
                prev_candidates.append("motion")
            source_cols: Dict[str, int] = {}
            for candidate in prev_candidates:
                if isinstance(old_rows.get(candidate), dict):
                    source_cols = old_rows[candidate]
                    break
            cols: Dict[str, int] = {}
            for next_cls in TIMING_CLASS_LABELS:
                next_candidates = [next_cls]
                if next_cls in ("motion_short", "motion_long"):
                    next_candidates.append("motion")
                value = None
                for candidate in next_candidates:
                    if candidate in source_cols:
                        value = int(source_cols[candidate])
                        break
                cols[next_cls] = default if value is None else value
            rows[prev_cls] = cols
        out[context] = rows
    return out


@dataclass
class JamoSpec:
    mode: str = "atomic"  # atomic | composite | raw | unmapped
    arm: str = "left"
    position: int = 1
    temporal: str = "S200"
    on1_ms: int = 200
    gap_ms: int = 100
    on2_ms: int = 200
    steps: List[Dict[str, Any]] = field(default_factory=list)
    raw_command: str = ""
    note: str = ""
    # auto | short | long | repeat | motion | composite | custom
    timing_class: str = "auto"



@dataclass
class DesignSetup:
    setup_name: str = "Default Hangul Design"
    setup_version: int = 1
    created_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    updated_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    baudrate: int = 115200
    left_marker: str = "#"
    right_marker: str = "@"
    logical_to_motor: Dict[str, int] = field(default_factory=lambda: dict(DEFAULT_LOGICAL_TO_MOTOR))

    # Legacy field names are retained only so old JSON files remain readable.
    # In timing semantics v6 every internal transition is end-to-start ISI.
    default_composite_gap_ms: int = 150
    cv_gap_ms: int = 250
    cvc_cv_gap_ms: int = 250
    cvc_vc_gap_ms: int = 250
    compound_final_gap_ms: int = 150
    inter_syllable_gap_ms: int = 350
    inter_word_gap_ms: int = 650

    inter_syllable_isi_ms: int = 350
    inter_word_isi_ms: int = 650

    # Duration thresholds only select which ISI row is used. They never alter
    # the meaning of the value: every selected value is still end-to-start ISI.
    cv_duration_split_ms: int = 300
    motion_duration_split_ms: int = 300
    cv_short_isi_ms: int = 250
    cv_long_isi_ms: int = 250
    cvc_cv_short_isi_ms: int = 250
    cvc_cv_long_isi_ms: int = 250
    use_duration_based_cv_isi: bool = True

    timing_semantics_version: int = 6
    # The generic names are kept for file compatibility. In v6 these dictionaries
    # store ISI values, not onset-to-onset SOA values.
    timing_defaults_ms: Dict[str, int] = field(default_factory=_default_timing_defaults_ms)
    timing_matrix_ms: Dict[str, Dict[str, Dict[str, int]]] = field(default_factory=dict)
    timing_pair_overrides_ms: Dict[str, int] = field(default_factory=dict)
    jamo: Dict[str, Dict[str, Any]] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        self.sync_duration_cv_fields_from_matrix()
        out = asdict(self)
        out["updated_at"] = datetime.now().isoformat(timespec="seconds")

        # Keep the familiar gap names synchronized for older analysis scripts.
        out["default_composite_gap_ms"] = int(self.timing_defaults_ms.get("composite", 0))
        out["cv_gap_ms"] = int(self.timing_defaults_ms.get("cv", 0))
        out["cvc_cv_gap_ms"] = int(self.timing_defaults_ms.get("cvc_cv", 0))
        out["cvc_vc_gap_ms"] = int(self.timing_defaults_ms.get("cvc_vc", 0))
        out["compound_final_gap_ms"] = int(self.timing_defaults_ms.get("compound_final", 0))
        out["inter_syllable_gap_ms"] = int(self.inter_syllable_isi_ms)
        out["inter_word_gap_ms"] = int(self.inter_word_isi_ms)

        out["timing_defaults_ms"] = {
            k: int(v) for k, v in self.timing_defaults_ms.items()
            if k in TIMING_CONTEXT_LABELS
        }
        out["timing_matrix_ms"] = {
            k: v for k, v in out.get("timing_matrix_ms", {}).items()
            if k in TIMING_CONTEXT_LABELS
        }
        out["timing_pair_overrides_ms"] = {
            k: int(v) for k, v in out.get("timing_pair_overrides_ms", {}).items()
            if k.split("|", 1)[0] in TIMING_CONTEXT_LABELS
        }
        return out

    @staticmethod
    def _legacy_step_duration_ms(step: str) -> int:
        durations = [int(x) for x in re.findall(r"/(\d+)", str(step))]
        low = str(step).lower()
        if re.search(r"/[s](?:@|#|$)", low):
            durations.append(200)
        if re.search(r"/[l](?:@|#|$)", low):
            durations.append(400)
        if re.search(r"/[r](?:@|#|$)", low):
            durations.append(500)
        return max(durations) if durations else 0

    @classmethod
    def _legacy_raw_duration_ms(cls, command: str) -> int:
        current_onset = 0
        final_end = 0
        for original_step in str(command).split("."):
            step = original_step.strip()
            if not step:
                continue
            if re.fullmatch(r"0/(\d+)", step):
                current_onset += int(step.split("/", 1)[1])
                final_end = max(final_end, current_onset)
                continue
            duration = cls._legacy_step_duration_ms(step)
            final_end = max(final_end, current_onset + duration)
            current_onset += duration
        return int(final_end)

    @classmethod
    def _legacy_jamo_duration_ms(
        cls,
        label: str,
        jamo_data: Dict[str, Dict[str, Any]],
        previous_version: int,
        memo: Optional[Dict[str, int]] = None,
        stack: Optional[List[str]] = None,
    ) -> int:
        memo = memo if memo is not None else {}
        stack = list(stack or [])
        if label in memo:
            return int(memo[label])
        if label in stack:
            return NOMINAL_TIMING_CLASS_DURATION_MS["composite"]
        if label in COMPOUND_FINALS and label not in jamo_data:
            total = 0
            for index, component in enumerate(COMPOUND_FINALS[label]):
                duration = cls._legacy_jamo_duration_ms(
                    component, jamo_data, previous_version, memo, stack + [label]
                )
                if index > 0:
                    total += LEGACY_ISI_DEFAULTS_MS["compound_final"]
                total += duration
            memo[label] = int(total)
            return int(total)

        raw = jamo_data.get(label, {"mode": "unmapped"})
        if not isinstance(raw, dict):
            return 0
        mode = str(raw.get("mode", "unmapped"))
        if mode == "atomic":
            temporal = str(raw.get("temporal", "S200"))
            preset = TEMPORAL_PRESETS.get(temporal, TEMPORAL_PRESETS["CUSTOM"])
            if temporal == "CUSTOM":
                on1 = int(raw.get("on1_ms", 200))
                gap = int(raw.get("gap_ms", 100))
                on2 = int(raw.get("on2_ms", 200))
            else:
                on1 = int(preset["on1_ms"])
                gap = int(preset["gap_ms"])
                on2 = int(preset["on2_ms"])
            duration = on1 + (max(0, gap) + on2 if on2 > 0 else 0)
        elif mode == "raw":
            duration = cls._legacy_raw_duration_ms(str(raw.get("raw_command", "")))
        elif mode == "composite":
            onset = 0
            final_end = 0
            previous_ref = ""
            steps = raw.get("steps") or []
            for index, step in enumerate(steps):
                if not isinstance(step, dict):
                    continue
                ref = str(step.get("ref", "")).strip()
                ref_duration = cls._legacy_jamo_duration_ms(
                    ref, jamo_data, previous_version, memo, stack + [label]
                )
                if index > 0:
                    if "isi_before_ms" in step:
                        onset += cls._legacy_jamo_duration_ms(
                            previous_ref, jamo_data, previous_version, memo, stack + [label]
                        ) + max(0, int(step.get("isi_before_ms", 0)))
                    elif previous_version < 2 and "gap_before_ms" in step:
                        onset += cls._legacy_jamo_duration_ms(
                            previous_ref, jamo_data, previous_version, memo, stack + [label]
                        ) + max(0, int(step.get("gap_before_ms", 0)))
                    elif "legacy_isi_ms" in step:
                        onset += cls._legacy_jamo_duration_ms(
                            previous_ref, jamo_data, previous_version, memo, stack + [label]
                        ) + max(0, int(step.get("legacy_isi_ms", 0)))
                    else:
                        onset += max(0, int(step.get(
                            "soa_before_ms",
                            step.get("gap_before_ms", 0),
                        )))
                final_end = max(final_end, onset + ref_duration)
                previous_ref = ref
            duration = final_end
        else:
            duration = 0
        memo[label] = max(0, int(duration))
        return memo[label]

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "DesignSetup":
        original = dict(data or {})
        previous_version = int(original.get("timing_semantics_version", 1) or 1)

        allowed = {f.name for f in cls.__dataclass_fields__.values()}
        clean = {k: v for k, v in original.items() if k in allowed}
        obj = cls(**clean)
        obj.logical_to_motor = {str(k): int(v) for k, v in obj.logical_to_motor.items()}
        obj.jamo = copy.deepcopy(obj.jamo or {})

        # Boundary fields were already true ISI in v4/v5. Earlier v2/v3 files
        # stored boundary onset-to-onset values and need the old nominal 200 ms
        # subtraction to recover an approximate physical gap.
        if "inter_syllable_isi_ms" in original:
            obj.inter_syllable_isi_ms = max(0, int(original["inter_syllable_isi_ms"]))
        elif previous_version < 2:
            obj.inter_syllable_isi_ms = max(0, int(original.get("inter_syllable_gap_ms", 350)))
        else:
            old_boundary = int((original.get("timing_defaults_ms") or {}).get(
                "inter_syllable", original.get("inter_syllable_gap_ms", 550)
            ))
            obj.inter_syllable_isi_ms = max(0, old_boundary - 200)

        if "inter_word_isi_ms" in original:
            obj.inter_word_isi_ms = max(0, int(original["inter_word_isi_ms"]))
        elif previous_version < 2:
            obj.inter_word_isi_ms = max(0, int(original.get("inter_word_gap_ms", 650)))
        else:
            old_boundary = int((original.get("timing_defaults_ms") or {}).get(
                "inter_word", original.get("inter_word_gap_ms", 850)
            ))
            obj.inter_word_isi_ms = max(0, old_boundary - 200)

        if previous_version >= 6:
            obj.timing_defaults_ms = {
                **_default_timing_defaults_ms(),
                **{str(k): max(0, int(v)) for k, v in (obj.timing_defaults_ms or {}).items()},
            }
            obj.timing_matrix_ms = _expand_motion_duration_matrix(
                _deep_copy_timing_matrix(obj.timing_matrix_ms),
                obj.timing_defaults_ms,
            )
            obj.timing_pair_overrides_ms = {
                str(k): max(0, int(v))
                for k, v in (obj.timing_pair_overrides_ms or {}).items()
                if str(k).split("|", 1)[0] in TIMING_CONTEXT_LABELS
            }
            obj._normalize_composite_step_keys_to_isi(previous_version)
            obj.sync_duration_cv_fields_from_matrix()
            obj.timing_semantics_version = 6
            obj.use_duration_based_cv_isi = True
            return obj

        legacy_jamo = copy.deepcopy(obj.jamo)
        legacy_defaults = {
            "composite": int(original.get("default_composite_gap_ms", 150)),
            "cv": int(original.get("cv_gap_ms", 250)),
            "cvc_cv": int(original.get("cvc_cv_gap_ms", 250)),
            "cvc_vc": int(original.get("cvc_vc_gap_ms", 250)),
            "compound_final": int(original.get("compound_final_gap_ms", 150)),
        }

        if previous_version < 2:
            # The oldest setup format already used end-to-start gaps.
            obj.timing_defaults_ms = {
                context: max(0, int(legacy_defaults[context]))
                for context in TIMING_CONTEXT_LABELS
            }
            obj.timing_matrix_ms = {
                context: {
                    prev_cls: {
                        next_cls: int(obj.timing_defaults_ms[context])
                        for next_cls in TIMING_CLASS_LABELS
                    }
                    for prev_cls in TIMING_CLASS_LABELS
                }
                for context in TIMING_CONTEXT_LABELS
            }
            obj.timing_pair_overrides_ms = {
                str(k): max(0, int(v))
                for k, v in (original.get("timing_pair_overrides_ms") or {}).items()
            }
        else:
            # v2-v5 stored onset-to-onset SOA. Switching to ISI is also a
            # simplification step: each non-CV context is collapsed to one clear
            # end-to-start gap based on the old short-row/default value. This
            # avoids carrying the old class-dependent overlap complexity into
            # the new model.
            old_defaults_raw = {
                **{
                    context: NOMINAL_TIMING_CLASS_DURATION_MS["short"]
                    + LEGACY_ISI_DEFAULTS_MS[context]
                    for context in TIMING_CONTEXT_LABELS
                },
                **{
                    str(k): int(v)
                    for k, v in (original.get("timing_defaults_ms") or {}).items()
                    if str(k) in TIMING_CONTEXT_LABELS
                },
            }
            converted_defaults = {
                context: max(0, int(old_defaults_raw[context]) - 200)
                for context in TIMING_CONTEXT_LABELS
            }
            converted_matrix: Dict[str, Dict[str, Dict[str, int]]] = {
                context: {
                    prev_cls: {
                        next_cls: int(converted_defaults[context])
                        for next_cls in TIMING_CLASS_LABELS
                    }
                    for prev_cls in TIMING_CLASS_LABELS
                }
                for context in TIMING_CONTEXT_LABELS
            }

            # The four visible CV/CVC controls were the authoritative values in
            # older versions. Preserve their approximate physical gaps.
            old_cv_short = int(original.get(
                "cv_short_soa_ms", old_defaults_raw.get("cv", 450)
            ))
            old_cv_long = int(original.get("cv_long_soa_ms", old_cv_short))
            old_cvc_short = int(original.get(
                "cvc_cv_short_soa_ms", old_defaults_raw.get("cvc_cv", 450)
            ))
            old_cvc_long = int(original.get("cvc_cv_long_soa_ms", old_cvc_short))
            converted_rows = {
                ("cv", "short"): max(0, old_cv_short - 200),
                ("cv", "long"): max(0, old_cv_long - 400),
                ("cvc_cv", "short"): max(0, old_cvc_short - 200),
                ("cvc_cv", "long"): max(0, old_cvc_long - 400),
            }
            for (context, prev_cls), value in converted_rows.items():
                for next_cls in TIMING_CLASS_LABELS:
                    converted_matrix[context][prev_cls][next_cls] = int(value)

            obj.timing_matrix_ms = converted_matrix
            obj.timing_defaults_ms = {
                context: int(converted_defaults[context])
                for context in TIMING_CONTEXT_LABELS
            }
            obj.timing_defaults_ms["cv"] = int(converted_rows[("cv", "short")])
            obj.timing_defaults_ms["cvc_cv"] = int(converted_rows[("cvc_cv", "short")])

            memo: Dict[str, int] = {}
            converted_overrides: Dict[str, int] = {}
            for key, value in (original.get("timing_pair_overrides_ms") or {}).items():
                parts = str(key).split("|", 2)
                if len(parts) != 3 or parts[0] not in TIMING_CONTEXT_LABELS:
                    continue
                previous_duration = cls._legacy_jamo_duration_ms(
                    parts[1], legacy_jamo, previous_version, memo
                )
                if previous_duration <= 0:
                    previous_class = obj.timing_class_for(parts[1])
                    previous_duration = int(
                        NOMINAL_TIMING_CLASS_DURATION_MS.get(previous_class, 300)
                    )
                converted_overrides[str(key)] = max(0, int(value) - previous_duration)
            obj.timing_pair_overrides_ms = converted_overrides

        obj._normalize_composite_step_keys_to_isi(
            previous_version,
            legacy_jamo=legacy_jamo,
        )
        obj.timing_matrix_ms = _expand_motion_duration_matrix(
            obj.timing_matrix_ms, obj.timing_defaults_ms
        )
        obj.sync_duration_cv_fields_from_matrix()
        obj.timing_semantics_version = 6
        obj.use_duration_based_cv_isi = True
        return obj

    def ensure_timing_matrix(self) -> None:
        self.timing_matrix_ms = _expand_motion_duration_matrix(
            self.timing_matrix_ms, self.timing_defaults_ms
        )

    def set_uniform_timing_row(self, context: str, prev_cls: str, value: int) -> None:
        """Set one complete previous-stimulus row to a single ISI value."""
        self.ensure_timing_matrix()
        if context not in TIMING_CONTEXT_LABELS or prev_cls not in TIMING_CLASS_LABELS:
            return
        row = self.timing_matrix_ms[context].setdefault(prev_cls, {})
        for next_cls in TIMING_CLASS_LABELS:
            row[next_cls] = max(0, int(value))

    def uniform_timing_row_value(self, context: str, prev_cls: str, fallback: int) -> int:
        self.ensure_timing_matrix()
        row = self.timing_matrix_ms.get(context, {}).get(prev_cls, {})
        for next_cls in TIMING_CLASS_LABELS:
            if next_cls in row:
                return max(0, int(row[next_cls]))
        return max(0, int(fallback))

    def sync_duration_cv_rows_from_fields(self) -> None:
        self.set_uniform_timing_row("cv", "short", self.cv_short_isi_ms)
        self.set_uniform_timing_row("cv", "long", self.cv_long_isi_ms)
        self.set_uniform_timing_row("cvc_cv", "short", self.cvc_cv_short_isi_ms)
        self.set_uniform_timing_row("cvc_cv", "long", self.cvc_cv_long_isi_ms)
        self.timing_defaults_ms["cv"] = max(0, int(self.cv_short_isi_ms))
        self.timing_defaults_ms["cvc_cv"] = max(0, int(self.cvc_cv_short_isi_ms))

    def sync_duration_cv_fields_from_matrix(self) -> None:
        self.cv_short_isi_ms = self.uniform_timing_row_value(
            "cv", "short", self.cv_short_isi_ms
        )
        self.cv_long_isi_ms = self.uniform_timing_row_value(
            "cv", "long", self.cv_long_isi_ms
        )
        self.cvc_cv_short_isi_ms = self.uniform_timing_row_value(
            "cvc_cv", "short", self.cvc_cv_short_isi_ms
        )
        self.cvc_cv_long_isi_ms = self.uniform_timing_row_value(
            "cvc_cv", "long", self.cvc_cv_long_isi_ms
        )
        self.timing_defaults_ms["cv"] = int(self.cv_short_isi_ms)
        self.timing_defaults_ms["cvc_cv"] = int(self.cvc_cv_short_isi_ms)

    def _normalize_composite_step_keys_to_isi(
        self,
        previous_version: int,
        legacy_jamo: Optional[Dict[str, Dict[str, Any]]] = None,
    ) -> None:
        legacy_jamo = copy.deepcopy(legacy_jamo if legacy_jamo is not None else self.jamo)
        memo: Dict[str, int] = {}
        for owner_label, raw_spec in self.jamo.items():
            if not isinstance(raw_spec, dict):
                continue
            steps = raw_spec.get("steps") or []
            previous_ref = ""
            for index, step in enumerate(steps):
                if not isinstance(step, dict):
                    continue
                if index == 0:
                    isi = 0
                elif "isi_before_ms" in step:
                    isi = max(0, int(step.get("isi_before_ms", 0)))
                elif previous_version < 2 and "gap_before_ms" in step:
                    isi = max(0, int(step.get("gap_before_ms", 0)))
                elif "legacy_isi_ms" in step:
                    isi = max(0, int(step.get("legacy_isi_ms", 0)))
                else:
                    old_soa = int(step.get(
                        "soa_before_ms",
                        step.get("gap_before_ms", self.timing_defaults_ms.get("composite", 0)),
                    ))
                    previous_duration = self._legacy_jamo_duration_ms(
                        previous_ref, legacy_jamo, previous_version, memo, [owner_label]
                    )
                    if previous_duration <= 0:
                        previous_class = self.timing_class_for(previous_ref)
                        previous_duration = int(
                            NOMINAL_TIMING_CLASS_DURATION_MS.get(previous_class, 300)
                        )
                    isi = max(0, old_soa - previous_duration)
                step["isi_before_ms"] = int(isi)
                step.pop("soa_before_ms", None)
                step.pop("gap_before_ms", None)
                step.pop("legacy_isi_ms", None)
                previous_ref = str(step.get("ref", "")).strip()

    def get_spec(self, label: str) -> JamoSpec:
        raw = self.jamo.get(label, {"mode": "unmapped"})
        allowed = {f.name for f in JamoSpec.__dataclass_fields__.values()}
        return JamoSpec(**{k: v for k, v in raw.items() if k in allowed})

    def set_spec(self, label: str, spec: JamoSpec) -> None:
        self.jamo[label] = asdict(spec)

    def marker_for_arm(self, arm: str) -> str:
        return self.left_marker if arm == "left" else self.right_marker

    def motor_for_position(self, position: int) -> int:
        return int(self.logical_to_motor.get(str(int(position)), int(position)))

    def timing_class_for(self, label: str, stack: Optional[List[str]] = None) -> str:
        stack = list(stack or [])
        if label in stack:
            return "composite"
        if label in COMPOUND_FINALS and label not in self.jamo:
            return "composite"
        spec = self.get_spec(label)
        if spec.timing_class in TIMING_CLASS_EDITOR_LABELS:
            return spec.timing_class
        if spec.mode == "composite":
            return "composite"
        if spec.mode == "raw":
            raw = spec.raw_command.lower()
            if "/i" in raw or "/d" in raw or "," in raw:
                return "motion"
            return "custom"
        if spec.mode == "atomic":
            if spec.temporal == "S200":
                return "short"
            if spec.temporal == "L400":
                return "long"
            if spec.temporal == "R200_100_200":
                return "repeat"
            if int(spec.on2_ms) > 0:
                return "repeat"
            return "short" if int(spec.on1_ms) <= 300 else "long"
        return "custom"

    def refined_timing_class(self, timing_class: str, duration_ms: Optional[int]) -> str:
        if timing_class == "motion":
            duration = int(duration_ms or 0)
            return (
                "motion_short"
                if duration <= int(self.motion_duration_split_ms)
                else "motion_long"
            )
        return timing_class if timing_class in TIMING_CLASS_LABELS else "custom"

    @staticmethod
    def pair_override_key(context: str, from_label: str, to_label: str) -> str:
        return f"{context}|{from_label}|{to_label}"

    def resolve_isi_ms(
        self,
        context: str,
        from_label: str,
        to_label: str,
        from_duration_ms: Optional[int] = None,
        to_duration_ms: Optional[int] = None,
    ) -> int:
        pair_key = self.pair_override_key(context, from_label, to_label)
        if pair_key in self.timing_pair_overrides_ms:
            return max(0, int(self.timing_pair_overrides_ms[pair_key]))

        prev_base = self.timing_class_for(from_label)
        next_base = self.timing_class_for(to_label)
        prev_cls = self.refined_timing_class(prev_base, from_duration_ms)
        next_cls = self.refined_timing_class(next_base, to_duration_ms)

        if (
            self.use_duration_based_cv_isi
            and context in ("cv", "cvc_cv")
            and from_duration_ms is not None
            and prev_base not in ("motion", "motion_short", "motion_long")
        ):
            prev_cls = (
                "short"
                if int(from_duration_ms) <= int(self.cv_duration_split_ms)
                else "long"
            )

        try:
            return max(0, int(self.timing_matrix_ms[context][prev_cls][next_cls]))
        except Exception:
            return max(0, int(self.timing_defaults_ms.get(context, 0)))

    def stable_hash(self) -> str:
        blob = json.dumps(self.to_dict(), ensure_ascii=False, sort_keys=True).encode("utf-8")
        return hashlib.sha256(blob).hexdigest()[:16]


class CompileError(RuntimeError):
    pass


@dataclass
class TimelineEvent:
    onset_ms: int
    command: str
    duration_ms: int
    label: str = ""

    @property
    def end_ms(self) -> int:
        return int(self.onset_ms) + int(self.duration_ms)


class HangulCompiler:
    """Compile Hangul into a tactile timeline for the current Arduino code.

    Raw ``/i`` and ``/d`` motion commands are preserved. Every generated
    transition uses end-to-start ISI, so the next jamo begins only after the
    complete previous tactile stimulus has ended and the configured ISI has
    elapsed. This remains fully representable by the unchanged Arduino dot
    syntax without overlap exceptions.
    """

    def __init__(self, setup: DesignSetup):
        self.setup = setup

    @staticmethod
    def join(parts: Sequence[str]) -> str:
        return ".".join(p.strip(".") for p in parts if str(p).strip("."))

    @staticmethod
    def _step_duration_ms(step: str) -> int:
        step = re.sub(r"\^(\d+)\s*$", "", str(step).strip())
        durations = [int(x) for x in re.findall(r"/(\d+)", step)]
        low = step.lower()
        if re.search(r"/[s](?:@|#|$)", low):
            durations.append(200)
        if re.search(r"/[l](?:@|#|$)", low):
            durations.append(400)
        if re.search(r"/[r](?:@|#|$)", low):
            durations.append(500)
        return max(durations) if durations else 0

    @staticmethod
    def _shift(events: Sequence[TimelineEvent], offset_ms: int) -> List[TimelineEvent]:
        return [
            TimelineEvent(int(e.onset_ms) + int(offset_ms), e.command, e.duration_ms, e.label)
            for e in events
        ]

    @staticmethod
    def timeline_duration_ms(events: Sequence[TimelineEvent]) -> int:
        return max((e.end_ms for e in events), default=0)

    def atomic_timeline(self, spec: JamoSpec, label: str = "") -> List[TimelineEvent]:
        marker = self.setup.marker_for_arm(spec.arm)
        motor = self.setup.motor_for_position(spec.position)
        preset = TEMPORAL_PRESETS.get(spec.temporal, TEMPORAL_PRESETS["CUSTOM"])
        if spec.temporal == "CUSTOM":
            on1, gap_ms, on2 = int(spec.on1_ms), int(spec.gap_ms), int(spec.on2_ms)
        else:
            on1, gap_ms, on2 = int(preset["on1_ms"]), int(preset["gap_ms"]), int(preset["on2_ms"])
        if on1 <= 0:
            raise CompileError("첫 번째 ON duration은 1 ms 이상이어야 합니다.")
        events = [TimelineEvent(0, f"{marker}{motor}/{on1}", on1, label)]
        if on2 > 0:
            onset2 = on1 + max(0, gap_ms)
            events.append(TimelineEvent(onset2, f"{marker}{motor}/{on2}", on2, label))
        return events

    def _raw_timeline(self, command: str, label: str = "") -> List[TimelineEvent]:
        command = str(command).strip().strip(".")
        if not command:
            raise CompileError(f"{label} raw command가 비어 있습니다.")
        events: List[TimelineEvent] = []
        current_onset = 0
        current_marker = ""
        for original_step in command.split("."):
            step = original_step.strip()
            if not step:
                continue
            explicit_advance: Optional[int] = None
            m = re.search(r"\^(\d+)\s*$", step)
            if m:
                raise CompileError(
                    "현재 Arduino 코드는 ^N onset 문법을 지원하지 않습니다. "
                    "Raw command에서 ^N을 제거하세요."
                )
            if re.fullmatch(r"0/(\d+)", step):
                rest_ms = int(step.split("/", 1)[1])
                current_onset += explicit_advance if explicit_advance is not None else rest_ms
                continue
            if step.startswith(("@", "#")):
                markers = re.findall(r"[@#]", step)
                if markers:
                    current_marker = markers[-1]
                normalized = step
            else:
                if not current_marker:
                    raise CompileError("Raw script는 @ 또는 # marker로 시작해야 합니다.")
                normalized = current_marker + step
            duration = self._step_duration_ms(normalized)
            events.append(TimelineEvent(current_onset, normalized, duration, label))
            current_onset += explicit_advance if explicit_advance is not None else duration
        return events

    def compile_jamo_timeline(self, label: str, stack: Optional[List[str]] = None) -> List[TimelineEvent]:
        stack = list(stack or [])
        if label in stack:
            raise CompileError("순환 참조: " + " → ".join(stack + [label]))
        if label in COMPOUND_FINALS and label not in self.setup.jamo:
            events: List[TimelineEvent] = []
            onset = 0
            previous = ""
            previous_duration: Optional[int] = None
            for index, comp in enumerate(COMPOUND_FINALS[label]):
                comp_events = self.compile_jamo_timeline(comp, stack + [label])
                comp_duration = self.timeline_duration_ms(comp_events)
                if index > 0:
                    isi = self.setup.resolve_isi_ms(
                        "compound_final", previous, comp,
                        from_duration_ms=previous_duration,
                        to_duration_ms=comp_duration,
                    )
                    onset += int(previous_duration or 0) + max(0, int(isi))
                events.extend(self._shift(comp_events, onset))
                previous = comp
                previous_duration = comp_duration
            return events

        spec = self.setup.get_spec(label)
        if spec.mode == "unmapped":
            raise CompileError(f"{label} 매핑이 없습니다.")
        if spec.mode == "atomic":
            return self.atomic_timeline(spec, label)
        if spec.mode == "raw":
            return self._raw_timeline(spec.raw_command, label)
        if spec.mode == "composite":
            if not spec.steps:
                raise CompileError(f"{label} composite step이 없습니다.")
            events: List[TimelineEvent] = []
            onset = 0
            previous_ref = ""
            previous_duration: Optional[int] = None
            for index, step in enumerate(spec.steps):
                ref = str(step.get("ref", "")).strip()
                if not ref:
                    raise CompileError(f"{label}의 {index + 1}번째 구성 자모가 비어 있습니다.")
                ref_events = self.compile_jamo_timeline(ref, stack + [label])
                ref_duration = self.timeline_duration_ms(ref_events)
                if index > 0:
                    if "isi_before_ms" in step:
                        isi = int(step.get("isi_before_ms", 0))
                    else:
                        isi = self.setup.resolve_isi_ms(
                            "composite", previous_ref, ref,
                            from_duration_ms=previous_duration,
                            to_duration_ms=ref_duration,
                        )
                    onset += int(previous_duration or 0) + max(0, int(isi))
                events.extend(self._shift(ref_events, onset))
                previous_ref = ref
                previous_duration = ref_duration
            return events
        raise CompileError(f"지원하지 않는 mode: {spec.mode}")

    def compile_syllable_timeline(self, ch: str) -> Tuple[List[TimelineEvent], str, str, int]:
        dec = self.decompose_syllable(ch)
        if dec is None:
            if ch in self.setup.jamo or ch in COMPOUND_FINALS:
                return self.compile_jamo_timeline(ch), ch, ch, 0
            raise CompileError(f"지원하지 않는 문자: {ch}")
        cho, jung, jong = dec
        cho_events = self.compile_jamo_timeline(cho)
        jung_events = self.compile_jamo_timeline(jung)
        cho_duration = self.timeline_duration_ms(cho_events)
        jung_duration = self.timeline_duration_ms(jung_events)
        jung_isi = self.setup.resolve_isi_ms(
            "cvc_cv" if jong else "cv", cho, jung,
            from_duration_ms=cho_duration,
            to_duration_ms=jung_duration,
        )
        jung_onset = cho_duration + max(0, int(jung_isi))
        events = list(cho_events)
        events.extend(self._shift(jung_events, jung_onset))
        last_label = jung
        last_jamo_onset = jung_onset
        if jong:
            jong_events = self.compile_jamo_timeline(jong)
            jong_duration = self.timeline_duration_ms(jong_events)
            jong_isi = self.setup.resolve_isi_ms(
                "cvc_vc", jung, jong,
                from_duration_ms=jung_duration,
                to_duration_ms=jong_duration,
            )
            jong_onset = jung_onset + jung_duration + max(0, int(jong_isi))
            events.extend(self._shift(jong_events, jong_onset))
            last_label = jong
            last_jamo_onset = jong_onset
        return events, cho, last_label, int(last_jamo_onset)

    @staticmethod
    def decompose_syllable(ch: str) -> Optional[Tuple[str, str, str]]:
        if not ch:
            return None
        code = ord(ch[0])
        if not (0xAC00 <= code <= 0xD7A3):
            return None
        sidx = code - 0xAC00
        return CHOSUNG[sidx // 588], JUNGSUNG[(sidx % 588) // 28], JONGSUNG[sidx % 28]

    @staticmethod
    def serialize_timeline(events: Sequence[TimelineEvent]) -> str:
        """Serialize using the user's unchanged Arduino dot-step syntax.

        Raw ramp tokens remain intact, for example::

            #5/150.#5/d,4/i/100.#4/150

        Empty end-to-start ISIs are emitted as ``0/N``. Equal-onset events are
        joined into one step. Generated ISI timelines are sequential, so an
        overlap here indicates a malformed raw/custom timeline rather than a
        normal timing-rule conflict.
        """
        clean = [e for e in events if int(e.duration_ms) > 0 and str(e.command).strip()]
        if not clean:
            return ""

        grouped: Dict[int, List[TimelineEvent]] = {}
        for event in clean:
            grouped.setdefault(int(event.onset_ms), []).append(event)

        parts: List[str] = []
        cursor = 0
        for onset in sorted(grouped):
            if onset < cursor:
                overlap = cursor - onset
                raise CompileError(
                    f"타임라인 내부에 {overlap} ms 중첩이 있습니다. "
                    "일반 ISI 규칙에서는 발생하지 않으므로 raw/custom command를 확인하세요."
                )
            if onset > cursor:
                parts.append(f"0/{onset - cursor}")
            group = grouped[onset]
            commands = [str(e.command).strip().strip(".") for e in group if str(e.command).strip(".")]
            if commands:
                parts.append("".join(commands))
            cursor = max(int(e.end_ms) for e in group)
        return ".".join(parts)

    def compile_jamo(self, label: str, stack: Optional[List[str]] = None) -> str:
        return self.serialize_timeline(self.compile_jamo_timeline(label, stack))

    def compile_syllable(self, ch: str) -> str:
        events, _first, _last, _last_onset = self.compile_syllable_timeline(ch)
        return self.serialize_timeline(events)

    def compile_text(self, text: str) -> Tuple[str, List[Dict[str, Any]]]:
        text = str(text)
        events: List[TimelineEvent] = []
        trace: List[Dict[str, Any]] = []
        previous_char_start = 0
        previous_end_ms = 0
        previous_last_label = ""
        have_previous = False
        pending_word_boundary = False

        for ch in text:
            if ch.isspace():
                if have_previous:
                    pending_word_boundary = True
                continue

            syllable_events, first_label, last_label, last_jamo_onset = self.compile_syllable_timeline(ch)
            if not have_previous:
                char_start = 0
            else:
                boundary_isi = (
                    self.setup.inter_word_isi_ms
                    if pending_word_boundary
                    else self.setup.inter_syllable_isi_ms
                )
                # Boundary timing is true ISI: wait after the complete previous
                # syllable has physically ended, then start the next syllable.
                char_start = int(previous_end_ms) + max(0, int(boundary_isi))
            shifted = self._shift(syllable_events, char_start)
            events.extend(shifted)
            relative_command = self.serialize_timeline(syllable_events)
            trace.append({
                "char": ch,
                "command": relative_command,
                "onset_ms": int(char_start),
                "first_label": first_label,
                "last_label": last_label,
            })
            previous_char_start = char_start
            previous_end_ms = max((e.end_ms for e in shifted), default=char_start)
            previous_last_label = last_label
            have_previous = True
            pending_word_boundary = False

        if not trace:
            raise CompileError("입력된 한글이 없습니다.")
        return self.serialize_timeline(events), trace

    @staticmethod
    def estimate_duration_ms(command: str) -> int:
        current_onset = 0
        final_end = 0
        for original_step in str(command).split("."):
            step = original_step.strip()
            if not step:
                continue
            explicit_advance: Optional[int] = None
            m = re.search(r"\^(\d+)\s*$", step)
            if m:
                explicit_advance = int(m.group(1))
                step = step[:m.start()].strip()
            if re.fullmatch(r"0/(\d+)", step):
                duration = int(step.split("/", 1)[1])
                current_onset += explicit_advance if explicit_advance is not None else duration
                final_end = max(final_end, current_onset)
                continue
            duration = HangulCompiler._step_duration_ms(step)
            final_end = max(final_end, current_onset + duration)
            current_onset += explicit_advance if explicit_advance is not None else duration
        return int(final_end)


# ----------------------------- default setup -----------------------------

def atomic(arm: str, pos: int, temporal: str, note: str = "") -> Dict[str, Any]:
    return asdict(JamoSpec(mode="atomic", arm=arm, position=pos, temporal=temporal, note=note))


def composite(refs: Sequence[str], gap: int = 150, note: str = "") -> Dict[str, Any]:
    """Create a composite with direct end-to-start ISI values."""
    steps = []
    for i, ref in enumerate(refs):
        steps.append({"ref": ref, "isi_before_ms": 0 if i == 0 else int(gap)})
    return asdict(JamoSpec(mode="composite", steps=steps, note=note))


def raw(command: str, note: str = "") -> Dict[str, Any]:
    return asdict(JamoSpec(mode="raw", raw_command=command, note=note))


def make_default_setup() -> DesignSetup:
    s = DesignSetup()
    # Preserve the earlier tactile-Hangul design using the current logical UI layout.
    s.jamo.update({
        "ㄱ": atomic("left", 1, "S200", "기본 ㄱ"),
        "ㄴ": atomic("left", 2, "S200", "기본 ㄴ"),
        "ㄷ": atomic("left", 3, "S200", "기본 ㄷ"),
        "ㅂ": atomic("left", 4, "S200", "기본 ㅂ"),
        "ㅅ": atomic("left", 6, "S200", "기본 ㅅ"),
        "ㅇ": atomic("left", 7, "S200", "기본 ㅇ"),
        "ㅈ": atomic("left", 9, "S200", "기본 ㅈ"),
        "ㅋ": atomic("left", 1, "L400", "ㄱ 위치 긴 자극"),
        "ㄹ": atomic("left", 2, "L400", "ㄴ 위치 긴 자극"),
        "ㅌ": atomic("left", 3, "L400", "ㄷ 위치 긴 자극"),
        "ㅍ": atomic("left", 4, "L400", "ㅂ 위치 긴 자극"),
        "ㅎ": atomic("left", 6, "L400", "ㅅ 위치 긴 자극"),
        "ㅁ": atomic("left", 7, "L400", "ㅇ 위치 긴 자극"),
        "ㅊ": atomic("left", 9, "L400", "ㅈ 위치 긴 자극"),
        "ㄲ": composite(["ㄱ", "ㄱ"], 150, "ㄱ 반복"),
        "ㄸ": composite(["ㄷ", "ㄷ"], 150, "ㄷ 반복"),
        "ㅃ": composite(["ㅂ", "ㅂ"], 150, "ㅂ 반복"),
        "ㅆ": composite(["ㅅ", "ㅅ"], 150, "ㅅ 반복"),
        "ㅉ": composite(["ㅈ", "ㅈ"], 150, "ㅈ 반복"),
        "ㅣ": atomic("left", 5, "S200", "중앙 짧은 자극"),
        "ㅡ": atomic("left", 5, "L400", "중앙 긴 자극"),
        # The old four directional vowels are retained as raw motion scripts.
        "ㅏ": raw("#5/150.#5/d,4/i/100.#4/150", "중앙→motor4 motion"),
        "ㅓ": raw("#5/150.#5/d,6/i/100.#6/150", "중앙→motor6 motion"),
        "ㅗ": raw("#5/150.#5/d,2/i/100.#2/150", "중앙→motor2 motion"),
        "ㅜ": raw("#5/150.#5/d,8/i/100.#8/150", "중앙→motor8 motion"),
        "ㅢ": composite(["ㅡ", "ㅣ"], 150, "ㅡ + ㅣ"),
        "ㅑ": composite(["ㅏ", "ㅏ"], 150),
        "ㅕ": composite(["ㅓ", "ㅓ"], 150),
        "ㅛ": composite(["ㅗ", "ㅗ"], 150),
        "ㅠ": composite(["ㅜ", "ㅜ"], 150),
        "ㅐ": composite(["ㅏ", "ㅣ"], 150),
        "ㅔ": composite(["ㅓ", "ㅣ"], 150),
        "ㅒ": composite(["ㅑ", "ㅣ"], 150),
        "ㅖ": composite(["ㅕ", "ㅣ"], 150),
        "ㅘ": composite(["ㅗ", "ㅏ"], 150),
        "ㅙ": composite(["ㅗ", "ㅏ", "ㅣ"], 150),
        "ㅚ": composite(["ㅗ", "ㅣ"], 150),
        "ㅝ": composite(["ㅜ", "ㅓ"], 150),
        "ㅞ": composite(["ㅜ", "ㅓ", "ㅣ"], 150),
        "ㅟ": composite(["ㅜ", "ㅣ"], 150),
    })

    # Every internal value is now direct end-to-start ISI.
    s.timing_defaults_ms = _default_timing_defaults_ms()
    s.timing_matrix_ms = {
        context: {
            prev_cls: {
                next_cls: int(s.timing_defaults_ms[context])
                for next_cls in TIMING_CLASS_LABELS
            }
            for prev_cls in TIMING_CLASS_LABELS
        }
        for context in TIMING_CONTEXT_LABELS
    }
    s.cv_short_isi_ms = int(s.timing_defaults_ms["cv"])
    s.cv_long_isi_ms = int(s.timing_defaults_ms["cv"])
    s.cvc_cv_short_isi_ms = int(s.timing_defaults_ms["cvc_cv"])
    s.cvc_cv_long_isi_ms = int(s.timing_defaults_ms["cvc_cv"])
    s.sync_duration_cv_rows_from_fields()
    s.timing_semantics_version = 6
    return s


# ----------------------------- serial -----------------------------

class SerialController:
    def __init__(self) -> None:
        self.port = None
        self.port_name = ""
        self.last_ack = ""

    def is_connected(self) -> bool:
        return bool(self.port and self.port.is_open)

    def list_ports(self) -> List[Tuple[str, str]]:
        if serial is None:
            return []
        return [(p.device, p.description) for p in serial.tools.list_ports.comports()]

    def connect(self, port_name: str, baudrate: int) -> None:
        if serial is None:
            raise RuntimeError("pyserial이 설치되지 않았습니다.")
        self.close()
        self.port = serial.Serial(port_name, int(baudrate), timeout=0.15, write_timeout=1.0)
        self.port_name = port_name
        time.sleep(1.5)
        self.port.reset_input_buffer()

    def close(self) -> None:
        if self.port is not None:
            try:
                self.port.close()
            except Exception:
                pass
        self.port = None
        self.port_name = ""

    def send(self, command: str, wait_ack: bool = True, timeout_sec: float = 2.0) -> str:
        command = command.strip()
        if not command:
            raise RuntimeError("빈 serial command입니다.")
        if not self.is_connected():
            self.last_ack = "DRY RUN"
            print("[DRY SERIAL]", command)
            return self.last_ack
        self.port.reset_input_buffer()
        self.port.write((command + "\n").encode("utf-8"))
        self.port.flush()
        if not wait_ack:
            return "SENT"
        deadline = time.perf_counter() + timeout_sec
        lines: List[str] = []
        while time.perf_counter() < deadline:
            raw = self.port.readline()
            if not raw:
                QApplication.processEvents()
                continue
            line = raw.decode("utf-8", errors="replace").strip()
            if line:
                lines.append(line)
            if line == "OK":
                self.last_ack = "OK"
                return "OK"
            if line.startswith("ERR"):
                self.last_ack = line
                raise RuntimeError(line)
        self.last_ack = "TIMEOUT: " + " | ".join(lines)
        raise RuntimeError(self.last_ack)

    def emergency_off(self) -> None:
        try:
            self.send("@0#0", wait_ack=False)
        except Exception:
            pass


# ----------------------------- utilities -----------------------------

def add_shadow(widget: QWidget, blur: int = 34, y: int = 9, alpha: int = 24) -> None:
    effect = QGraphicsDropShadowEffect(widget)
    effect.setBlurRadius(blur)
    effect.setOffset(0, y)
    effect.setColor(QColor(0, 0, 0, alpha))
    widget.setGraphicsEffect(effect)


def card() -> QFrame:
    f = QFrame()
    f.setObjectName("Card")
    add_shadow(f)
    return f


def button(text: str, kind: str = "secondary") -> QPushButton:
    b = QPushButton(text)
    b.setProperty("kind", kind)
    b.setCursor(Qt.PointingHandCursor)
    return b


def clear_layout(layout) -> None:
    while layout.count():
        item = layout.takeAt(0)
        w = item.widget()
        if w is not None:
            w.deleteLater()
        child = item.layout()
        if child is not None:
            clear_layout(child)


def load_syllables(path: Path, limit: int = 200) -> List[str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl이 필요합니다.")
    if not path.exists():
        raise RuntimeError(f"syllable 파일이 없습니다: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)
    if first is None:
        return []
    header = [str(v).strip().lower() if v is not None else "" for v in first]
    has_header = any(h in header for h in ("syl", "syll", "syllable", "count"))
    if has_header:
        if "syl" in header:
            idx = header.index("syl")
        elif "syll" in header:
            idx = header.index("syll")
        elif "syllable" in header:
            idx = header.index("syllable")
        else:
            idx = 0
        all_rows = rows
    else:
        idx = 0
        all_rows = iter([first] + list(rows))
    labels: List[str] = []
    for row in all_rows:
        if row and len(row) > idx and row[idx] is not None:
            txt = str(row[idx]).strip()
            if txt:
                labels.append(txt[0])
        if len(labels) >= int(limit):
            break
    return list(dict.fromkeys(labels))


def load_general_quiz_syllables(path: Path) -> List[str]:
    """Load the strict source inventory for the typed CV/CVC quiz.

    The preferred syllable column is detected from common English/Korean header
    names. If the selected cells contain words rather than one-character entries,
    every modern Hangul syllable is split out in source order. Duplicate syllables
    are removed while keeping the first occurrence.
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl이 필요합니다.")
    path = Path(path)
    if not path.exists():
        raise RuntimeError(
            f"일반 퀴즈용 syllable_top200.xlsx가 없습니다: {path}\n"
            "프로그램 폴더에 파일을 넣거나 일반 퀴즈 화면에서 XLSX를 선택하세요."
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)
    if first is None:
        return []

    preferred_headers = {
        "syl", "syll", "syllable", "letter", "char",
        "음절", "글자", "문자", "단어", "word",
    }
    normalized = [str(value).strip().lower() if value is not None else "" for value in first]
    source_column = next((i for i, value in enumerate(normalized) if value in preferred_headers), None)
    has_header = source_column is not None
    if source_column is None:
        source_column = 0

    source_rows = rows if has_header else iter([first] + list(rows))
    syllables: List[str] = []
    seen: set[str] = set()
    for row in source_rows:
        if not row or len(row) <= source_column or row[source_column] is None:
            continue
        text = str(row[source_column]).strip()
        for ch in text:
            if 0xAC00 <= ord(ch) <= 0xD7A3 and ch not in seen:
                seen.add(ch)
                syllables.append(ch)
    return syllables


# ----------------------------- editor widgets -----------------------------

class PositionGrid(QWidget):
    def __init__(self, on_change=None):
        super().__init__()
        self.on_change = on_change
        self.value = 1
        self.buttons: Dict[int, QPushButton] = {}
        lay = QGridLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)
        n = 1
        for r in range(3):
            for c in range(3):
                b = QPushButton(str(n))
                b.setObjectName("PositionButton")
                b.setCheckable(True)
                b.setMinimumSize(72, 62)
                b.clicked.connect(lambda _=False, v=n: self.set_value(v))
                self.buttons[n] = b
                lay.addWidget(b, r, c)
                n += 1
        self.set_value(1, emit=False)

    def set_value(self, value: int, emit: bool = True) -> None:
        self.value = int(value)
        for n, b in self.buttons.items():
            b.setChecked(n == self.value)
        if emit and self.on_change:
            self.on_change(self.value)



class TimingRulesDialog(QDialog):
    """Advanced duration thresholds, ISI matrix, and pair overrides."""

    def __init__(self, app: "MainWindow"):
        super().__init__(app)
        self.app = app
        self.setWindowTitle("세부 ISI 규칙")
        self.resize(1280, 880)
        self.setMinimumSize(1000, 720)
        self.matrix_data = _deep_copy_timing_matrix(app.setup.timing_matrix_ms)
        self.pair_data = dict(app.setup.timing_pair_overrides_ms)
        self.jamo_data = copy.deepcopy(app.setup.jamo)
        self.current_context = ""
        self._loading = False

        # Keep the complete advanced editor in one vertical scroll area and
        # leave save/cancel actions fixed at the bottom for high-DPI displays.
        outer_root = QVBoxLayout(self)
        outer_root.setContentsMargins(14, 12, 14, 12)
        outer_root.setSpacing(10)

        editor_scroll = QScrollArea()
        editor_scroll.setObjectName("TimingRulesScroll")
        editor_scroll.setWidgetResizable(True)
        editor_scroll.setFrameShape(QFrame.NoFrame)
        editor_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        editor_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        editor_body = QWidget()
        editor_body.setObjectName("TimingRulesScrollBody")
        root = QVBoxLayout(editor_body)
        root.setContentsMargins(10, 10, 10, 12)
        root.setSpacing(14)
        editor_scroll.setWidget(editor_body)
        outer_root.addWidget(editor_scroll, 1)

        title = QLabel("자극 종류별 end-to-start ISI")
        title.setObjectName("SectionTitle")
        root.addWidget(title)
        hint = QLabel(
            "모든 값은 앞 자극이 완전히 끝난 뒤 다음 자극을 시작하기까지 기다리는 ISI입니다. "
            "예를 들어 앞 자극이 1000 ms이고 ISI가 200 ms이면 다음 자극은 1200 ms에 시작합니다. "
            "따라서 자극 길이가 아무리 길어도 겹침 예외가 생기지 않습니다. "
            "자모 pair override가 있으면 아래 class 행렬보다 우선합니다. "
            "CV와 CVC 첫 자음 구간에서는 실제 자음 duration으로 ‘짧음/김’ 행을 선택하며, "
            "해당 행의 셀 하나를 바꾸면 행 전체가 함께 바뀌고 오른쪽 기본 설정과도 동기화됩니다."
        )
        hint.setObjectName("Hint")
        hint.setWordWrap(True)
        root.addWidget(hint)

        duration_card = QFrame()
        duration_card.setObjectName("EstimatorCard")
        duration_grid = QGridLayout(duration_card)
        duration_grid.setContentsMargins(16, 13, 16, 13)
        duration_grid.setHorizontalSpacing(16)
        duration_grid.setVerticalSpacing(8)
        duration_title = QLabel("Duration 분기 기준")
        duration_title.setObjectName("EstimatorTitle")
        duration_grid.addWidget(duration_title, 0, 0, 1, 4)
        self.consonant_duration_split = QSpinBox()
        self.consonant_duration_split.setRange(1, 5000)
        self.consonant_duration_split.setSuffix(" ms")
        self.consonant_duration_split.setValue(int(app.setup.cv_duration_split_ms))
        self.consonant_duration_split.setToolTip("이 값 이하이면 짧은 자음 ISI 행, 초과이면 긴 자음 ISI 행을 사용합니다.")
        self.motion_duration_split = QSpinBox()
        self.motion_duration_split.setRange(1, 5000)
        self.motion_duration_split.setSuffix(" ms")
        self.motion_duration_split.setValue(int(app.setup.motion_duration_split_ms))
        self.motion_duration_split.setToolTip("이 값 이하의 motion은 짧은 모션, 초과이면 긴 모션 행/열을 사용합니다.")
        duration_grid.addWidget(QLabel("자음 duration 기준"), 1, 0)
        duration_grid.addWidget(self.consonant_duration_split, 1, 1)
        duration_grid.addWidget(QLabel("모션 duration 기준"), 1, 2)
        duration_grid.addWidget(self.motion_duration_split, 1, 3)
        duration_note = QLabel(
            "기본값은 둘 다 300 ms입니다. 경계값과 같은 duration은 ‘짧은’ 쪽으로 분류됩니다."
        )
        duration_note.setObjectName("Hint")
        duration_note.setWordWrap(True)
        duration_grid.addWidget(duration_note, 2, 0, 1, 4)
        root.addWidget(duration_card)

        context_row = QHBoxLayout()
        context_row.addWidget(QLabel("적용 구간"))
        self.context_combo = QComboBox()
        for key, label in TIMING_CONTEXT_LABELS.items():
            self.context_combo.addItem(label, key)
        context_row.addWidget(self.context_combo, 1)
        fill_btn = button("현재 기본 ISI로 전체 채우기", "ghost")
        fill_btn.clicked.connect(self.fill_matrix_with_default)
        context_row.addWidget(fill_btn)
        root.addLayout(context_row)

        class_note = QLabel("행 = 앞 자극 종류 · 열 = 뒤 자극 종류")
        class_note.setObjectName("Hint")
        root.addWidget(class_note)
        self.matrix_rule_note = QLabel("")
        self.matrix_rule_note.setObjectName("MiniCard")
        self.matrix_rule_note.setWordWrap(True)
        root.addWidget(self.matrix_rule_note)
        classes = list(TIMING_CLASS_LABELS)
        self.matrix_table = QTableWidget(len(classes), len(classes))
        self.matrix_table.setObjectName("TimingMatrixTable")
        self.matrix_table.setHorizontalHeaderLabels([TIMING_CLASS_LABELS[x] for x in classes])
        self.matrix_table.setVerticalHeaderLabels([TIMING_CLASS_LABELS[x] for x in classes])

        # Do not stretch seven columns into cells that are too narrow for
        # ``450 ms`` under Windows 125/150% DPI scaling.  Fixed logical widths
        # preserve a readable editor, while the table supplies a horizontal
        # scroll bar on smaller displays.
        h_header = self.matrix_table.horizontalHeader()
        h_header.setSectionResizeMode(QHeaderView.Fixed)
        h_header.setMinimumSectionSize(132)
        h_header.setDefaultSectionSize(144)
        h_header.setStretchLastSection(False)
        v_header = self.matrix_table.verticalHeader()
        v_header.setSectionResizeMode(QHeaderView.Fixed)
        v_header.setMinimumSectionSize(46)
        v_header.setDefaultSectionSize(46)
        v_header.setMinimumWidth(110)
        self.matrix_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.matrix_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.matrix_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.matrix_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.matrix_table.setMinimumHeight(365)
        self.matrix_table.setMaximumHeight(405)

        self.matrix_spins: Dict[Tuple[str, str], QSpinBox] = {}
        for r, prev_cls in enumerate(classes):
            self.matrix_table.setRowHeight(r, 46)
            for c, next_cls in enumerate(classes):
                sp = QSpinBox()
                sp.setObjectName("TimingMatrixSpin")
                sp.setRange(0, 5000)
                sp.setSuffix(" ms")
                sp.setAlignment(Qt.AlignCenter)
                sp.setMinimumWidth(116)
                sp.setMinimumHeight(38)
                sp.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                sp.setToolTip(
                    f"{TIMING_CLASS_LABELS[prev_cls]} 종료 → {TIMING_CLASS_LABELS[next_cls]} 시작 ISI"
                )
                self.matrix_table.setCellWidget(r, c, sp)
                self.matrix_spins[(prev_cls, next_cls)] = sp
                sp.valueChanged.connect(
                    lambda value, p=prev_cls, n=next_cls:
                    self._matrix_value_changed(p, n, value)
                )
        root.addWidget(self.matrix_table)

        pair_title_row = QHBoxLayout()
        pair_title = QLabel("특정 자모 pair override")
        pair_title.setObjectName("EstimatorTitle")
        pair_title_row.addWidget(pair_title)
        pair_title_row.addStretch(1)
        self.pair_from = QComboBox(); self.pair_from.addItems(CONSONANTS + VOWELS + list(COMPOUND_FINALS))
        self.pair_to = QComboBox(); self.pair_to.addItems(CONSONANTS + VOWELS + list(COMPOUND_FINALS))
        self.pair_isi = QSpinBox(); self.pair_isi.setRange(0, 5000); self.pair_isi.setSuffix(" ms")
        add_pair = button("추가", "secondary")
        del_pair = button("선택 삭제", "ghost")
        add_pair.clicked.connect(self.add_pair_override)
        del_pair.clicked.connect(self.delete_pair_override)
        pair_title_row.addWidget(self.pair_from)
        pair_title_row.addWidget(QLabel("→"))
        pair_title_row.addWidget(self.pair_to)
        pair_title_row.addWidget(self.pair_isi)
        pair_title_row.addWidget(add_pair)
        pair_title_row.addWidget(del_pair)
        root.addLayout(pair_title_row)

        self.pair_table = QTableWidget(0, 3)
        self.pair_table.setHorizontalHeaderLabels(["앞 자모", "뒤 자모", "ISI"])
        self.pair_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.pair_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.pair_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.pair_table.verticalHeader().setVisible(False)
        self.pair_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.pair_table.setMinimumHeight(96)
        self.pair_table.setMaximumHeight(150)
        root.addWidget(self.pair_table)


        actions = QHBoxLayout()
        actions.addStretch(1)
        cancel_btn = button("취소", "ghost")
        save_btn = button("ISI 규칙 저장", "primary")
        cancel_btn.clicked.connect(self.reject)
        save_btn.clicked.connect(self.accept_rules)
        actions.addWidget(cancel_btn)
        actions.addWidget(save_btn)
        outer_root.addLayout(actions)

        # Preserve natural content height on short/high-DPI displays.
        editor_body.setMinimumHeight(max(980, editor_body.sizeHint().height()))

        self.context_combo.currentIndexChanged.connect(self._switch_context)
        self._switch_context(0)

    def _matrix_value_changed(self, prev_cls: str, _next_cls: str, value: int) -> None:
        """Keep duration-based CV rows uniform and visibly synchronized."""
        if self._loading:
            return
        context = self.current_context
        if context in ("cv", "cvc_cv") and prev_cls in ("short", "long"):
            self._loading = True
            try:
                for next_cls in TIMING_CLASS_LABELS:
                    spin = self.matrix_spins[(prev_cls, next_cls)]
                    spin.blockSignals(True)
                    spin.setValue(int(value))
                    spin.blockSignals(False)
            finally:
                self._loading = False

    def _save_current_context(self) -> None:
        if self._loading or not self.current_context:
            return
        context = self.current_context
        rows: Dict[str, Dict[str, int]] = {}
        for prev_cls in TIMING_CLASS_LABELS:
            rows[prev_cls] = {}
            for next_cls in TIMING_CLASS_LABELS:
                rows[prev_cls][next_cls] = int(self.matrix_spins[(prev_cls, next_cls)].value())
        self.matrix_data[context] = rows

        prefix = context + "|"
        for key in [k for k in self.pair_data if k.startswith(prefix)]:
            self.pair_data.pop(key, None)
        for r in range(self.pair_table.rowCount()):
            from_lab = self.pair_table.item(r, 0).text()
            to_lab = self.pair_table.item(r, 1).text()
            isi_item = self.pair_table.item(r, 2)
            raw_isi = isi_item.data(Qt.UserRole)
            if raw_isi is None:
                match = re.search(r"\d+", isi_item.text())
                raw_isi = int(match.group(0)) if match else 0
            isi = int(raw_isi)
            key = DesignSetup.pair_override_key(context, from_lab, to_lab)
            self.pair_data[key] = isi

    def _switch_context(self, _index: int) -> None:
        if self.current_context:
            self._save_current_context()
        context = str(self.context_combo.currentData())
        self.current_context = context
        self._loading = True
        default = int(self.app.setup.timing_defaults_ms.get(context, 0))
        matrix = self.matrix_data.get(context, {})
        for prev_cls in TIMING_CLASS_LABELS:
            for next_cls in TIMING_CLASS_LABELS:
                value = matrix.get(prev_cls, {}).get(next_cls, default)
                self.matrix_spins[(prev_cls, next_cls)].setValue(int(value))
        self._load_pairs(context)
        if context in ("cv", "cvc_cv"):
            self.matrix_rule_note.setText(
                "이 구간의 ‘짧음’과 ‘김’ 행은 앞 자음의 실제 duration으로 선택됩니다. "
                "셀 하나를 수정하면 해당 행 전체가 같은 값으로 바뀌고, 저장 시 기본 화면의 "
                "짧은/긴 ISI 값에도 그대로 반영됩니다."
            )
            self.matrix_rule_note.show()
        else:
            self.matrix_rule_note.setText(
                "이 구간은 각 행·열 셀을 독립적으로 설정할 수 있습니다. "
                "특정 자모 pair override가 있으면 행렬보다 먼저 적용됩니다."
            )
            self.matrix_rule_note.show()
        self._loading = False

    def _load_pairs(self, context: str) -> None:
        self.pair_table.setRowCount(0)
        prefix = context + "|"
        items = []
        for key, value in self.pair_data.items():
            if not key.startswith(prefix):
                continue
            parts = key.split("|", 2)
            if len(parts) == 3:
                items.append((parts[1], parts[2], int(value)))
        for from_lab, to_lab, isi in sorted(items):
            self._append_pair_row(from_lab, to_lab, isi)

    def _append_pair_row(self, from_lab: str, to_lab: str, isi: int) -> None:
        r = self.pair_table.rowCount()
        self.pair_table.insertRow(r)
        self.pair_table.setItem(r, 0, QTableWidgetItem(from_lab))
        self.pair_table.setItem(r, 1, QTableWidgetItem(to_lab))
        item = QTableWidgetItem(f"{int(isi)} ms")
        item.setData(Qt.UserRole, int(isi))
        self.pair_table.setItem(r, 2, item)

    def add_pair_override(self) -> None:
        from_lab, to_lab, isi = self.pair_from.currentText(), self.pair_to.currentText(), self.pair_isi.value()
        for r in range(self.pair_table.rowCount()):
            if self.pair_table.item(r, 0).text() == from_lab and self.pair_table.item(r, 1).text() == to_lab:
                self.pair_table.item(r, 2).setText(f"{isi} ms")
                self.pair_table.item(r, 2).setData(Qt.UserRole, int(isi))
                self.pair_table.selectRow(r)
                return
        self._append_pair_row(from_lab, to_lab, isi)
        self.pair_table.selectRow(self.pair_table.rowCount() - 1)

    def delete_pair_override(self) -> None:
        r = self.pair_table.currentRow()
        if r >= 0:
            self.pair_table.removeRow(r)
    
    def fill_matrix_with_default(self) -> None:
        context = str(self.context_combo.currentData())
        default = int(self.app.setup.timing_defaults_ms.get(context, 0))
        self._loading = True
        try:
            for sp in self.matrix_spins.values():
                sp.setValue(default)

            # The two duration rows are linked to the basic controls, so a
            # reset must restore those values rather than replacing both with
            # the short/default value.
            if context == "cv":
                row_values = {
                    "short": int(self.app.cv_short_isi.value()),
                    "long": int(self.app.cv_long_isi.value()),
                }
            elif context == "cvc_cv":
                row_values = {
                    "short": int(self.app.cvc_cv_short_isi.value()),
                    "long": int(self.app.cvc_cv_long_isi.value()),
                }
            else:
                row_values = {}

            for prev_cls, value in row_values.items():
                for next_cls in TIMING_CLASS_LABELS:
                    self.matrix_spins[(prev_cls, next_cls)].setValue(value)
        finally:
            self._loading = False

    def accept_rules(self) -> None:
        self._save_current_context()
        setup = self.app.setup
        setup.timing_matrix_ms = _expand_motion_duration_matrix(
            self.matrix_data, setup.timing_defaults_ms
        )
        setup.timing_pair_overrides_ms = dict(self.pair_data)
        setup.jamo = copy.deepcopy(self.jamo_data)
        setup.cv_duration_split_ms = int(self.consonant_duration_split.value())
        setup.motion_duration_split_ms = int(self.motion_duration_split.value())
        setup.timing_semantics_version = 6

        # The advanced matrix is authoritative. Mirror its synchronized rows
        # back to the four simple controls without triggering a second rewrite.
        setup.sync_duration_cv_fields_from_matrix()
        for widget, value in (
            (self.app.cv_short_isi, setup.cv_short_isi_ms),
            (self.app.cv_long_isi, setup.cv_long_isi_ms),
            (self.app.cvc_cv_short_isi, setup.cvc_cv_short_isi_ms),
            (self.app.cvc_cv_long_isi, setup.cvc_cv_long_isi_ms),
        ):
            widget.blockSignals(True)
            widget.setValue(int(value))
            widget.blockSignals(False)

        self.app.update_duration_rule_summary()
        self.app.mark_dirty()
        self.app.update_duration_estimate()
        self.app.mapping_editor.preview_current(show_error=False)
        self.accept()


class MappingEditor(QWidget):
    def __init__(self, app: "MainWindow"):
        super().__init__()
        self.app = app
        self.current_label = "ㄱ"
        self._loading = False
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(16)

        title_row = QHBoxLayout()
        self.title = QLabel("ㄱ 디자인")
        self.title.setObjectName("SectionTitle")
        title_row.addWidget(self.title)
        title_row.addStretch(1)
        title_row.addWidget(QLabel("고급 ISI 유형"))
        self.timing_class_combo = QComboBox()
        self.timing_class_combo.addItem("자동 판별", "auto")
        for key, label in TIMING_CLASS_EDITOR_LABELS.items():
            self.timing_class_combo.addItem(label, key)
        self.timing_class_combo.setToolTip(
            "대부분 자동 판별로 둡니다. 모션은 실제 duration과 고급 설정의 모션 기준으로 "
            "짧은 모션/긴 모션 ISI 행렬을 자동 선택합니다."
        )
        title_row.addWidget(self.timing_class_combo)
        self.mode_combo = QComboBox()
        self.mode_combo.addItem("Atomic · 한 위치/시간 패턴", "atomic")
        self.mode_combo.addItem("Composite · 다른 자모 조합", "composite")
        self.mode_combo.addItem("Raw · 직접 serial script", "raw")
        self.mode_combo.addItem("Unmapped", "unmapped")
        self.mode_combo.currentIndexChanged.connect(self.update_mode_visibility)
        title_row.addWidget(self.mode_combo)
        root.addLayout(title_row)

        self.mode_stack = QStackedWidget()
        root.addWidget(self.mode_stack, 1)

        # Atomic page
        atomic_page = QWidget()
        ag = QGridLayout(atomic_page)
        ag.setContentsMargins(4, 4, 4, 4)
        ag.setHorizontalSpacing(18)
        ag.setVerticalSpacing(12)
        ag.addWidget(QLabel("팔"), 0, 0)
        self.arm_combo = QComboBox()
        self.arm_combo.addItem("왼팔 · #", "left")
        self.arm_combo.addItem("오른팔 · @", "right")
        ag.addWidget(self.arm_combo, 0, 1)
        ag.addWidget(QLabel("Temporal factor"), 1, 0)
        self.temporal_combo = QComboBox()
        for key, cfg in TEMPORAL_PRESETS.items():
            self.temporal_combo.addItem(cfg["label"], key)
        self.temporal_combo.currentIndexChanged.connect(self.update_custom_visibility)
        ag.addWidget(self.temporal_combo, 1, 1)
        ag.addWidget(QLabel("논리적 위치"), 2, 0, Qt.AlignTop)
        self.pos_grid = PositionGrid()
        ag.addWidget(self.pos_grid, 2, 1)
        self.custom_frame = QFrame()
        custom = QHBoxLayout(self.custom_frame)
        custom.setContentsMargins(0, 0, 0, 0)
        self.on1_spin = QSpinBox(); self.on1_spin.setRange(1, 3000); self.on1_spin.setValue(200); self.on1_spin.setSuffix(" ms")
        self.gap_spin = QSpinBox(); self.gap_spin.setRange(0, 3000); self.gap_spin.setValue(100); self.gap_spin.setSuffix(" ms")
        self.on2_spin = QSpinBox(); self.on2_spin.setRange(0, 3000); self.on2_spin.setValue(200); self.on2_spin.setSuffix(" ms")
        custom.addWidget(QLabel("ON1")); custom.addWidget(self.on1_spin)
        custom.addWidget(QLabel("Gap")); custom.addWidget(self.gap_spin)
        custom.addWidget(QLabel("ON2")); custom.addWidget(self.on2_spin)
        ag.addWidget(self.custom_frame, 3, 1)
        ag.setRowStretch(4, 1)
        self.mode_stack.addWidget(atomic_page)

        # Composite page
        composite_page = QWidget()
        cg = QVBoxLayout(composite_page)
        cg.setContentsMargins(4, 4, 4, 4)
        hint = QLabel(
            "예: ㅢ = ㅡ + ㅣ. 각 step은 저장된 자모 디자인을 참조합니다. "
            "오른쪽의 Composite 기본 ISI는 앞 Step이 완전히 끝난 뒤 기다리는 시간이며, "
            "이미 만든 행의 개별 ISI는 자동으로 덮어쓰지 않습니다."
        )
        hint.setWordWrap(True)
        hint.setObjectName("Hint")
        cg.addWidget(hint)
        self.steps_table = QTableWidget(0, 2)
        self.steps_table.setObjectName("StepsTable")
        self.steps_table.setHorizontalHeaderLabels(["구성 자모", "앞 Step 종료 후 ISI (ms)"])
        table_header = self.steps_table.horizontalHeader()
        table_header.setSectionResizeMode(0, QHeaderView.Stretch)
        table_header.setSectionResizeMode(1, QHeaderView.Fixed)
        table_header.resizeSection(1, 196)
        table_header.setMinimumHeight(50)
        table_header.setDefaultAlignment(Qt.AlignCenter)
        self.steps_table.verticalHeader().setVisible(False)
        self.steps_table.verticalHeader().setDefaultSectionSize(60)
        self.steps_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.steps_table.setSelectionMode(QTableWidget.SingleSelection)
        self.steps_table.setShowGrid(False)
        self.steps_table.setMinimumHeight(260)
        cg.addWidget(self.steps_table, 1)
        cr = QHBoxLayout()
        self.ref_combo = QComboBox(); self.ref_combo.addItems(CONSONANTS + VOWELS)
        self.ref_combo.setToolTip("새로 추가할 구성 자모")
        self.step_gap_spin = QSpinBox()
        self.step_gap_spin.setRange(0, 2000)
        self.step_gap_spin.setValue(int(self.app.setup.timing_defaults_ms.get("composite", 350)))
        self.step_gap_spin.setSuffix(" ms")
        self.step_gap_spin.setToolTip("앞 Step이 완전히 끝난 뒤 새 Step 시작까지 기다릴 ISI입니다. 오른쪽 Composite 기본 ISI와 동기화됩니다.")
        add_btn = button("＋ Step", "secondary")
        del_btn = button("삭제", "ghost")
        apply_gap_btn = button("기존 Step에 적용", "ghost")
        apply_gap_btn.setToolTip("현재 ISI 값을 이미 만든 2번째 이후 Step에 한 번에 적용합니다.")
        up_btn = QPushButton()
        up_btn.setObjectName("RoundIconButton")
        up_btn.setIcon(QIcon(str(ARROW_UP_SVG)))
        up_btn.setIconSize(QSize(22, 22))
        up_btn.setFixedSize(48, 48)
        up_btn.setCursor(Qt.PointingHandCursor)
        up_btn.setToolTip("선택한 step을 위로 이동")
        down_btn = QPushButton()
        down_btn.setObjectName("RoundIconButton")
        down_btn.setIcon(QIcon(str(ARROW_DOWN_SVG)))
        down_btn.setIconSize(QSize(22, 22))
        down_btn.setFixedSize(48, 48)
        down_btn.setCursor(Qt.PointingHandCursor)
        down_btn.setToolTip("선택한 step을 아래로 이동")
        add_btn.clicked.connect(self.add_step)
        del_btn.clicked.connect(self.delete_step)
        apply_gap_btn.clicked.connect(self.apply_step_gap_to_existing)
        up_btn.clicked.connect(lambda: self.move_step(-1))
        down_btn.clicked.connect(lambda: self.move_step(1))
        cr.addWidget(QLabel("추가할 자모"))
        cr.addWidget(self.ref_combo)
        cr.addWidget(QLabel("새 Step ISI"))
        cr.addWidget(self.step_gap_spin)
        cr.addWidget(add_btn)
        cr.addWidget(del_btn)
        cr.addWidget(apply_gap_btn)
        cr.addStretch(1)
        cr.addWidget(up_btn)
        cr.addWidget(down_btn)
        cg.addLayout(cr)
        self.mode_stack.addWidget(composite_page)

        # Raw page
        raw_page = QWidget()
        rg = QVBoxLayout(raw_page)
        rg.setContentsMargins(4, 4, 4, 4)
        raw_hint = QLabel("고급 기능: @/# marker를 포함한 전체 serial script를 직접 입력합니다.")
        raw_hint.setObjectName("Hint")
        self.raw_edit = QPlainTextEdit()
        self.raw_edit.setPlaceholderText("예: #5/150.#5/d,4/i/100.#4/150")
        rg.addWidget(raw_hint)
        rg.addWidget(self.raw_edit, 1)
        self.mode_stack.addWidget(raw_page)

        # Unmapped page
        unmapped = QWidget()
        ul = QVBoxLayout(unmapped)
        lab = QLabel("이 자모를 사용하지 않습니다. 이 자모가 포함된 음절은 컴파일되지 않습니다.")
        lab.setObjectName("Hint")
        lab.setWordWrap(True)
        ul.addWidget(lab)
        ul.addStretch(1)
        self.mode_stack.addWidget(unmapped)

        note_row = QHBoxLayout()
        note_row.addWidget(QLabel("메모"))
        self.note_edit = QLineEdit()
        self.note_edit.setPlaceholderText("디자인 의도 또는 설명")
        note_row.addWidget(self.note_edit, 1)
        root.addLayout(note_row)

        self.preview = QPlainTextEdit()
        self.preview.setReadOnly(True)
        self.preview.setMaximumHeight(105)
        self.preview.setPlaceholderText("컴파일된 command가 표시됩니다.")
        root.addWidget(self.preview)

        action = QHBoxLayout()
        save_btn = button("매핑 저장", "primary")
        preview_btn = button("미리보기", "secondary")
        test_btn = button("이 자모 느껴보기", "secondary")
        save_btn.clicked.connect(self.save_current)
        preview_btn.clicked.connect(self.preview_current)
        test_btn.clicked.connect(self.test_current)
        action.addWidget(preview_btn)
        action.addWidget(test_btn)
        action.addStretch(1)
        action.addWidget(save_btn)
        root.addLayout(action)

        # The command preview is live. Every mapping control refreshes both the
        # compiled command and the estimated physical playback duration.
        self.mode_combo.currentIndexChanged.connect(self._on_editor_value_changed)
        self.timing_class_combo.currentIndexChanged.connect(self._on_editor_value_changed)
        self.arm_combo.currentIndexChanged.connect(self._on_editor_value_changed)
        self.temporal_combo.currentIndexChanged.connect(self._on_editor_value_changed)
        self.pos_grid.on_change = lambda _value: self._on_editor_value_changed()
        for spin in (self.on1_spin, self.gap_spin, self.on2_spin):
            spin.valueChanged.connect(self._on_editor_value_changed)
        self.raw_edit.textChanged.connect(self._on_editor_value_changed)

    def _on_editor_value_changed(self, *_args) -> None:
        if self._loading:
            return
        self.preview_current(show_error=False)

    def update_mode_visibility(self) -> None:
        mode = self.mode_combo.currentData()
        index = {"atomic": 0, "composite": 1, "raw": 2, "unmapped": 3}.get(mode, 3)
        self.mode_stack.setCurrentIndex(index)
        self.update_custom_visibility()

    def update_custom_visibility(self) -> None:
        self.custom_frame.setVisible(self.temporal_combo.currentData() == "CUSTOM")

    def load_label(self, label: str) -> None:
        self._loading = True
        self.current_label = label
        self.title.setText(f"{label} 디자인")
        spec = self.app.setup.get_spec(label)
        idx = self.mode_combo.findData(spec.mode)
        self.mode_combo.setCurrentIndex(max(0, idx))
        timing_idx = self.timing_class_combo.findData(spec.timing_class)
        self.timing_class_combo.setCurrentIndex(max(0, timing_idx))
        self.arm_combo.setCurrentIndex(max(0, self.arm_combo.findData(spec.arm)))
        self.pos_grid.set_value(spec.position, emit=False)
        self.temporal_combo.setCurrentIndex(max(0, self.temporal_combo.findData(spec.temporal)))
        self.on1_spin.setValue(int(spec.on1_ms))
        self.gap_spin.setValue(int(spec.gap_ms))
        self.on2_spin.setValue(int(spec.on2_ms))
        self.steps_table.setRowCount(0)
        for step in spec.steps:
            self._append_step_row(str(step.get("ref", "")), int(step.get("isi_before_ms", 0)))
        self.raw_edit.setPlainText(spec.raw_command)
        self.note_edit.setText(spec.note)
        self.update_mode_visibility()
        self.set_default_composite_gap(self.app.setup.timing_defaults_ms.get("composite", 350))
        self.preview_current(show_error=False)
        self._loading = False

    def set_default_composite_gap(self, value: int) -> None:
        """Synchronize the new-Step gap control with the global timing default.

        Existing table rows keep their explicit per-step values. This prevents a
        global default change from silently destroying a deliberately customized
        composite pattern.
        """
        value = max(0, int(value))
        self.step_gap_spin.blockSignals(True)
        self.step_gap_spin.setValue(value)
        self.step_gap_spin.blockSignals(False)

    def _append_step_row(self, ref: str, isi_before: int) -> None:
        r = self.steps_table.rowCount()
        self.steps_table.insertRow(r)
        self.steps_table.setRowHeight(r, 60)
        combo = QComboBox()
        combo.setObjectName("TableComboBox")
        combo.addItems(CONSONANTS + VOWELS)
        combo.setCurrentText(ref)
        combo.setMinimumHeight(44)
        spin = QSpinBox()
        spin.setObjectName("TableSpinBox")
        spin.setRange(0, 2000)
        spin.setSuffix(" ms")
        spin.setValue(int(isi_before))
        spin.setMinimumHeight(44)
        spin.setMinimumWidth(184)
        combo.currentTextChanged.connect(self._on_editor_value_changed)
        spin.valueChanged.connect(self._on_editor_value_changed)
        self.steps_table.setCellWidget(r, 0, combo)
        self.steps_table.setCellWidget(r, 1, spin)

    def add_step(self) -> None:
        gap = 0 if self.steps_table.rowCount() == 0 else self.step_gap_spin.value()
        self._append_step_row(self.ref_combo.currentText(), gap)
        self.steps_table.selectRow(self.steps_table.rowCount() - 1)
        self._on_editor_value_changed()

    def delete_step(self) -> None:
        r = self.steps_table.currentRow()
        if r >= 0:
            self.steps_table.removeRow(r)
            self._on_editor_value_changed()

    def apply_step_gap_to_existing(self) -> None:
        """Apply the new-Step gap value to every existing step after the first."""
        gap = int(self.step_gap_spin.value())
        for r in range(self.steps_table.rowCount()):
            spin = self.steps_table.cellWidget(r, 1)
            if spin is not None:
                spin.setValue(0 if r == 0 else gap)
        self._on_editor_value_changed()

    def move_step(self, delta: int) -> None:
        r = self.steps_table.currentRow()
        nr = r + delta
        if r < 0 or nr < 0 or nr >= self.steps_table.rowCount():
            return
        a = self._step_at(r); b = self._step_at(nr)
        self._set_step_at(r, b); self._set_step_at(nr, a)
        self.steps_table.selectRow(nr)
        self._on_editor_value_changed()

    def _step_at(self, r: int) -> Dict[str, Any]:
        return {
            "ref": self.steps_table.cellWidget(r, 0).currentText(),
            "isi_before_ms": self.steps_table.cellWidget(r, 1).value(),
        }

    def _set_step_at(self, r: int, step: Dict[str, Any]) -> None:
        self.steps_table.cellWidget(r, 0).setCurrentText(str(step["ref"]))
        self.steps_table.cellWidget(r, 1).setValue(int(step.get("isi_before_ms", 0)))

    def collect_spec(self) -> JamoSpec:
        mode = self.mode_combo.currentData()
        steps = [self._step_at(r) for r in range(self.steps_table.rowCount())]
        return JamoSpec(
            mode=mode,
            arm=self.arm_combo.currentData(),
            position=self.pos_grid.value,
            temporal=self.temporal_combo.currentData(),
            on1_ms=self.on1_spin.value(),
            gap_ms=self.gap_spin.value(),
            on2_ms=self.on2_spin.value(),
            steps=steps,
            raw_command=self.raw_edit.toPlainText().strip(),
            note=self.note_edit.text().strip(),
            timing_class=self.timing_class_combo.currentData(),
        )

    def save_current(self) -> None:
        self.app.setup.set_spec(self.current_label, self.collect_spec())
        self.app.mark_dirty()
        self.app.refresh_jamo_lists()
        self.preview_current()
        self.app.update_duration_estimate()
        self.app.toast(f"{self.current_label} 매핑 저장됨")

    def preview_current(self, show_error: bool = True) -> None:
        backup = self.app.setup.jamo.get(self.current_label)
        self.app.setup.set_spec(self.current_label, self.collect_spec())
        try:
            cmd = HangulCompiler(self.app.setup).compile_jamo(self.current_label)
            dur = HangulCompiler.estimate_duration_ms(cmd)
            timing_class = self.app.setup.timing_class_for(self.current_label)
            class_label = TIMING_CLASS_EDITOR_LABELS.get(timing_class, timing_class)
            self.preview.setPlainText(f"{cmd}\n예상 duration: {dur} ms · 고급 ISI 유형: {class_label}")
        except Exception as exc:
            self.preview.setPlainText(f"컴파일 오류: {exc}")
            if show_error:
                self.app.toast(str(exc), error=True)
        finally:
            if backup is None:
                self.app.setup.jamo.pop(self.current_label, None)
            else:
                self.app.setup.jamo[self.current_label] = backup

    def test_current(self) -> None:
        self.save_current()
        try:
            cmd = HangulCompiler(self.app.setup).compile_jamo(self.current_label)
            self.app.send_command(cmd)
        except Exception as exc:
            QMessageBox.critical(self, "자극 오류", str(exc))



def load_startup_default_setup() -> Tuple[DesignSetup, Path, str]:
    """Load ``hangul_tactile_setups/default.json`` at application startup.

    The lookup is case-insensitive so ``Default.json`` also works. If no user
    default exists yet, create one from the bundled default setup (or from the
    built-in setup as a final fallback). A malformed user file never prevents
    the program from opening; it falls back to the built-in setup and reports
    the reason after the UI is ready.
    """
    SETUP_DIR.mkdir(parents=True, exist_ok=True)
    canonical = SETUP_DIR / "default.json"
    candidates = sorted(
        (p for p in SETUP_DIR.glob("*.json") if p.stem.casefold() == "default"),
        key=lambda p: (p.name.casefold() != "default.json", p.name.casefold()),
    )
    selected = candidates[0] if candidates else canonical

    if selected.exists():
        try:
            with selected.open("r", encoding="utf-8") as handle:
                raw_setup = json.load(handle)
            previous_version = int(raw_setup.get("timing_semantics_version", 1) or 1)
            setup = DesignSetup.from_dict(raw_setup)
            if previous_version < 6:
                message = (
                    f"{selected.name}의 기존 SOA 타이밍을 end-to-start ISI로 자동 변환했습니다. "
                    "확인 후 저장하면 v17 형식으로 저장됩니다."
                )
            else:
                message = f"시작 기본 셋업 불러옴 · {selected.name}"
            return setup, selected, message
        except Exception as exc:
            fallback = make_default_setup()
            return (
                fallback,
                canonical,
                f"{selected.name}을 불러오지 못해 내장 기본값을 사용합니다: {exc}",
            )

    # Seed the user-visible default.json once. Prefer the package's JSON copy
    # because it is easy to inspect and edit outside the application.
    setup = make_default_setup()
    bundled_default = RESOURCE_DIR / "hangul_tactile_default_setup.json"
    if bundled_default.exists():
        try:
            with bundled_default.open("r", encoding="utf-8") as handle:
                setup = DesignSetup.from_dict(json.load(handle))
        except Exception:
            setup = make_default_setup()

    try:
        with canonical.open("w", encoding="utf-8") as handle:
            json.dump(setup.to_dict(), handle, ensure_ascii=False, indent=2)
        message = "hangul_tactile_setups/default.json 생성 및 자동 불러오기 완료"
    except Exception as exc:
        message = f"default.json 생성 실패 · 내장 기본값 사용: {exc}"
    return setup, canonical, message


# ----------------------------- main window -----------------------------

class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Hangul Tactile Designer")
        self.resize(1680, 980)
        self.setMinimumSize(1280, 760)
        self.setup, startup_setup_path, self._startup_setup_message = load_startup_default_setup()
        self.setup_path: Optional[Path] = startup_setup_path
        self.dirty = False
        self.serial_ctl = SerialController()
        self.voice_backend = None
        self.voice_backend_path = RESOURCE_DIR / "hangul_voice_backend.py"
        if not self.voice_backend_path.exists():
            legacy_backend = RESOURCE_DIR / "hangul_voice_backend_v20.py"
            if legacy_backend.exists():
                self.voice_backend_path = legacy_backend
        self.quiz_labels: List[str] = []
        self.quiz_plan: List[str] = []
        self.quiz_index = -1
        self.quiz_target = ""
        self.quiz_command = ""
        self.quiz_worker = None
        self.quiz_started_perf = 0.0
        self.quiz_rows: List[Dict[str, Any]] = []
        self.quiz_csv_path: Optional[Path] = None

        # Text-answer CV/CVC tactile quiz state.
        self.general_quiz_plan: List[Dict[str, Any]] = []
        self.general_quiz_index = -1
        self.general_quiz_target = ""
        self.general_quiz_command = ""
        self.general_quiz_started_perf = 0.0
        self.general_quiz_replays = 0
        self.general_quiz_answered = False
        self.general_quiz_rows: List[Dict[str, Any]] = []
        self.general_quiz_csv_path: Optional[Path] = None

        # Multiple-choice tactile learning state for selected jamo groups.
        self.learning_plan: List[Dict[str, Any]] = []
        self.learning_index = -1
        self.learning_target = ""
        self.learning_command = ""
        self.learning_started_perf = 0.0
        self.learning_replays = 0
        self.learning_answered = False
        self.learning_rows: List[Dict[str, Any]] = []
        self.learning_csv_path: Optional[Path] = None
        self.learning_choice_buttons: List[QPushButton] = []
        self.build_ui()
        self.apply_style()
        self.refresh_ports()
        self.refresh_jamo_lists()
        self.mapping_editor.load_label("ㄱ")
        self.sync_setup_controls_from_model()
        self.update_setup_title()
        if self._startup_setup_message:
            is_error = "실패" in self._startup_setup_message or "못해" in self._startup_setup_message
            QTimer.singleShot(
                250,
                lambda message=self._startup_setup_message, error=is_error:
                self.toast(message, error=error),
            )

    # --------- UI shell ---------
    def build_ui(self) -> None:
        central = QWidget()
        central.setObjectName("Root")
        self.setCentralWidget(central)
        outer = QHBoxLayout(central)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        sidebar = QFrame()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(246)
        sl = QVBoxLayout(sidebar)
        sl.setContentsMargins(22, 28, 22, 24)
        sl.setSpacing(10)
        brand = QLabel("Tactile\nHangul")
        brand.setObjectName("Brand")
        sl.addWidget(brand)
        subtitle = QLabel("DESIGN STUDIO")
        subtitle.setObjectName("Eyebrow")
        sl.addWidget(subtitle)
        sl.addSpacing(22)
        self.nav_buttons: List[QPushButton] = []
        navs = [
            ("디자인", 0),
            ("심플 테스트", 1),
            ("선택 학습", 2),
            ("일반 퀴즈", 3),
            ("보이스 퀴즈", 4),
            ("하드웨어 · 설정", 5),
        ]
        for text, idx in navs:
            b = QPushButton(text)
            b.setObjectName("NavButton")
            b.setCheckable(True)
            b.clicked.connect(lambda _=False, i=idx: self.switch_page(i))
            self.nav_buttons.append(b)
            sl.addWidget(b)
        sl.addStretch(1)
        self.sidebar_status = QLabel("DRY RUN")
        self.sidebar_status.setObjectName("StatusPill")
        self.sidebar_status.setAlignment(Qt.AlignCenter)
        sl.addWidget(self.sidebar_status)
        emergency = button("Emergency OFF", "danger")
        emergency.clicked.connect(self.emergency_off)
        sl.addWidget(emergency)
        outer.addWidget(sidebar)

        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(28, 22, 30, 28)
        rl.setSpacing(14)
        top = QFrame(); top.setObjectName("TopBar")
        tl = QHBoxLayout(top); tl.setContentsMargins(22, 14, 16, 14)
        self.setup_title = QLabel("Default Hangul Design")
        self.setup_title.setObjectName("TopTitle")
        tl.addWidget(self.setup_title)
        tl.addStretch(1)
        new_btn = button("새 디자인", "ghost")
        load_btn = button("불러오기", "secondary")
        save_btn = button("저장", "primary")
        save_as_btn = button("다른 이름으로", "secondary")
        new_btn.clicked.connect(self.new_setup)
        load_btn.clicked.connect(self.load_setup_dialog)
        save_btn.clicked.connect(self.save_setup)
        save_as_btn.clicked.connect(lambda: self.save_setup(save_as=True))
        for b in (new_btn, load_btn, save_as_btn, save_btn):
            tl.addWidget(b)
        rl.addWidget(top)

        self.stack = QStackedWidget()
        self.stack.addWidget(self.build_design_page())
        self.stack.addWidget(self.build_simple_page())
        self.stack.addWidget(self.build_learning_page())
        self.stack.addWidget(self.build_general_quiz_page())
        self.stack.addWidget(self.build_quiz_page())
        self.stack.addWidget(self.build_hardware_page())
        rl.addWidget(self.stack, 1)
        self.toast_label = QLabel("")
        self.toast_label.setObjectName("Toast")
        self.toast_label.setAlignment(Qt.AlignCenter)
        self.toast_label.hide()
        rl.addWidget(self.toast_label)
        outer.addWidget(right, 1)
        self.switch_page(0)

    def build_design_page(self) -> QWidget:
        page = QWidget()
        root = QHBoxLayout(page)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(16)

        left = card(); left.setMinimumWidth(340); left.setMaximumWidth(390)
        ll = QVBoxLayout(left); ll.setContentsMargins(22, 20, 22, 20); ll.setSpacing(12)
        t = QLabel("자모 라이브러리"); t.setObjectName("SectionTitle"); ll.addWidget(t)
        hint = QLabel("자모를 선택하고 atomic / composite / raw 방식으로 디자인하세요.")
        hint.setObjectName("Hint"); hint.setWordWrap(True); ll.addWidget(hint)
        self.jamo_tabs = QTabWidget()
        self.consonant_list = QListWidget(); self.vowel_list = QListWidget()
        self.consonant_list.itemClicked.connect(lambda item: self.mapping_editor.load_label(item.data(Qt.UserRole)))
        self.vowel_list.itemClicked.connect(lambda item: self.mapping_editor.load_label(item.data(Qt.UserRole)))
        self.jamo_tabs.setObjectName("JamoSegmentedTabs")
        self.jamo_tabs.addTab(self.consonant_list, "자음")
        self.jamo_tabs.addTab(self.vowel_list, "모음")
        self.jamo_tabs.setDocumentMode(True)
        self.jamo_tabs.setUsesScrollButtons(False)
        self.jamo_tabs.tabBar().setExpanding(True)
        self.jamo_tabs.currentChanged.connect(self.update_jamo_tab_state)
        self.update_jamo_tab_state(0)
        ll.addWidget(self.jamo_tabs, 1)
        validate = button("전체 디자인 검증", "secondary")
        validate.clicked.connect(self.validate_design)
        ll.addWidget(validate)
        root.addWidget(left)

        center = card()
        cl = QVBoxLayout(center); cl.setContentsMargins(24, 22, 24, 22)
        self.mapping_editor = MappingEditor(self)
        cl.addWidget(self.mapping_editor)
        root.addWidget(center, 1)

        # Keep the timing controls usable on laptops or Windows displays with
        # large DPI scaling.  The previous layout allowed Qt to vertically
        # squeeze the estimator card when the available height was short,
        # causing its labels and spin box to overlap.  A dedicated scroll area
        # now preserves each control's natural height instead.
        timing = card(); timing.setMinimumWidth(360); timing.setMaximumWidth(430)
        timing_outer = QVBoxLayout(timing)
        timing_outer.setContentsMargins(0, 0, 0, 0)
        timing_scroll = QScrollArea()
        timing_scroll.setObjectName("TimingScroll")
        timing_scroll.setWidgetResizable(True)
        timing_scroll.setFrameShape(QFrame.NoFrame)
        timing_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        timing_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        timing_body = QWidget()
        timing_body.setObjectName("TimingScrollBody")
        timing_body.setMinimumHeight(1520)
        gl = QVBoxLayout(timing_body); gl.setContentsMargins(18, 20, 18, 20); gl.setSpacing(12)
        timing_scroll.setWidget(timing_body)
        timing_outer.addWidget(timing_scroll)
        gt = QLabel("음절 타이밍"); gt.setObjectName("SectionTitle"); gl.addWidget(gt)
        gh = QLabel(
            "모든 타이밍 값은 동일한 end-to-start ISI입니다. 앞 자모의 전체 촉각 자극이 끝난 뒤 "
            "설정한 시간만큼 기다리고 다음 자모를 시작합니다. 자극 길이가 길어도 겹침 예외가 생기지 않습니다."
        )
        gh.setObjectName("Hint"); gh.setWordWrap(True); gl.addWidget(gh)
        self.default_comp_gap = self._labeled_spin(gl, "Composite 기본 ISI", 0, 5000, 150)

        duration_hint = QLabel(
            "자→모 ISI는 앞 자음의 실제 duration으로 짧음/김 행을 선택합니다. "
            "선택된 값도 항상 앞 자극 종료 후 기다리는 시간입니다."
        )
        duration_hint.setObjectName("Hint"); duration_hint.setWordWrap(True); gl.addWidget(duration_hint)
        self.duration_rule_summary = QLabel("")
        self.duration_rule_summary.setObjectName("Hint")
        self.duration_rule_summary.setWordWrap(True)
        gl.addWidget(self.duration_rule_summary)
        self.cv_short_isi = self._labeled_spin(gl, "CV · 짧은 자음→모음 ISI", 0, 5000, 250)
        self.cv_long_isi = self._labeled_spin(gl, "CV · 긴 자음→모음 ISI", 0, 5000, 250)
        self.cvc_cv_short_isi = self._labeled_spin(gl, "CVC · 짧은 첫 자음→모음 ISI", 0, 5000, 250)
        self.cvc_cv_long_isi = self._labeled_spin(gl, "CVC · 긴 첫 자음→모음 ISI", 0, 5000, 250)
        # Compatibility aliases used by a few older helper methods.
        self.cv_gap = self.cv_short_isi
        self.cvc_cv_gap = self.cvc_cv_short_isi

        self.cvc_vc_gap = self._labeled_spin(gl, "CVC · 모음→끝 자음 ISI", 0, 5000, 250)
        self.comp_final_gap = self._labeled_spin(gl, "복합 종성 내부 ISI", 0, 5000, 150)
        self.syllable_gap = self._labeled_spin(gl, "음절 경계 ISI", 0, 7000, 350)
        self.word_gap = self._labeled_spin(gl, "단어 경계 ISI", 0, 10000, 650)
        self.default_comp_gap.setToolTip(
            "Composite에서 앞 Step이 완전히 끝난 뒤 새 Step을 시작하기까지의 기본 대기시간입니다. "
            "기존 행의 개별 ISI는 자동으로 바뀌지 않습니다."
        )
        self.cv_short_isi.setToolTip("고급 CV 행렬의 ‘짧음’ 행 전체와 연결됩니다.")
        self.cv_long_isi.setToolTip("고급 CV 행렬의 ‘김’ 행 전체와 연결됩니다.")
        self.cvc_cv_short_isi.setToolTip("고급 CVC 첫 자음→모음 행렬의 ‘짧음’ 행 전체와 연결됩니다.")
        self.cvc_cv_long_isi.setToolTip("고급 CVC 첫 자음→모음 행렬의 ‘김’ 행 전체와 연결됩니다.")
        self.cvc_vc_gap.setToolTip("받침 있는 CVC에서 모음 자극 종료 후 종성 시작까지 기다리는 시간")
        self.comp_final_gap.setToolTip("ㄳ, ㄵ, ㄺ 등에서 첫 종성 자극 종료 후 둘째 종성 시작까지 기다리는 시간")
        self.syllable_gap.setToolTip("앞 음절의 마지막 모터 자극이 완전히 끝난 뒤 다음 음절 초성 시작까지")
        self.word_gap.setToolTip("앞 단어의 마지막 모터 자극이 완전히 끝난 뒤 다음 단어 첫 초성 시작까지")
        self.default_comp_gap.valueChanged.connect(self.on_default_composite_gap_changed)
        for sp in (
            self.cv_short_isi, self.cv_long_isi,
            self.cvc_cv_short_isi, self.cvc_cv_long_isi,
            self.cvc_vc_gap, self.comp_final_gap, self.syllable_gap, self.word_gap,
        ):
            sp.valueChanged.connect(self.sync_model_from_setup_controls)
        advanced_btn = button("고급 duration·유형·pair ISI", "secondary")
        advanced_btn.setToolTip(
            "자음/모션 duration 분기 기준, 자극 종류별 ISI 행렬, 특정 자모 pair override를 설정합니다."
        )
        advanced_btn.clicked.connect(self.open_timing_rules)
        gl.addWidget(advanced_btn)

        gl.addSpacing(4)
        estimate_card = QFrame()
        estimate_card.setObjectName("EstimatorCard")
        estimate_card.setMinimumHeight(178)
        estimate_layout = QVBoxLayout(estimate_card)
        estimate_layout.setContentsMargins(16, 15, 16, 15)
        estimate_layout.setSpacing(7)
        estimate_title = QLabel("현재 셋업 속도 추정")
        estimate_title.setObjectName("EstimatorTitle")
        estimate_layout.addWidget(estimate_title)
        length_row = QHBoxLayout()
        length_row.addWidget(QLabel("평균 단어 길이"))
        length_row.addStretch(1)
        self.word_length_spin = QSpinBox()
        self.word_length_spin.setRange(1, 10)
        self.word_length_spin.setValue(2)
        self.word_length_spin.setSuffix(" 음절")
        self.word_length_spin.setMaximumWidth(145)
        self.word_length_spin.valueChanged.connect(self.update_duration_estimate)
        length_row.addWidget(self.word_length_spin)
        estimate_layout.addLayout(length_row)
        self.duration_estimate_value = QLabel("계산 중…")
        self.duration_estimate_value.setObjectName("EstimatorValue")
        self.duration_estimate_value.setWordWrap(True)
        estimate_layout.addWidget(self.duration_estimate_value)
        self.duration_estimate_note = QLabel("")
        self.duration_estimate_note.setObjectName("EstimatorNote")
        self.duration_estimate_note.setWordWrap(True)
        estimate_layout.addWidget(self.duration_estimate_note)
        gl.addWidget(estimate_card)

        explain = QLabel(
            "예: 자음 duration 1000 ms + ISI 200 ms → 다음 자모는 1200 ms에 시작\n"
            "짧음/김/모션 유형은 어떤 ISI 행을 고를지만 결정\n"
            "Composite·CV·CVC·음절·단어 모두 같은 ISI 의미"
        )
        explain.setObjectName("MiniCard")
        explain.setWordWrap(True)
        explain.setMinimumHeight(82)
        gl.addWidget(explain)
        gl.addStretch(1)
        root.addWidget(timing)
        return page

    def build_simple_page(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(16)
        hero = card(); hl = QVBoxLayout(hero); hl.setContentsMargins(32, 28, 32, 28); hl.setSpacing(14)
        title = QLabel("입력한 한글을 바로 느껴보기"); title.setObjectName("HeroTitle"); hl.addWidget(title)
        sub = QLabel("예: 각, 한글, 의. 현재 저장된 디자인과 end-to-start ISI 규칙으로 command를 생성합니다.")
        sub.setObjectName("Hint"); hl.addWidget(sub)
        row = QHBoxLayout()
        self.simple_input = QLineEdit(); self.simple_input.setPlaceholderText("각"); self.simple_input.setMinimumHeight(58)
        compile_btn = button("Command 만들기", "secondary")
        play_btn = button("자극 재생", "primary")
        compile_btn.clicked.connect(self.compile_simple)
        play_btn.clicked.connect(self.play_simple)
        self.simple_input.returnPressed.connect(self.play_simple)
        row.addWidget(self.simple_input, 1); row.addWidget(compile_btn); row.addWidget(play_btn)
        hl.addLayout(row)
        root.addWidget(hero)

        details = card(); dl = QVBoxLayout(details); dl.setContentsMargins(24, 22, 24, 22)
        self.simple_summary = QLabel("아직 컴파일된 자극이 없습니다."); self.simple_summary.setObjectName("SectionTitle"); dl.addWidget(self.simple_summary)
        self.simple_command = QPlainTextEdit(); self.simple_command.setReadOnly(True); dl.addWidget(self.simple_command, 1)
        root.addWidget(details, 1)
        return page


    def build_learning_page(self) -> QWidget:
        """Build a multiple-choice learning page for selected jamo groups."""
        page = QScrollArea()
        page.setObjectName("LearningScroll")
        page.setWidgetResizable(True)
        page.setFrameShape(QFrame.NoFrame)
        page.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        page.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        content = QWidget()
        content.setObjectName("LearningScrollBody")
        root = QVBoxLayout(content)
        root.setContentsMargins(0, 0, 8, 0)
        root.setSpacing(14)
        page.setWidget(content)

        config = card()
        cl = QGridLayout(config)
        cl.setContentsMargins(24, 20, 24, 20)
        cl.setHorizontalSpacing(14)
        cl.setVerticalSpacing(10)

        title = QLabel("선택 자모 학습")
        title.setObjectName("SectionTitle")
        cl.addWidget(title, 0, 0, 1, 6)

        intro = QLabel(
            "학습할 자모 그룹을 선택하면 해당 자모의 촉각 자극을 무작위로 제시하고, "
            "화면에 나타난 보기 중 하나를 눌러 맞히는 연습을 진행합니다."
        )
        intro.setObjectName("Hint")
        intro.setWordWrap(True)
        cl.addWidget(intro, 1, 0, 1, 6)

        cl.addWidget(QLabel("학습 그룹"), 2, 0)
        self.learning_basic_consonant_check = QCheckBox("기본 자음 14개")
        self.learning_basic_consonant_check.setChecked(True)
        self.learning_double_consonant_check = QCheckBox("쌍자음 5개")
        self.learning_basic_vowel_check = QCheckBox("기본 모음 14개")
        self.learning_compound_vowel_check = QCheckBox("복합 모음 7개")
        cl.addWidget(self.learning_basic_consonant_check, 2, 1)
        cl.addWidget(self.learning_double_consonant_check, 2, 2)
        cl.addWidget(self.learning_basic_vowel_check, 3, 1)
        cl.addWidget(self.learning_compound_vowel_check, 3, 2)

        cl.addWidget(QLabel("문제 수"), 2, 4)
        self.learning_trials = QSpinBox()
        self.learning_trials.setRange(1, 1000)
        self.learning_trials.setValue(40)
        self.learning_trials.setSuffix(" 회")
        cl.addWidget(self.learning_trials, 2, 5)

        cl.addWidget(QLabel("보기 수"), 3, 4)
        self.learning_choice_count = QSpinBox()
        self.learning_choice_count.setRange(2, 8)
        self.learning_choice_count.setValue(4)
        self.learning_choice_count.setSuffix(" 개")
        self.learning_choice_count.setToolTip(
            "정답을 포함한 보기 개수입니다. 선택한 전체 자모 수보다 많으면 가능한 수만 표시합니다."
        )
        cl.addWidget(self.learning_choice_count, 3, 5)

        cl.addWidget(QLabel("세션 ID"), 4, 0)
        self.learning_session_edit = QLineEdit()
        self.learning_session_edit.setPlaceholderText("선택 사항 · 예: jamo_practice01")
        cl.addWidget(self.learning_session_edit, 4, 1, 1, 3)
        self.learning_start_btn = button("선택 학습 시작", "primary")
        self.learning_start_btn.clicked.connect(self.start_learning_session)
        cl.addWidget(self.learning_start_btn, 4, 4, 1, 2)

        group_hint = QLabel(
            "여러 그룹을 함께 선택할 수 있습니다. 선택된 그룹은 문제 수 안에서 가능한 한 균형 있게 출제되며, "
            "같은 그룹 안에서는 모든 자모가 한 번씩 나온 뒤 다시 섞입니다."
        )
        group_hint.setObjectName("Hint")
        group_hint.setWordWrap(True)
        cl.addWidget(group_hint, 5, 0, 1, 6)
        root.addWidget(config)

        stage = card()
        st = QVBoxLayout(stage)
        st.setContentsMargins(26, 22, 26, 22)
        st.setSpacing(14)

        self.learning_progress = QLabel("학습을 시작하세요.")
        self.learning_progress.setObjectName("Eyebrow")
        st.addWidget(self.learning_progress)

        self.learning_instruction = QLabel("학습 그룹을 선택한 뒤 선택 학습 시작을 누르세요.")
        self.learning_instruction.setObjectName("HeroTitle")
        self.learning_instruction.setAlignment(Qt.AlignCenter)
        self.learning_instruction.setWordWrap(True)
        st.addWidget(self.learning_instruction)

        self.learning_feedback = QLabel("")
        self.learning_feedback.setObjectName("Feedback")
        self.learning_feedback.setAlignment(Qt.AlignCenter)
        self.learning_feedback.setWordWrap(True)
        st.addWidget(self.learning_feedback)

        choice_title = QLabel("보기")
        choice_title.setObjectName("EstimatorTitle")
        st.addWidget(choice_title)

        self.learning_choice_widget = QWidget()
        self.learning_choice_layout = QGridLayout(self.learning_choice_widget)
        self.learning_choice_layout.setContentsMargins(0, 0, 0, 0)
        self.learning_choice_layout.setHorizontalSpacing(12)
        self.learning_choice_layout.setVerticalSpacing(12)
        st.addWidget(self.learning_choice_widget, 1)

        action_row = QHBoxLayout()
        self.learning_replay_btn = button("자극 다시 재생", "secondary")
        self.learning_replay_btn.setEnabled(False)
        self.learning_replay_btn.clicked.connect(lambda: self.play_learning_stimulus(replay=True))
        self.learning_next_btn = button("다음 문제", "primary")
        self.learning_next_btn.setEnabled(False)
        self.learning_next_btn.clicked.connect(self.next_learning_trial)
        action_row.addWidget(self.learning_replay_btn)
        action_row.addStretch(1)
        action_row.addWidget(self.learning_next_btn)
        st.addLayout(action_row)

        stage.setMinimumHeight(480)
        root.addWidget(stage, 1)
        return page

    def build_general_quiz_page(self) -> QWidget:
        """Build a text-answer tactile quiz for random CV/CVC syllables."""
        page = QScrollArea()
        page.setObjectName("GeneralQuizScroll")
        page.setWidgetResizable(True)
        page.setFrameShape(QFrame.NoFrame)
        page.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        page.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        content = QWidget()
        content.setObjectName("GeneralQuizScrollBody")
        root = QVBoxLayout(content)
        root.setContentsMargins(0, 0, 8, 0)
        root.setSpacing(14)
        page.setWidget(content)

        config = card()
        cl = QGridLayout(config)
        cl.setContentsMargins(24, 20, 24, 20)
        cl.setHorizontalSpacing(14)
        cl.setVerticalSpacing(10)
        title = QLabel("일반 촉각 퀴즈")
        title.setObjectName("SectionTitle")
        cl.addWidget(title, 0, 0, 1, 6)

        intro = QLabel(
            "syllable_top200.xlsx에 실제로 들어 있는 음절만 사용합니다. 파일 셀에 단어가 있으면 "
            "한글 음절 단위로 모두 잘라 중복 없이 출제하며, 선택한 CV/CVC 및 자모 그룹으로 다시 필터링합니다. "
            "답은 아래 채팅 입력창에 직접 입력합니다."
        )
        intro.setObjectName("Hint")
        intro.setWordWrap(True)
        cl.addWidget(intro, 1, 0, 1, 6)

        cl.addWidget(QLabel("음절 구조"), 2, 0)
        self.gq_cv_check = QCheckBox("자+모 (CV)")
        self.gq_cv_check.setChecked(True)
        self.gq_cvc_check = QCheckBox("자+모+자 (CVC)")
        self.gq_cvc_check.setChecked(True)
        cl.addWidget(self.gq_cv_check, 2, 1)
        cl.addWidget(self.gq_cvc_check, 2, 2)
        cl.addWidget(QLabel("문제 수"), 2, 4)
        self.gq_trials = QSpinBox()
        self.gq_trials.setRange(1, 1000)
        self.gq_trials.setValue(40)
        self.gq_trials.setSuffix(" 회")
        cl.addWidget(self.gq_trials, 2, 5)

        cl.addWidget(QLabel("초성/종성"), 3, 0)
        self.gq_basic_consonant_check = QCheckBox("기본 자음 14개")
        self.gq_basic_consonant_check.setChecked(True)
        self.gq_double_consonant_check = QCheckBox("쌍자음 5개")
        self.gq_double_consonant_check.setChecked(True)
        cl.addWidget(self.gq_basic_consonant_check, 3, 1)
        cl.addWidget(self.gq_double_consonant_check, 3, 2)

        cl.addWidget(QLabel("중성"), 4, 0)
        self.gq_basic_vowel_check = QCheckBox("기본 모음 14개")
        self.gq_basic_vowel_check.setChecked(True)
        self.gq_diphthong_check = QCheckBox("이중모음 7개")
        self.gq_diphthong_check.setChecked(True)
        cl.addWidget(self.gq_basic_vowel_check, 4, 1)
        cl.addWidget(self.gq_diphthong_check, 4, 2)

        cl.addWidget(QLabel("출제 음절 XLSX"), 5, 0)
        self.gq_syllable_edit = QLineEdit(str(APP_DIR / "syllable_top200.xlsx"))
        self.gq_syllable_edit.setPlaceholderText("syllable_top200.xlsx 경로")
        cl.addWidget(self.gq_syllable_edit, 5, 1, 1, 4)
        gq_syllable_btn = button("찾기", "ghost")
        gq_syllable_btn.clicked.connect(self.choose_general_quiz_syllable)
        cl.addWidget(gq_syllable_btn, 5, 5)

        cl.addWidget(QLabel("세션 ID"), 6, 0)
        self.gq_session_edit = QLineEdit()
        self.gq_session_edit.setPlaceholderText("선택 사항 · 예: practice01")
        cl.addWidget(self.gq_session_edit, 6, 1, 1, 3)
        self.gq_start_btn = button("Top200 퀴즈 시작", "primary")
        self.gq_start_btn.clicked.connect(self.start_general_quiz)
        cl.addWidget(self.gq_start_btn, 6, 4, 1, 2)

        final_hint = QLabel(
            "출제 후보는 XLSX에서 읽은 음절에 한정됩니다. CVC 종성은 선택한 자음 가운데 실제 한글 종성으로 "
            "가능한 한 글자 자음만 사용하며, ㄸ·ㅃ·ㅉ과 ㄳ·ㄺ 같은 복합 종성은 제외됩니다."
        )
        final_hint.setObjectName("Hint")
        final_hint.setWordWrap(True)
        cl.addWidget(final_hint, 7, 0, 1, 6)
        root.addWidget(config)

        stage = card()
        st = QVBoxLayout(stage)
        st.setContentsMargins(26, 22, 26, 22)
        st.setSpacing(12)
        self.gq_progress = QLabel("퀴즈를 시작하세요.")
        self.gq_progress.setObjectName("Eyebrow")
        st.addWidget(self.gq_progress)
        self.gq_instruction = QLabel("옵션을 선택한 뒤 랜덤 퀴즈 시작을 누르세요.")
        self.gq_instruction.setObjectName("HeroTitle")
        self.gq_instruction.setWordWrap(True)
        st.addWidget(self.gq_instruction)

        self.gq_feedback = QLabel("")
        self.gq_feedback.setObjectName("Feedback")
        self.gq_feedback.setAlignment(Qt.AlignCenter)
        self.gq_feedback.setWordWrap(True)
        st.addWidget(self.gq_feedback)

        self.gq_chat = QPlainTextEdit()
        self.gq_chat.setObjectName("GeneralQuizChat")
        self.gq_chat.setReadOnly(True)
        self.gq_chat.setMinimumHeight(250)
        self.gq_chat.setPlaceholderText("퀴즈 대화가 여기에 표시됩니다.")
        st.addWidget(self.gq_chat, 1)

        answer_row = QHBoxLayout()
        self.gq_answer = QLineEdit()
        self.gq_answer.setPlaceholderText("느낀 음절을 입력하세요. 예: 가 또는 ㄱㅏ")
        self.gq_answer.setMinimumHeight(56)
        self.gq_answer.setEnabled(False)
        self.gq_answer.returnPressed.connect(self.submit_general_quiz_answer)
        self.gq_submit_btn = button("답변 전송", "primary")
        self.gq_submit_btn.setEnabled(False)
        self.gq_submit_btn.clicked.connect(self.submit_general_quiz_answer)
        answer_row.addWidget(self.gq_answer, 1)
        answer_row.addWidget(self.gq_submit_btn)
        st.addLayout(answer_row)

        action_row = QHBoxLayout()
        self.gq_replay_btn = button("자극 다시 재생", "secondary")
        self.gq_replay_btn.setEnabled(False)
        self.gq_replay_btn.clicked.connect(lambda: self.play_general_quiz_stimulus(replay=True))
        self.gq_next_btn = button("다음 문제", "secondary")
        self.gq_next_btn.setEnabled(False)
        self.gq_next_btn.clicked.connect(self.next_general_quiz_trial)
        action_row.addWidget(self.gq_replay_btn)
        action_row.addStretch(1)
        action_row.addWidget(self.gq_next_btn)
        st.addLayout(action_row)
        stage.setMinimumHeight(470)
        root.addWidget(stage, 1)
        return page

    def build_quiz_page(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(14)
        config = card(); cl = QGridLayout(config); cl.setContentsMargins(24, 20, 24, 20); cl.setHorizontalSpacing(12); cl.setVerticalSpacing(10)
        title = QLabel("Voice Quiz"); title.setObjectName("SectionTitle"); cl.addWidget(title, 0, 0, 1, 6)
        cl.addWidget(QLabel("Subject"), 1, 0)
        self.subject_edit = QLineEdit(); self.subject_edit.setPlaceholderText("s01"); cl.addWidget(self.subject_edit, 1, 1)
        cl.addWidget(QLabel("Quiz set"), 1, 2)
        self.quiz_type = QComboBox(); self.quiz_type.addItem("저장된 전체 자모", "jamo_all"); self.quiz_type.addItem("자음", "consonant"); self.quiz_type.addItem("모음", "vowel"); self.quiz_type.addItem("Syllable XLSX", "syllable")
        cl.addWidget(self.quiz_type, 1, 3)
        cl.addWidget(QLabel("Trials"), 1, 4)
        self.quiz_trials = QSpinBox(); self.quiz_trials.setRange(1, 1000); self.quiz_trials.setValue(40); cl.addWidget(self.quiz_trials, 1, 5)
        cl.addWidget(QLabel("Voice backend"), 2, 0)
        self.backend_edit = QLineEdit(str(self.voice_backend_path)); cl.addWidget(self.backend_edit, 2, 1, 1, 4)
        backend_btn = button("찾기", "ghost"); backend_btn.clicked.connect(self.choose_backend); cl.addWidget(backend_btn, 2, 5)
        cl.addWidget(QLabel("Syllable XLSX"), 3, 0)
        self.syllable_edit = QLineEdit(str(APP_DIR / "syllable_top200.xlsx")); cl.addWidget(self.syllable_edit, 3, 1, 1, 3)
        self.syllable_edit.textChanged.connect(self.update_duration_estimate)
        syl_btn = button("찾기", "ghost"); syl_btn.clicked.connect(self.choose_syllable); cl.addWidget(syl_btn, 3, 4)
        self.syllable_limit = QSpinBox(); self.syllable_limit.setRange(1, 2000); self.syllable_limit.setValue(200); cl.addWidget(self.syllable_limit, 3, 5)
        self.syllable_limit.valueChanged.connect(self.update_duration_estimate)
        cl.addWidget(QLabel("Voice profiles"), 4, 0)
        self.voice_profiles_edit = QLineEdit(str(APP_DIR / "voice_profiles")); cl.addWidget(self.voice_profiles_edit, 4, 1, 1, 4)
        profile_btn = button("폴더", "ghost"); profile_btn.clicked.connect(self.choose_voice_profiles); cl.addWidget(profile_btn, 4, 5)
        check_btn = button("음성 모델 확인", "secondary"); check_btn.clicked.connect(self.check_voice_model)
        pooled_btn = button("Pooled 모델 생성", "secondary"); pooled_btn.clicked.connect(self.build_pooled_model)
        start_btn = button("퀴즈 시작", "primary"); start_btn.clicked.connect(self.start_quiz)
        cl.addWidget(check_btn, 5, 0, 1, 2); cl.addWidget(pooled_btn, 5, 2, 1, 2); cl.addWidget(start_btn, 5, 4, 1, 2)
        root.addWidget(config)

        stage = card(); st = QVBoxLayout(stage); st.setContentsMargins(26, 22, 26, 22); st.setSpacing(12)
        self.quiz_progress = QLabel("퀴즈를 시작하세요."); self.quiz_progress.setObjectName("Eyebrow"); st.addWidget(self.quiz_progress)
        self.quiz_instruction = QLabel("현재 디자인으로 촉각 자극을 제시하고 음성 응답 후보를 표시합니다.")
        self.quiz_instruction.setObjectName("HeroTitle"); self.quiz_instruction.setWordWrap(True); st.addWidget(self.quiz_instruction)
        self.quiz_feedback = QLabel(""); self.quiz_feedback.setObjectName("Feedback"); self.quiz_feedback.setAlignment(Qt.AlignCenter); st.addWidget(self.quiz_feedback)
        self.quiz_action = button("자극 시작 + 음성 듣기", "primary"); self.quiz_action.setMinimumHeight(62); self.quiz_action.clicked.connect(self.quiz_action_clicked); self.quiz_action.setEnabled(False); st.addWidget(self.quiz_action)
        self.candidate_widget = QWidget(); self.candidate_layout = QGridLayout(self.candidate_widget); self.candidate_layout.setSpacing(10); st.addWidget(self.candidate_widget)
        manual_row = QHBoxLayout(); self.quiz_manual = QLineEdit(); self.quiz_manual.setPlaceholderText("후보에 없으면 직접 입력"); manual_btn = button("직접 응답", "secondary"); manual_btn.clicked.connect(self.submit_manual_quiz); self.quiz_manual.returnPressed.connect(self.submit_manual_quiz); manual_row.addWidget(self.quiz_manual, 1); manual_row.addWidget(manual_btn); st.addLayout(manual_row)
        root.addWidget(stage, 1)
        return page

    def build_hardware_page(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(16)
        serial_card = card(); sl = QGridLayout(serial_card); sl.setContentsMargins(24, 22, 24, 22); sl.setHorizontalSpacing(12); sl.setVerticalSpacing(10)
        t = QLabel("Serial · Device mapping"); t.setObjectName("SectionTitle"); sl.addWidget(t, 0, 0, 1, 6)
        sl.addWidget(QLabel("COM"), 1, 0)
        self.port_combo = QComboBox(); sl.addWidget(self.port_combo, 1, 1, 1, 2)
        refresh = button("새로고침", "ghost"); refresh.clicked.connect(self.refresh_ports); sl.addWidget(refresh, 1, 3)
        self.connect_btn = button("연결", "primary"); self.connect_btn.clicked.connect(self.toggle_serial); sl.addWidget(self.connect_btn, 1, 4, 1, 2)
        sl.addWidget(QLabel("왼팔 marker"), 2, 0); self.left_marker_edit = QLineEdit("#"); self.left_marker_edit.setMaxLength(1); sl.addWidget(self.left_marker_edit, 2, 1)
        sl.addWidget(QLabel("오른팔 marker"), 2, 2); self.right_marker_edit = QLineEdit("@"); self.right_marker_edit.setMaxLength(1); sl.addWidget(self.right_marker_edit, 2, 3)
        sl.addWidget(QLabel("Baud"), 2, 4); self.baud_spin = QSpinBox(); self.baud_spin.setRange(9600, 2000000); self.baud_spin.setValue(115200); sl.addWidget(self.baud_spin, 2, 5)
        sl.addWidget(QLabel("논리 1..9 → motor"), 3, 0)
        self.motor_map_edit = QLineEdit("3,2,1,6,5,4,9,8,7"); sl.addWidget(self.motor_map_edit, 3, 1, 1, 4)
        apply_btn = button("mapping 적용", "secondary"); apply_btn.clicked.connect(self.apply_hardware_settings); sl.addWidget(apply_btn, 3, 5)
        self.ack_check = QCheckBox("Arduino ACK 확인")
        self.ack_check.setChecked(True)
        self.ack_check.setToolTip("ON이면 Arduino의 OK/ERR 응답을 확인한 뒤 다음 동작으로 넘어갑니다.")
        sl.addWidget(self.ack_check, 4, 0, 1, 3)
        ack_hint = QLabel("ON 상태는 초록색 스위치로 표시됩니다.")
        ack_hint.setObjectName("Hint")
        sl.addWidget(ack_hint, 4, 3, 1, 3)
        soa_fw_hint = QLabel(
            "현재 프로그램은 사용 중인 기존 Arduino 코드를 그대로 사용하며 raw의 /i, /d를 PWM 숫자로 바꾸지 않습니다. "
            "모든 조합 타이밍은 end-to-start ISI라서 새 자모가 앞 자모 종료 전에 겹쳐 시작하지 않습니다. "
            "Arduino 코드는 지금 사용 중인 버전을 그대로 두면 됩니다."
        )
        soa_fw_hint.setObjectName("Hint")
        soa_fw_hint.setWordWrap(True)
        sl.addWidget(soa_fw_hint, 5, 0, 1, 6)
        root.addWidget(serial_card)

        test_card = card(); tl = QVBoxLayout(test_card); tl.setContentsMargins(24, 22, 24, 22)
        top = QHBoxLayout(); tt = QLabel("Hardware test"); tt.setObjectName("SectionTitle"); top.addWidget(tt); top.addStretch(1)
        self.hw_arm = QComboBox(); self.hw_arm.addItem("왼팔", "left"); self.hw_arm.addItem("오른팔", "right")
        self.hw_temporal = QComboBox();
        for k, v in TEMPORAL_PRESETS.items(): self.hw_temporal.addItem(v["label"], k)
        top.addWidget(self.hw_arm); top.addWidget(self.hw_temporal); tl.addLayout(top)
        grid = QGridLayout(); grid.setSpacing(12)
        n = 1
        for r in range(3):
            for c in range(3):
                b = QPushButton(str(n)); b.setObjectName("HardwareButton"); b.setMinimumSize(120, 82); b.clicked.connect(lambda _=False, p=n: self.test_hardware_position(p)); grid.addWidget(b, r, c); n += 1
        tl.addLayout(grid)
        root.addWidget(test_card, 1)
        return page

    def _labeled_spin(self, layout: QVBoxLayout, label: str, lo: int, hi: int, value: int) -> QSpinBox:
        """Create a timing field that remains readable in a narrow side card.

        Long Korean labels and a 180 px spin box could not reliably fit on one
        row at Windows high-DPI scaling.  Stacking the label above a full-width
        editor removes the horizontal clipping while the containing timing
        panel continues to scroll vertically.
        """
        field = QFrame()
        field.setObjectName("TimingField")
        field_layout = QVBoxLayout(field)
        field_layout.setContentsMargins(0, 2, 0, 4)
        field_layout.setSpacing(6)

        label_widget = QLabel(label)
        label_widget.setObjectName("TimingFieldLabel")
        label_widget.setWordWrap(True)
        label_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        field_layout.addWidget(label_widget)

        sp = QSpinBox()
        sp.setObjectName("TimingFieldSpin")
        sp.setRange(lo, hi)
        sp.setValue(value)
        sp.setSuffix(" ms")
        sp.setMinimumWidth(0)
        sp.setMaximumWidth(16777215)
        sp.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        field_layout.addWidget(sp)

        layout.addWidget(field)
        return sp

    # --------- style ---------
    def apply_style(self) -> None:
        self.setFont(QFont("Segoe UI Variable", 10))
        up_url = ARROW_UP_SVG.as_posix()
        down_url = ARROW_DOWN_SVG.as_posix()
        toggle_on_url = TOGGLE_ON_SVG.as_posix()
        toggle_off_url = TOGGLE_OFF_SVG.as_posix()
        self.setStyleSheet(f"""
        QWidget#Root {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #F5F5F7, stop:0.52 #F1F2F5, stop:1 #ECEEF2);
            color: #1C1C1E;
        }}
        QWidget {{
            font-family: "Segoe UI Variable", "Segoe UI", "Apple SD Gothic Neo", "Malgun Gothic";
            font-size: 15px;
            color: #1C1C1E;
        }}
        QFrame#Sidebar {{
            background: rgba(247, 247, 249, 246);
            border-right: 1px solid #DEDEE3;
        }}
        QLabel#Brand {{ font-size: 31px; font-weight: 780; letter-spacing: -1px; color: #1C1C1E; }}
        QLabel#Eyebrow {{ color: #8E8E93; font-size: 12px; font-weight: 760; letter-spacing: 1.5px; }}
        QPushButton#NavButton {{
            text-align: left; border: 1px solid transparent; border-radius: 18px;
            padding: 15px 17px; color: #636366; font-weight: 680; background: transparent;
        }}
        QPushButton#NavButton:hover {{ background: rgba(229,229,234,190); color: #1C1C1E; }}
        QPushButton#NavButton:checked {{
            background: #3A3A3C; color: #FFFFFF; border: 1px solid #3A3A3C;
        }}

        QFrame#TopBar, QFrame#Card {{
            background: rgba(255, 255, 255, 235);
            border: 1px solid rgba(209, 209, 214, 205);
            border-radius: 28px;
        }}
        QFrame#TopBar {{ background: rgba(255,255,255,245); }}
        QLabel#TopTitle {{ font-size: 19px; font-weight: 760; color: #1C1C1E; }}
        QLabel#SectionTitle {{ font-size: 21px; font-weight: 760; letter-spacing: -0.3px; color: #1C1C1E; }}
        QLabel#HeroTitle {{ font-size: 29px; font-weight: 790; letter-spacing: -0.8px; color: #1C1C1E; }}
        QLabel#Hint {{ color: #6E6E73; line-height: 1.35; }}
        QLabel#MiniCard {{
            background: rgba(242,242,247,220); border: 1px solid #E1E1E6;
            border-radius: 18px; padding: 15px; color: #636366;
        }}
        QLabel#StatusPill {{
            background: #E9E9ED; border: 1px solid #D8D8DC;
            border-radius: 14px; padding: 9px; color: #48484A; font-weight: 760;
        }}
        QLabel#Feedback {{
            background: #F2F2F7; border: 1px solid #E1E1E6;
            border-radius: 18px; padding: 13px; color: #48484A; font-weight: 680;
        }}
        QLabel#Toast {{ background: #3A3A3C; color: white; border-radius: 15px; padding: 11px; font-weight: 700; }}

        QPushButton {{ min-height: 40px; border-radius: 15px; padding: 9px 17px; font-weight: 700; }}
        QPushButton[kind="primary"] {{
            background: #3A3A3C; color: white; border: 1px solid #3A3A3C;
        }}
        QPushButton[kind="primary"]:hover {{ background: #2C2C2E; border-color: #2C2C2E; }}
        QPushButton[kind="primary"]:pressed {{ background: #1C1C1E; border-color: #1C1C1E; }}
        QPushButton[kind="secondary"] {{
            background: #E9E9ED; color: #1C1C1E; border: 1px solid #D8D8DC;
        }}
        QPushButton[kind="secondary"]:hover {{ background: #DEDEE3; border-color: #C7C7CC; }}
        QPushButton[kind="secondary"]:pressed {{ background: #CFCFD4; border-color: #B8B8BD; }}
        QPushButton[kind="ghost"] {{
            background: rgba(255,255,255,150); color: #3A3A3C; border: 1px solid #E1E1E6;
        }}
        QPushButton[kind="ghost"]:hover {{ background: #F0F0F3; color: #1C1C1E; }}
        QPushButton[kind="ghost"]:pressed {{ background: #E2E2E7; }}
        QPushButton[kind="danger"] {{ background: #FFF0EF; color: #D70015; border: 1px solid #FFD0CD; }}
        QPushButton[kind="danger"]:hover {{ background: #FFE4E1; border-color: #FFB8B2; }}
        QPushButton:disabled {{ background: #F0F0F2; color: #AEAEB2; border-color: #E4E4E8; }}

        QPushButton#RoundIconButton {{
            background: #E9E9ED; border: 1px solid #D8D8DC; border-radius: 17px; padding: 0px;
        }}
        QPushButton#RoundIconButton:hover {{ background: #DEDEE3; border: 1px solid #C7C7CC; }}
        QPushButton#RoundIconButton:pressed {{ background: #CFCFD4; border: 1px solid #B8B8BD; }}

        QLineEdit, QComboBox, QSpinBox, QPlainTextEdit, QListWidget, QTableWidget {{
            background: rgba(255,255,255,245);
            border: 1px solid #D8D8DC;
            border-radius: 15px; padding: 9px; color: #1C1C1E;
            selection-background-color: #D0E7FF; selection-color: #1C1C1E;
        }}
        QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QPlainTextEdit:focus, QListWidget:focus, QTableWidget:focus {{
            border: 2px solid #8AB8F5; background: #FFFFFF;
        }}
        QComboBox {{ padding-right: 42px; min-height: 38px; }}
        QComboBox::drop-down {{
            subcontrol-origin: padding; subcontrol-position: top right; width: 38px;
            border-left: 1px solid #E1E1E6; border-top-right-radius: 14px; border-bottom-right-radius: 14px;
            background: #F2F2F7;
        }}
        QComboBox::drop-down:hover {{ background: #E5E5EA; }}
        QComboBox::down-arrow {{ image: url("{down_url}"); width: 16px; height: 16px; }}
        QComboBox QAbstractItemView {{
            background: #FFFFFF; color: #1C1C1E; border: 1px solid #D1D1D6;
            selection-background-color: #E5E5EA; selection-color: #1C1C1E; outline: none;
        }}

        QSpinBox {{
            padding: 0px 58px 0px 16px;
            min-height: 52px;
        }}
        QSpinBox::up-button {{
            subcontrol-origin: border; subcontrol-position: top right;
            width: 50px; height: 26px;
            background: #F2F2F7;
            border-left: 1px solid #E1E1E6;
            border-bottom: 1px solid #E1E1E6;
            border-top-right-radius: 14px;
        }}
        QSpinBox::down-button {{
            subcontrol-origin: border; subcontrol-position: bottom right;
            width: 50px; height: 26px;
            background: #F2F2F7;
            border-left: 1px solid #E1E1E6;
            border-bottom-right-radius: 14px;
        }}
        QSpinBox::up-button:hover, QSpinBox::down-button:hover {{ background: #E5E5EA; }}
        QSpinBox::up-button:pressed, QSpinBox::down-button:pressed {{ background: #D1D1D6; }}
        QSpinBox::up-arrow {{ image: url("{up_url}"); width: 16px; height: 16px; }}
        QSpinBox::down-arrow {{ image: url("{down_url}"); width: 16px; height: 16px; }}

        /* Full-width fields in the narrow right timing panel. */
        QFrame#TimingField {{ background: transparent; border: none; }}
        QLabel#TimingFieldLabel {{
            color: #3A3A3C; font-size: 14px; font-weight: 650; padding: 0px 2px;
        }}
        QSpinBox#TimingFieldSpin {{
            padding: 0px 48px 0px 14px;
            min-height: 50px;
        }}
        QSpinBox#TimingFieldSpin::up-button {{ width: 40px; height: 25px; }}
        QSpinBox#TimingFieldSpin::down-button {{ width: 40px; height: 25px; }}

        /* Compact spin boxes used only inside the seven-column ISI matrix.
           The global 58 px padding + 50 px buttons hid the last digit and
           suffix when DPI scaling made a matrix cell narrow. */
        QTableWidget#TimingMatrixTable QSpinBox#TimingMatrixSpin {{
            margin: 4px 5px;
            padding: 0px 34px 0px 8px;
            min-height: 38px;
            border-radius: 10px;
        }}
        QTableWidget#TimingMatrixTable QSpinBox#TimingMatrixSpin::up-button {{
            width: 28px; height: 19px; border-top-right-radius: 9px;
        }}
        QTableWidget#TimingMatrixTable QSpinBox#TimingMatrixSpin::down-button {{
            width: 28px; height: 19px; border-bottom-right-radius: 9px;
        }}

        QTableWidget#StepsTable {{
            padding: 0px;
            border-radius: 16px;
        }}
        QTableWidget#StepsTable QComboBox,
        QTableWidget#StepsTable QSpinBox {{
            margin: 5px 7px;
            min-height: 42px;
        }}
        QTableWidget#StepsTable QSpinBox {{
            padding: 0px 52px 0px 13px;
        }}
        QTableWidget#StepsTable QSpinBox::up-button {{ width: 45px; height: 21px; }}
        QTableWidget#StepsTable QSpinBox::down-button {{ width: 45px; height: 21px; }}

        QCheckBox {{ spacing: 13px; color: #1C1C1E; font-weight: 700; min-height: 38px; }}
        QCheckBox::indicator {{ width: 56px; height: 34px; image: url("{toggle_off_url}"); }}
        QCheckBox::indicator:checked {{ image: url("{toggle_on_url}"); }}

        QListWidget {{ padding: 7px; }}
        QListWidget::item {{ padding: 12px 13px; margin: 3px; border: 1px solid transparent; border-radius: 13px; color: #48484A; }}
        QListWidget::item:hover {{ background: #EFEFF2; color: #1C1C1E; }}
        QListWidget::item:selected {{
            background: #3A3A3C; color: white; border: 1px solid #3A3A3C; font-weight: 780;
        }}

        QTabWidget#JamoSegmentedTabs::pane {{
            border: 1px solid #E1E1E6; border-radius: 20px;
            background: rgba(255,255,255,220); top: -1px;
        }}
        QTabWidget#JamoSegmentedTabs QTabBar::tab {{
            background: #ECECEF; color: #48484A;
            min-height: 32px; padding: 11px 22px; margin: 4px;
            border: 1px solid transparent; border-radius: 16px; font-size: 16px; font-weight: 700;
        }}
        QTabWidget#JamoSegmentedTabs QTabBar::tab:hover {{ background: #E1E1E5; color: #1C1C1E; }}
        QTabWidget#JamoSegmentedTabs QTabBar::tab:selected {{
            background: #3A3A3C; color: white; border: 1px solid #3A3A3C; font-weight: 820;
        }}

        QPushButton#PositionButton {{
            background: #EFEFF2; border: 1px solid #DEDEE3;
            border-radius: 21px; font-size: 20px; font-weight: 780; color: #48484A;
        }}
        QPushButton#PositionButton:hover {{ background: #E1E1E5; border-color: #C7C7CC; color: #1C1C1E; }}
        QPushButton#PositionButton:checked {{
            background: #3A3A3C; color: white; border: 2px solid #3A3A3C; font-weight: 860;
        }}
        QPushButton#HardwareButton {{
            background: #EFEFF2; border: 1px solid #DEDEE3;
            border-radius: 25px; font-size: 24px; font-weight: 790; color: #3A3A3C;
        }}
        QPushButton#HardwareButton:hover {{ background: #E1E1E5; border: 1px solid #C7C7CC; }}
        QPushButton#HardwareButton:pressed {{ background: #3A3A3C; color: white; border: 2px solid #3A3A3C; }}

        QHeaderView::section {{
            background: #F2F2F7; color: #48484A; border: none;
            border-bottom: 1px solid #DEDEE3; padding: 10px 9px; font-weight: 760;
            min-height: 28px;
        }}
        QScrollArea#TimingScroll,
        QScrollArea#TimingRulesScroll {{
            background: transparent;
            border: none;
        }}
        QScrollArea#TimingScroll > QWidget > QWidget,
        QScrollArea#TimingRulesScroll > QWidget > QWidget {{
            background: transparent;
        }}
        QWidget#TimingScrollBody,
        QWidget#TimingRulesScrollBody {{ background: transparent; }}
        QFrame#EstimatorCard {{
            background: #F6F6F8;
            border: 1px solid #E1E1E6;
            border-radius: 18px;
        }}
        QLabel#EstimatorTitle {{ color: #3A3A3C; font-size: 15px; font-weight: 760; }}
        QLabel#EstimatorValue {{ color: #1C1C1E; font-size: 22px; font-weight: 800; }}
        QLabel#EstimatorNote {{ color: #6E6E73; font-size: 12px; }}
        QTableCornerButton::section {{ background: #F2F2F7; border: none; }}
        QScrollBar:vertical {{ background: transparent; width: 10px; margin: 4px; }}
        QScrollBar::handle:vertical {{ background: #C7C7CC; border-radius: 5px; min-height: 32px; }}
        QScrollBar::handle:vertical:hover {{ background: #AEAEB2; }}
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0px; }}
        QScrollArea#GeneralQuizScroll {{ background: transparent; border: none; }}
        QScrollArea#GeneralQuizScroll > QWidget > QWidget {{ background: transparent; }}
        QWidget#GeneralQuizScrollBody {{ background: transparent; }}
        QPlainTextEdit#GeneralQuizChat {{
            background: #F6F6F8; border: 1px solid #E1E1E6; border-radius: 18px;
            padding: 16px; font-size: 16px; line-height: 1.45;
        }}
        QToolTip {{ background: #3A3A3C; color: white; border: none; border-radius: 8px; padding: 6px; }}
        """)
        for b in self.findChildren(QPushButton):
            kind = b.property("kind")
            if kind:
                b.style().unpolish(b); b.style().polish(b)

    # --------- general ---------
    def update_jamo_tab_state(self, index: int) -> None:
        """Keep labels stable; the selected segment is shown by fill and contrast."""
        if not hasattr(self, "jamo_tabs"):
            return
        self.jamo_tabs.setTabText(0, "자음")
        self.jamo_tabs.setTabText(1, "모음")

    def switch_page(self, index: int) -> None:
        self.stack.setCurrentIndex(index)
        for i, b in enumerate(self.nav_buttons): b.setChecked(i == index)

    def toast(self, text: str, error: bool = False) -> None:
        self.toast_label.setText(text)
        self.toast_label.setStyleSheet("background:#B3261E;color:white;border-radius:13px;padding:10px;font-weight:650;" if error else "")
        self.toast_label.show()
        QTimer.singleShot(2600, self.toast_label.hide)

    def mark_dirty(self) -> None:
        self.dirty = True
        self.update_setup_title()

    def update_setup_title(self) -> None:
        suffix = "  •  수정됨" if self.dirty else ""
        self.setup_title.setText(f"{self.setup.setup_name}{suffix}")

    def new_setup(self) -> None:
        if self.dirty and QMessageBox.question(self, "새 디자인", "저장하지 않은 변경사항을 버릴까요?") != QMessageBox.Yes:
            return
        self.setup = make_default_setup()
        self.setup.setup_name = "Untitled Hangul Design"
        self.setup_path = None
        self.dirty = True
        self.sync_setup_controls_from_model()
        self.refresh_jamo_lists()
        self.mapping_editor.load_label("ㄱ")
        self.update_setup_title()

    def save_setup(self, save_as: bool = False) -> None:
        self.sync_model_from_setup_controls()
        if save_as or self.setup_path is None:
            default = SETUP_DIR / (re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", self.setup.setup_name) + ".json")
            path, _ = QFileDialog.getSaveFileName(self, "디자인 저장", str(default), "JSON (*.json)")
            if not path:
                return
            self.setup_path = Path(path)
            self.setup.setup_name = self.setup_path.stem
        self.setup_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.setup_path, "w", encoding="utf-8") as f:
            json.dump(self.setup.to_dict(), f, ensure_ascii=False, indent=2)
        self.dirty = False
        self.update_setup_title()
        self.toast(f"저장됨 · {self.setup_path.name}")

    def load_setup_dialog(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "디자인 불러오기", str(SETUP_DIR), "JSON (*.json)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw_setup = json.load(f)
            previous_version = int(raw_setup.get("timing_semantics_version", 1) or 1)
            self.setup = DesignSetup.from_dict(raw_setup)
            self.setup_path = Path(path)
            self.dirty = previous_version < 6
            self.sync_setup_controls_from_model()
            self.refresh_jamo_lists()
            self.mapping_editor.load_label("ㄱ")
            self.update_setup_title()
            if previous_version < 6:
                self.toast("기존 SOA 셋업을 ISI로 변환해 불러왔습니다. 저장하면 v17 형식으로 저장됩니다.")
            else:
                self.toast(f"불러옴 · {self.setup_path.name}")
        except Exception as exc:
            QMessageBox.critical(self, "불러오기 실패", str(exc))

    def sync_setup_controls_from_model(self) -> None:
        defaults = self.setup.timing_defaults_ms
        for widget, value in [
            (self.default_comp_gap, defaults.get("composite", 150)),
            (self.cv_short_isi, self.setup.cv_short_isi_ms),
            (self.cv_long_isi, self.setup.cv_long_isi_ms),
            (self.cvc_cv_short_isi, self.setup.cvc_cv_short_isi_ms),
            (self.cvc_cv_long_isi, self.setup.cvc_cv_long_isi_ms),
            (self.cvc_vc_gap, defaults.get("cvc_vc", 250)),
            (self.comp_final_gap, defaults.get("compound_final", 150)),
            (self.syllable_gap, self.setup.inter_syllable_isi_ms),
            (self.word_gap, self.setup.inter_word_isi_ms),
        ]:
            widget.blockSignals(True); widget.setValue(int(value)); widget.blockSignals(False)
        self.left_marker_edit.setText(self.setup.left_marker)
        self.right_marker_edit.setText(self.setup.right_marker)
        self.baud_spin.setValue(self.setup.baudrate)
        self.motor_map_edit.setText(",".join(str(self.setup.logical_to_motor[str(i)]) for i in range(1, 10)))
        if hasattr(self, "mapping_editor"):
            self.mapping_editor.set_default_composite_gap(self.setup.timing_defaults_ms.get("composite", 150))
        self.update_duration_rule_summary()
        self.update_duration_estimate()

    def on_default_composite_gap_changed(self, value: int) -> None:
        """Update the global default and the editor's new-Step gap immediately."""
        if hasattr(self, "mapping_editor"):
            self.mapping_editor.set_default_composite_gap(int(value))
        self.sync_model_from_setup_controls()

    def sync_model_from_setup_controls(self) -> None:
        if not hasattr(self, "cv_gap"):
            return

        setup = self.setup
        setup.timing_semantics_version = 6
        setup.ensure_timing_matrix()

        old_cv_values = {
            ("cv", "short"): int(setup.cv_short_isi_ms),
            ("cv", "long"): int(setup.cv_long_isi_ms),
            ("cvc_cv", "short"): int(setup.cvc_cv_short_isi_ms),
            ("cvc_cv", "long"): int(setup.cvc_cv_long_isi_ms),
        }
        new_cv_values = {
            ("cv", "short"): int(self.cv_short_isi.value()),
            ("cv", "long"): int(self.cv_long_isi.value()),
            ("cvc_cv", "short"): int(self.cvc_cv_short_isi.value()),
            ("cvc_cv", "long"): int(self.cvc_cv_long_isi.value()),
        }

        # Non-CV basic defaults retain the earlier inherited-cell behavior:
        # cells that still equal the old default follow the new default, while
        # deliberately customized advanced cells remain untouched.
        non_cv_defaults = {
            "composite": int(self.default_comp_gap.value()),
            "cvc_vc": int(self.cvc_vc_gap.value()),
            "compound_final": int(self.comp_final_gap.value()),
        }
        for context, new_value in non_cv_defaults.items():
            old_value = int(setup.timing_defaults_ms.get(context, new_value))
            rows = setup.timing_matrix_ms.get(context, {})
            for cols in rows.values():
                for next_cls, cell_value in list(cols.items()):
                    if int(cell_value) == old_value:
                        cols[next_cls] = int(new_value)
            setup.timing_defaults_ms[context] = int(new_value)

        # The four CV controls are explicit row-wide controls. Only a control
        # that actually changed rewrites its linked row, preventing unrelated
        # settings from erasing advanced edits.
        for key, new_value in new_cv_values.items():
            if new_value != old_cv_values[key]:
                setup.set_uniform_timing_row(key[0], key[1], new_value)

        setup.cv_short_isi_ms = new_cv_values[("cv", "short")]
        setup.cv_long_isi_ms = new_cv_values[("cv", "long")]
        setup.cvc_cv_short_isi_ms = new_cv_values[("cvc_cv", "short")]
        setup.cvc_cv_long_isi_ms = new_cv_values[("cvc_cv", "long")]
        setup.timing_defaults_ms["cv"] = int(setup.cv_short_isi_ms)
        setup.timing_defaults_ms["cvc_cv"] = int(setup.cvc_cv_short_isi_ms)
        setup.use_duration_based_cv_isi = True
        setup.inter_syllable_isi_ms = self.syllable_gap.value()
        setup.inter_word_isi_ms = self.word_gap.value()

        # Keep deprecated fields synchronized only for human readability in old tools.
        setup.default_composite_gap_ms = self.default_comp_gap.value()
        setup.cv_gap_ms = self.cv_gap.value()
        setup.cvc_cv_gap_ms = self.cvc_cv_gap.value()
        setup.cvc_vc_gap_ms = self.cvc_vc_gap.value()
        setup.compound_final_gap_ms = self.comp_final_gap.value()
        setup.inter_syllable_gap_ms = setup.inter_syllable_isi_ms
        setup.inter_word_gap_ms = setup.inter_word_isi_ms
        self.mark_dirty()
        self.update_duration_estimate()
        if hasattr(self, "mapping_editor"):
            self.mapping_editor.preview_current(show_error=False)

    def _speed_estimate_syllables(self) -> Tuple[List[str], str]:
        """Return syllables used for the live speed estimate and a source label."""
        candidate_paths: List[Path] = []
        if hasattr(self, "syllable_edit"):
            text = self.syllable_edit.text().strip()
            if text:
                candidate_paths.append(Path(text))
        candidate_paths.extend([
            APP_DIR / "syllable_top200.xlsx",
            APP_DIR / "syllable.xlsx",
        ])
        seen: set[str] = set()
        for path in candidate_paths:
            key = str(path.resolve()) if path.exists() else str(path)
            if key in seen:
                continue
            seen.add(key)
            if path.exists():
                try:
                    labels = load_syllables(path, self.syllable_limit.value() if hasattr(self, "syllable_limit") else 200)
                    if labels:
                        return labels, path.name
                except Exception:
                    pass

        # File-independent fallback: a broad set of CV and frequent-final CVC
        # syllables. This keeps the estimate available even before an XLSX is copied.
        labels: List[str] = []
        common_finals = ["", "ㄱ", "ㄴ", "ㄹ", "ㅁ", "ㅂ", "ㅅ", "ㅇ"]
        jong_index = {j: i for i, j in enumerate(JONGSUNG)}
        for cho_idx in range(len(CHOSUNG)):
            for jung_idx in range(len(JUNGSUNG)):
                for jong in common_finals:
                    code = 0xAC00 + (cho_idx * 21 + jung_idx) * 28 + jong_index[jong]
                    labels.append(chr(code))
        return labels, "내장 CV/CVC 표본"

    def update_duration_estimate(self, *_args) -> None:
        if not hasattr(self, "duration_estimate_value"):
            return
        compiler = HangulCompiler(self.setup)
        labels, source = self._speed_estimate_syllables()
        valid: List[str] = []
        syllable_durations: List[int] = []
        for label in labels:
            try:
                command = compiler.compile_syllable(label)
                syllable_durations.append(HangulCompiler.estimate_duration_ms(command))
                valid.append(label)
            except Exception:
                continue
        if not valid:
            self.duration_estimate_value.setText("계산 불가")
            self.duration_estimate_note.setText("컴파일 가능한 음절이 없습니다. 자모 매핑을 확인하세요.")
            return

        word_len = self.word_length_spin.value() if hasattr(self, "word_length_spin") else 2
        sample_count = min(120, max(24, len(valid)))
        word_durations: List[int] = []
        word_onset_cycles: List[int] = []
        n = len(valid)
        for i in range(sample_count):
            word = "".join(valid[(i + j * 37) % n] for j in range(word_len))
            next_word = "".join(valid[(i + 19 + j * 53) % n] for j in range(word_len))
            try:
                command, _trace = compiler.compile_text(word)
                word_durations.append(HangulCompiler.estimate_duration_ms(command))

                _pair_command, pair_trace = compiler.compile_text(word + " " + next_word)
                if len(pair_trace) > word_len:
                    word_onset_cycles.append(int(pair_trace[word_len]["onset_ms"]))
            except Exception:
                continue

        if not word_durations:
            self.duration_estimate_value.setText("계산 불가")
            self.duration_estimate_note.setText("현재 ISI 규칙으로 단어를 컴파일할 수 없습니다.")
            return

        avg_syllable_ms = sum(syllable_durations) / len(syllable_durations)
        avg_word_ms = sum(word_durations) / len(word_durations)
        avg_cycle_ms = (sum(word_onset_cycles) / len(word_onset_cycles)) if word_onset_cycles else avg_word_ms
        words_per_min = 60000.0 / avg_cycle_ms if avg_cycle_ms > 0 else 0.0
        self.duration_estimate_value.setText(
            f"{word_len}음절 단어 평균  {avg_word_ms / 1000.0:.2f}초"
        )
        self.duration_estimate_note.setText(
            f"평균 음절 물리 duration {avg_syllable_ms / 1000.0:.2f}초 · "
            f"단어 onset 주기 {avg_cycle_ms / 1000.0:.2f}초 · 약 {words_per_min:.1f}단어/분\n"
            f"모든 내부·경계 ISI 반영 · 기준: {source} · 컴파일 성공 {len(valid)}/{len(labels)}개"
        )

    def update_duration_rule_summary(self) -> None:
        if hasattr(self, "duration_rule_summary"):
            self.duration_rule_summary.setText(
                f"현재 분기 기준 · 자음 {int(self.setup.cv_duration_split_ms)} ms · "
                f"모션 {int(self.setup.motion_duration_split_ms)} ms "
                "(기준 이하 = 짧음) · CV 짧음/김 행은 고급 행렬과 동일"
            )

    def open_timing_rules(self) -> None:
        self.sync_model_from_setup_controls()
        dialog = TimingRulesDialog(self)
        dialog.exec()
        self.update_duration_rule_summary()

    def apply_hardware_settings(self) -> None:
        try:
            vals = [int(x.strip()) for x in self.motor_map_edit.text().split(",")]
            if len(vals) != 9 or any(v < 1 or v > 9 for v in vals):
                raise ValueError("motor mapping은 1~9 숫자 9개여야 합니다.")
            lm, rm = self.left_marker_edit.text().strip(), self.right_marker_edit.text().strip()
            if lm not in ("@", "#") or rm not in ("@", "#") or lm == rm:
                raise ValueError("marker는 @와 #를 서로 다르게 지정해야 합니다.")
            self.setup.left_marker = lm; self.setup.right_marker = rm
            self.setup.baudrate = self.baud_spin.value()
            self.setup.logical_to_motor = {str(i + 1): vals[i] for i in range(9)}
            self.mark_dirty(); self.mapping_editor.preview_current(show_error=False)
            self.toast("하드웨어 mapping 적용됨")
        except Exception as exc:
            QMessageBox.warning(self, "Mapping 오류", str(exc))

    def refresh_jamo_lists(self) -> None:
        if not hasattr(self, "consonant_list"):
            return
        def fill(widget: QListWidget, labels: Sequence[str]) -> None:
            current = widget.currentItem().data(Qt.UserRole) if widget.currentItem() else None
            widget.clear()
            for lab in labels:
                spec = self.setup.get_spec(lab)
                if spec.mode == "atomic":
                    detail = f"{('L' if spec.arm == 'left' else 'R')}{spec.position} · {spec.temporal}"
                elif spec.mode == "composite":
                    detail = " + ".join(str(x.get("ref", "?")) for x in spec.steps) or "empty"
                elif spec.mode == "raw": detail = "RAW"
                else: detail = "미매핑"
                item = QListWidgetItem(f"{lab}     {detail}")
                item.setData(Qt.UserRole, lab)
                widget.addItem(item)
                if lab == current: widget.setCurrentItem(item)
        fill(self.consonant_list, CONSONANTS)
        fill(self.vowel_list, VOWELS)

    def validate_design(self) -> None:
        compiler = HangulCompiler(self.setup)
        errors: List[str] = []
        for lab in CONSONANTS + VOWELS:
            try:
                compiler.compile_jamo(lab)
            except Exception as exc:
                errors.append(f"{lab}: {exc}")
        if errors:
            QMessageBox.warning(self, "검증 결과", f"오류 {len(errors)}개\n\n" + "\n".join(errors[:30]))
        else:
            QMessageBox.information(self, "검증 결과", "모든 자음과 모음이 정상적으로 컴파일됩니다.")

    # --------- serial ---------
    def refresh_ports(self) -> None:
        if not hasattr(self, "port_combo"):
            return
        self.port_combo.clear()
        ports = self.serial_ctl.list_ports()
        for dev, desc in ports:
            self.port_combo.addItem(f"{dev} · {desc}", dev)
        if not ports:
            self.port_combo.addItem("포트 없음 · DRY RUN", "")

    def toggle_serial(self) -> None:
        if self.serial_ctl.is_connected():
            self.serial_ctl.close(); self.connect_btn.setText("연결"); self.sidebar_status.setText("DRY RUN"); return
        port = self.port_combo.currentData()
        if not port:
            QMessageBox.warning(self, "COM", "COM 포트를 선택하세요."); return
        try:
            self.apply_hardware_settings()
            self.serial_ctl.connect(port, self.setup.baudrate)
            self.connect_btn.setText("연결 해제")
            self.sidebar_status.setText(f"CONNECTED · {port}")
            self.toast("Arduino 연결됨")
        except Exception as exc:
            QMessageBox.critical(self, "연결 실패", str(exc))

    def send_command(self, command: str) -> str:
        if len(command) > 560:
            raise RuntimeError(f"Command가 Arduino 수신 buffer에 비해 너무 깁니다 ({len(command)} chars). 입력을 줄여주세요.")
        wait_ack = self.ack_check.isChecked() if hasattr(self, "ack_check") else True
        ack = self.serial_ctl.send(command, wait_ack=wait_ack)
        self.toast(f"자극 전송 · {ack}")
        return ack

    def emergency_off(self) -> None:
        self.serial_ctl.emergency_off(); self.toast("모든 모터 OFF")

    def test_hardware_position(self, position: int) -> None:
        spec = JamoSpec(mode="atomic", arm=self.hw_arm.currentData(), position=position, temporal=self.hw_temporal.currentData())
        try:
            self.send_command(HangulCompiler.serialize_timeline(HangulCompiler(self.setup).atomic_timeline(spec)))
        except Exception as exc:
            QMessageBox.critical(self, "Hardware test", str(exc))

    # --------- simple test ---------
    def compile_simple(self) -> Optional[str]:
        self.sync_model_from_setup_controls()
        try:
            command, trace = HangulCompiler(self.setup).compile_text(self.simple_input.text())
            duration = HangulCompiler.estimate_duration_ms(command)
            self.simple_command.setPlainText(
                command + "\n\n" +
                "\n".join(f"t={x['onset_ms']} ms · {x['char']}  →  {x['command']}" for x in trace)
            )
            self.simple_summary.setText(f"{len(trace)}글자 · 예상 {duration} ms · command {len(command)} chars")
            return command
        except Exception as exc:
            self.simple_command.setPlainText(f"컴파일 오류: {exc}")
            self.simple_summary.setText("컴파일 실패")
            self.toast(str(exc), error=True)
            return None

    def play_simple(self) -> None:
        command = self.compile_simple()
        if command:
            try: self.send_command(command)
            except Exception as exc: QMessageBox.critical(self, "자극 오류", str(exc))


    # --------- selected jamo learning ---------
    def _selected_learning_groups(self) -> Dict[str, List[str]]:
        groups: Dict[str, List[str]] = {}
        if self.learning_basic_consonant_check.isChecked():
            groups["기본 자음"] = list(BASIC_CONSONANTS)
        if self.learning_double_consonant_check.isChecked():
            groups["쌍자음"] = list(DOUBLE_CONSONANTS)
        if self.learning_basic_vowel_check.isChecked():
            groups["기본 모음"] = list(BASIC_VOWELS)
        if self.learning_compound_vowel_check.isChecked():
            groups["복합 모음"] = list(DIPHTHONG_VOWELS)
        if not groups:
            raise RuntimeError("학습할 자모 그룹을 하나 이상 선택하세요.")
        return groups

    def _build_learning_candidates(
        self,
        groups: Dict[str, List[str]],
    ) -> Dict[str, List[Dict[str, Any]]]:
        compiler = HangulCompiler(self.setup)
        candidates: Dict[str, List[Dict[str, Any]]] = {}
        compile_errors: List[str] = []
        for group_name, labels in groups.items():
            items: List[Dict[str, Any]] = []
            for label in labels:
                try:
                    command = compiler.compile_jamo(label)
                    if command:
                        items.append({
                            "group": group_name,
                            "target": label,
                            "command": command,
                        })
                except Exception as exc:
                    compile_errors.append(f"{label}: {exc}")
            candidates[group_name] = items

        empty_groups = [name for name, items in candidates.items() if not items]
        if empty_groups:
            detail = "\n".join(compile_errors[:12])
            message = "현재 디자인으로 학습 가능한 자모가 없는 그룹: " + ", ".join(empty_groups)
            if detail:
                message += "\n\n컴파일 오류 예시:\n" + detail
            raise RuntimeError(message)
        return candidates

    @staticmethod
    def _make_learning_plan(
        candidates: Dict[str, List[Dict[str, Any]]],
        trial_count: int,
    ) -> List[Dict[str, Any]]:
        available_groups = [name for name, items in candidates.items() if items]
        if not available_groups:
            return []
        rng = random.Random(time.time_ns())
        bags: Dict[str, List[Dict[str, Any]]] = {name: [] for name in available_groups}
        plan: List[Dict[str, Any]] = []

        while len(plan) < int(trial_count):
            group_cycle = list(available_groups)
            rng.shuffle(group_cycle)
            for group_name in group_cycle:
                if len(plan) >= int(trial_count):
                    break
                if not bags[group_name]:
                    bags[group_name] = [dict(item) for item in candidates[group_name]]
                    rng.shuffle(bags[group_name])
                item = bags[group_name].pop()
                if plan and item["target"] == plan[-1]["target"] and bags[group_name]:
                    alternate = bags[group_name].pop()
                    bags[group_name].append(item)
                    item = alternate
                plan.append(item)
        return plan

    def _learning_all_labels(self) -> List[str]:
        labels: List[str] = []
        for group_items in getattr(self, "_learning_candidates", {}).values():
            labels.extend(str(item["target"]) for item in group_items)
        return list(dict.fromkeys(labels))

    def _learning_choices_for_target(self, target: str, count: int) -> List[str]:
        pool = self._learning_all_labels()
        if target not in pool:
            pool.append(target)
        choice_count = min(max(2, int(count)), len(pool))
        distractors = [label for label in pool if label != target]
        rng = random.Random(time.time_ns() ^ (self.learning_index + 1))
        rng.shuffle(distractors)
        choices = [target] + distractors[:max(0, choice_count - 1)]
        rng.shuffle(choices)
        return choices

    def _render_learning_choices(self, choices: Sequence[str]) -> None:
        clear_layout(self.learning_choice_layout)
        self.learning_choice_buttons = []
        count = len(choices)
        columns = 2 if count <= 4 else 4
        for index, label in enumerate(choices):
            choice_btn = button(str(label), "secondary")
            choice_btn.setObjectName("LearningChoiceButton")
            choice_btn.setMinimumHeight(78)
            choice_btn.setProperty("learning_label", str(label))
            choice_btn.clicked.connect(
                lambda _checked=False, selected=str(label): self.submit_learning_choice(selected)
            )
            self.learning_choice_layout.addWidget(choice_btn, index // columns, index % columns)
            self.learning_choice_buttons.append(choice_btn)

    def start_learning_session(self) -> None:
        self.sync_model_from_setup_controls()
        try:
            groups = self._selected_learning_groups()
            self.learning_feedback.setStyleSheet("")
            self.learning_feedback.setText("선택한 자모를 컴파일하는 중…")
            QApplication.processEvents()

            candidates = self._build_learning_candidates(groups)
            total_labels = sum(len(items) for items in candidates.values())
            if total_labels < 2:
                raise RuntimeError("보기 학습을 위해 컴파일 가능한 자모가 최소 2개 필요합니다.")

            plan = self._make_learning_plan(candidates, self.learning_trials.value())
            if not plan:
                raise RuntimeError("학습 문제를 만들 수 없습니다.")

            self._learning_candidates = candidates
            self._learning_selected_group_names = list(groups)
            self.learning_plan = plan
            self.learning_index = -1
            self.learning_rows = []
            self.learning_target = ""
            self.learning_command = ""
            self.learning_answered = False
            self.learning_replays = 0

            session = re.sub(
                r"[^0-9A-Za-z가-힣_-]+",
                "_",
                self.learning_session_edit.text().strip() or "learning",
            ).strip("_") or "learning"
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.learning_csv_path = DATA_DIR / f"jamo_learning_{session}_{stamp}.csv"

            self.learning_feedback.setText(
                f"{', '.join(groups)} · 총 {total_labels}개 자모 · "
                f"보기 최대 {min(self.learning_choice_count.value(), total_labels)}개"
            )
            self.next_learning_trial()
        except Exception as exc:
            self.learning_feedback.setStyleSheet("")
            self.learning_feedback.setText("")
            QMessageBox.critical(self, "선택 학습 시작 실패", str(exc))

    def next_learning_trial(self) -> None:
        self.learning_index += 1
        if self.learning_index >= len(self.learning_plan):
            total = len(self.learning_rows)
            correct = sum(int(row.get("correct", 0)) for row in self.learning_rows)
            accuracy = (100.0 * correct / total) if total else 0.0
            self.learning_progress.setText("LEARNING COMPLETE")
            self.learning_instruction.setText("선택 자모 학습 완료")
            self.learning_feedback.setStyleSheet(
                "background:#EAF8EF;color:#147A3D;border-radius:16px;padding:12px;font-weight:700;"
            )
            self.learning_feedback.setText(
                f"정답 {correct}/{total} · Accuracy {accuracy:.1f}% · 결과: {self.learning_csv_path}"
            )
            clear_layout(self.learning_choice_layout)
            self.learning_choice_buttons = []
            self.learning_replay_btn.setEnabled(False)
            self.learning_next_btn.setEnabled(False)
            return

        item = self.learning_plan[self.learning_index]
        self.learning_target = str(item["target"])
        self.learning_command = str(item["command"])
        self.learning_answered = False
        self.learning_replays = 0
        self.learning_started_perf = 0.0

        choices = self._learning_choices_for_target(
            self.learning_target,
            self.learning_choice_count.value(),
        )
        self._current_learning_choices = list(choices)
        self._render_learning_choices(choices)

        self.learning_progress.setText(
            f"TRIAL {self.learning_index + 1} / {len(self.learning_plan)} · {item['group']}"
        )
        self.learning_instruction.setText("촉각 자극을 느끼고 해당하는 자모 보기를 누르세요.")
        self.learning_feedback.setStyleSheet("")
        self.learning_feedback.setText("자극을 재생합니다…")
        self.learning_replay_btn.setEnabled(True)
        self.learning_next_btn.setEnabled(False)
        QTimer.singleShot(80, lambda: self.play_learning_stimulus(replay=False))

    def play_learning_stimulus(self, replay: bool = True) -> None:
        if not self.learning_command or self.learning_answered:
            return
        try:
            if replay:
                self.learning_replays += 1
            self.send_command(self.learning_command)
            self.learning_started_perf = time.perf_counter()
            self.learning_feedback.setText("자극 재생 완료 · 보기에서 답을 선택하세요.")
        except Exception as exc:
            self.learning_feedback.setText(f"자극 재생 실패: {exc}")

    def submit_learning_choice(self, selected: str) -> None:
        if self.learning_answered or not self.learning_target:
            return
        self.learning_answered = True
        elapsed_ms: Any = ""
        if self.learning_started_perf:
            elapsed_ms = round((time.perf_counter() - self.learning_started_perf) * 1000.0, 1)

        correct = int(str(selected) == self.learning_target)
        for choice_btn in self.learning_choice_buttons:
            label = str(choice_btn.property("learning_label") or choice_btn.text())
            choice_btn.setCursor(Qt.ArrowCursor)
            if label == self.learning_target:
                choice_btn.setProperty("kind", "primary")
            elif label == str(selected) and not correct:
                choice_btn.setProperty("kind", "danger")
            choice_btn.style().unpolish(choice_btn)
            choice_btn.style().polish(choice_btn)

        if correct:
            message = f"정답입니다 ✓  {self.learning_target}"
            if elapsed_ms != "":
                message += f" · {elapsed_ms} ms"
            self.learning_feedback.setStyleSheet(
                "background:#EAF8EF;color:#147A3D;border-radius:16px;padding:12px;font-weight:700;"
            )
        else:
            message = f"오답입니다. 선택 {selected} · 정답 {self.learning_target}"
            self.learning_feedback.setStyleSheet(
                "background:#FFF0F0;color:#B3261E;border-radius:16px;padding:12px;font-weight:700;"
            )
        self.learning_feedback.setText(message)

        item = self.learning_plan[self.learning_index]
        selected_groups = list(getattr(self, "_learning_selected_group_names", []))
        row = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "app_version": APP_VERSION,
            "session_id": self.learning_session_edit.text().strip(),
            "setup_name": self.setup.setup_name,
            "setup_hash": self.setup.stable_hash(),
            "setup_path": str(self.setup_path or ""),
            "trial_index": self.learning_index + 1,
            "target_group": item["group"],
            "target": self.learning_target,
            "response": str(selected),
            "correct": correct,
            "elapsed_ms_from_last_playback": elapsed_ms,
            "replay_count": self.learning_replays,
            "choice_count": len(getattr(self, "_current_learning_choices", [])),
            "choices": "|".join(getattr(self, "_current_learning_choices", [])),
            "selected_groups": "|".join(selected_groups),
            "command": self.learning_command,
        }
        self.learning_rows.append(row)
        self._append_learning_csv(row)
        self.learning_replay_btn.setEnabled(False)
        self.learning_next_btn.setEnabled(True)
        self.learning_next_btn.setFocus()

    def _append_learning_csv(self, row: Dict[str, Any]) -> None:
        if self.learning_csv_path is None:
            return
        exists = self.learning_csv_path.exists()
        with open(self.learning_csv_path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            if not exists:
                writer.writeheader()
            writer.writerow(row)

    # --------- general text-answer quiz ---------
    def _general_quiz_options(self) -> Tuple[List[str], List[str], List[str], List[str]]:
        structures: List[str] = []
        if self.gq_cv_check.isChecked():
            structures.append("CV")
        if self.gq_cvc_check.isChecked():
            structures.append("CVC")
        if not structures:
            raise RuntimeError("CV 또는 CVC 구조를 하나 이상 선택하세요.")

        initials: List[str] = []
        if self.gq_basic_consonant_check.isChecked():
            initials.extend(BASIC_CONSONANTS)
        if self.gq_double_consonant_check.isChecked():
            initials.extend(DOUBLE_CONSONANTS)
        initials = list(dict.fromkeys(initials))
        if not initials:
            raise RuntimeError("기본 자음 또는 쌍자음을 하나 이상 선택하세요.")

        vowels: List[str] = []
        if self.gq_basic_vowel_check.isChecked():
            vowels.extend(BASIC_VOWELS)
        if self.gq_diphthong_check.isChecked():
            vowels.extend(DIPHTHONG_VOWELS)
        vowels = list(dict.fromkeys(vowels))
        if not vowels:
            raise RuntimeError("기본 모음 또는 이중모음을 하나 이상 선택하세요.")

        finals = [lab for lab in initials if lab in SINGLE_FINAL_CONSONANTS]
        if "CVC" in structures and not finals:
            raise RuntimeError("선택한 자음 중 CVC 종성으로 사용할 수 있는 자음이 없습니다.")
        return structures, initials, vowels, finals

    def _build_general_quiz_candidates(
        self,
        source_syllables: Sequence[str],
        structures: Sequence[str],
        initials: Sequence[str],
        vowels: Sequence[str],
        finals: Sequence[str],
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Filter the XLSX inventory; never synthesize an out-of-file syllable."""
        compiler = HangulCompiler(self.setup)
        candidates: Dict[str, List[Dict[str, Any]]] = {key: [] for key in structures}
        initial_set = set(initials)
        vowel_set = set(vowels)
        final_set = set(finals)
        seen: set[str] = set()

        for source_rank, syllable in enumerate(source_syllables, start=1):
            if syllable in seen:
                continue
            seen.add(syllable)
            dec = compiler.decompose_syllable(syllable)
            if dec is None:
                continue
            cho, jung, jong = dec
            structure = "CVC" if jong else "CV"
            if structure not in candidates:
                continue
            if cho not in initial_set or jung not in vowel_set:
                continue
            if structure == "CVC" and jong not in final_set:
                continue
            try:
                command = compiler.compile_syllable(syllable)
            except Exception:
                continue
            candidates[structure].append({
                "syllable": syllable,
                "structure": structure,
                "cho": cho,
                "jung": jung,
                "jong": jong,
                "command": command,
                "source_rank": source_rank,
            })
        return candidates

    @staticmethod
    def _make_general_quiz_plan(
        candidates: Dict[str, List[Dict[str, Any]]],
        trial_count: int,
    ) -> List[Dict[str, Any]]:
        available = [key for key, values in candidates.items() if values]
        if not available:
            return []
        rng = random.Random(time.time_ns())
        bags: Dict[str, List[Dict[str, Any]]] = {key: [] for key in available}
        plan: List[Dict[str, Any]] = []
        while len(plan) < int(trial_count):
            cycle = list(available)
            rng.shuffle(cycle)
            for structure in cycle:
                if len(plan) >= int(trial_count):
                    break
                if not bags[structure]:
                    bags[structure] = [dict(item) for item in candidates[structure]]
                    rng.shuffle(bags[structure])
                item = bags[structure].pop()
                if plan and item["syllable"] == plan[-1]["syllable"] and bags[structure]:
                    alternate = bags[structure].pop()
                    bags[structure].append(item)
                    item = alternate
                plan.append(item)
        return plan

    def _append_general_quiz_chat(self, speaker: str, text: str) -> None:
        if not hasattr(self, "gq_chat"):
            return
        current = self.gq_chat.toPlainText().rstrip()
        block = f"{speaker}\n{text}"
        self.gq_chat.setPlainText((current + "\n\n" + block).strip())
        bar = self.gq_chat.verticalScrollBar()
        bar.setValue(bar.maximum())

    def start_general_quiz(self) -> None:
        self.sync_model_from_setup_controls()
        try:
            structures, initials, vowels, finals = self._general_quiz_options()
            self.gq_feedback.setStyleSheet("")
            self.gq_feedback.setText("Top200 XLSX를 읽고 현재 조건에 맞는 음절을 확인하는 중…")
            QApplication.processEvents()
            source_path = Path(self.gq_syllable_edit.text().strip())
            source_syllables = load_general_quiz_syllables(source_path)
            if not source_syllables:
                raise RuntimeError(
                    "선택한 XLSX에서 현대 한글 음절을 찾지 못했습니다. "
                    "음절/글자/단어 열 또는 첫 번째 열을 확인하세요."
                )
            candidates = self._build_general_quiz_candidates(
                source_syllables, structures, initials, vowels, finals
            )
            missing = [key for key in structures if not candidates.get(key)]
            if missing:
                raise RuntimeError(
                    "syllable_top200.xlsx 안에서 현재 자모 선택과 ISI 설정을 모두 만족하는 "
                    + ", ".join(missing)
                    + " 음절이 없습니다. 구조/자모 체크 또는 XLSX 내용을 확인하세요."
                )
            plan = self._make_general_quiz_plan(candidates, self.gq_trials.value())
            if not plan:
                raise RuntimeError("퀴즈 문제를 만들 수 없습니다.")

            self.general_quiz_plan = plan
            self.general_quiz_index = -1
            self.general_quiz_rows = []
            self.general_quiz_target = ""
            self.general_quiz_command = ""
            self.general_quiz_answered = False
            self.general_quiz_replays = 0
            session = re.sub(
                r"[^0-9A-Za-z가-힣_-]+",
                "_",
                self.gq_session_edit.text().strip() or "practice",
            ).strip("_") or "practice"
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.general_quiz_csv_path = DATA_DIR / f"general_quiz_{session}_{stamp}.csv"

            self.gq_chat.clear()
            self._append_general_quiz_chat(
                "시스템",
                f"Top200 기반 일반 촉각 퀴즈를 시작합니다. 구조 {', '.join(structures)} · "
                f"XLSX 음절 {len(source_syllables)}개 · 조건 통과 후보 "
                f"{sum(len(values) for values in candidates.values())}개 · 문제 {len(plan)}개 · "
                f"파일 {source_path.name}",
            )
            self.gq_feedback.setText(
                "자극을 느낀 뒤 한글 음절 하나를 입력하세요. 완성형 음절과 ㄱㅏ 같은 자모 입력을 모두 지원합니다."
            )
            self.next_general_quiz_trial()
        except Exception as exc:
            self.gq_feedback.setStyleSheet("")
            self.gq_feedback.setText("")
            QMessageBox.critical(self, "일반 퀴즈 시작 실패", str(exc))

    def next_general_quiz_trial(self) -> None:
        self.general_quiz_index += 1
        if self.general_quiz_index >= len(self.general_quiz_plan):
            total = len(self.general_quiz_rows)
            correct = sum(int(row.get("correct", 0)) for row in self.general_quiz_rows)
            accuracy = (100.0 * correct / total) if total else 0.0
            self.gq_progress.setText("QUIZ COMPLETE")
            self.gq_instruction.setText("일반 촉각 퀴즈 완료")
            self.gq_feedback.setStyleSheet(
                "background:#EAF8EF;color:#147A3D;border-radius:16px;padding:12px;font-weight:700;"
            )
            self.gq_feedback.setText(
                f"정답 {correct}/{total} · Accuracy {accuracy:.1f}% · "
                f"결과: {self.general_quiz_csv_path}"
            )
            self._append_general_quiz_chat(
                "시스템",
                f"퀴즈가 끝났습니다. 정답 {correct}/{total}, 정확도 {accuracy:.1f}%입니다.",
            )
            self.gq_answer.setEnabled(False)
            self.gq_submit_btn.setEnabled(False)
            self.gq_replay_btn.setEnabled(False)
            self.gq_next_btn.setEnabled(False)
            return

        item = self.general_quiz_plan[self.general_quiz_index]
        self.general_quiz_target = str(item["syllable"])
        self.general_quiz_command = str(item["command"])
        self.general_quiz_answered = False
        self.general_quiz_replays = 0
        self.general_quiz_started_perf = 0.0
        self.gq_progress.setText(
            f"TRIAL {self.general_quiz_index + 1} / {len(self.general_quiz_plan)} · {item['structure']}"
        )
        self.gq_instruction.setText("촉각 자극을 해석하고 채팅 입력창에 음절을 입력하세요.")
        self.gq_feedback.setStyleSheet("")
        self.gq_feedback.setText("자극을 재생합니다…")
        self.gq_answer.clear()
        self.gq_answer.setEnabled(True)
        self.gq_submit_btn.setEnabled(True)
        self.gq_replay_btn.setEnabled(True)
        self.gq_next_btn.setEnabled(False)
        self._append_general_quiz_chat(
            "시스템",
            f"문제 {self.general_quiz_index + 1}: 촉각 자극을 재생합니다. 느낀 음절을 입력하세요.",
        )
        QTimer.singleShot(80, lambda: self.play_general_quiz_stimulus(replay=False))

    def play_general_quiz_stimulus(self, replay: bool = True) -> None:
        if not self.general_quiz_command or self.general_quiz_answered:
            return
        try:
            if replay:
                self.general_quiz_replays += 1
                self._append_general_quiz_chat("시스템", "촉각 자극을 다시 재생했습니다.")
            self.send_command(self.general_quiz_command)
            self.general_quiz_started_perf = time.perf_counter()
            self.gq_feedback.setText("자극 재생 완료 · 답변을 입력하세요.")
            self.gq_answer.setFocus()
        except Exception as exc:
            self.gq_feedback.setText(f"자극 재생 실패: {exc}")

    @staticmethod
    def _normalize_general_quiz_answer(text: str) -> str:
        compact = re.sub(r"\s+", "", str(text))
        if len(compact) == 1 and 0xAC00 <= ord(compact) <= 0xD7A3:
            return compact
        if len(compact) in (2, 3):
            cho, jung = compact[0], compact[1]
            jong = compact[2] if len(compact) == 3 else ""
            if cho in CHOSUNG and jung in JUNGSUNG and jong in JONGSUNG:
                return compose_hangul_syllable(cho, jung, jong)
        return compact

    def submit_general_quiz_answer(self) -> None:
        if self.general_quiz_answered or not self.general_quiz_target:
            return
        raw_answer = self.gq_answer.text().strip()
        if not raw_answer:
            self.gq_feedback.setText("답변을 입력하세요.")
            return
        answer = self._normalize_general_quiz_answer(raw_answer)
        if len(answer) != 1 or not (0xAC00 <= ord(answer) <= 0xD7A3):
            self.gq_feedback.setText("한글 음절 하나를 입력하세요. 예: 가 또는 ㄱㅏ")
            return

        self.general_quiz_answered = True
        elapsed_ms: Any = ""
        if self.general_quiz_started_perf:
            elapsed_ms = round((time.perf_counter() - self.general_quiz_started_perf) * 1000.0, 1)
        item = self.general_quiz_plan[self.general_quiz_index]
        correct = int(answer == self.general_quiz_target)
        self._append_general_quiz_chat("나", answer)
        if correct:
            response_text = f"정답입니다 ✓  ({elapsed_ms} ms)" if elapsed_ms != "" else "정답입니다 ✓"
            self.gq_feedback.setStyleSheet(
                "background:#EAF8EF;color:#147A3D;border-radius:16px;padding:12px;font-weight:700;"
            )
        else:
            response_text = f"오답입니다. 정답은 {self.general_quiz_target}입니다."
            self.gq_feedback.setStyleSheet(
                "background:#FFF0F0;color:#B3261E;border-radius:16px;padding:12px;font-weight:700;"
            )
        self._append_general_quiz_chat("시스템", response_text)
        self.gq_feedback.setText(response_text)

        row = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "app_version": APP_VERSION,
            "session_id": self.gq_session_edit.text().strip(),
            "setup_name": self.setup.setup_name,
            "setup_hash": self.setup.stable_hash(),
            "setup_path": str(self.setup_path or ""),
            "trial_index": self.general_quiz_index + 1,
            "structure": item["structure"],
            "target": self.general_quiz_target,
            "target_cho": item["cho"],
            "target_jung": item["jung"],
            "target_jong": item["jong"],
            "source_xlsx": self.gq_syllable_edit.text().strip(),
            "source_rank": item.get("source_rank", ""),
            "response_raw": raw_answer,
            "response": answer,
            "correct": correct,
            "elapsed_ms_from_last_playback": elapsed_ms,
            "replay_count": self.general_quiz_replays,
            "command": self.general_quiz_command,
            "include_basic_consonants": int(self.gq_basic_consonant_check.isChecked()),
            "include_double_consonants": int(self.gq_double_consonant_check.isChecked()),
            "include_basic_vowels": int(self.gq_basic_vowel_check.isChecked()),
            "include_diphthongs": int(self.gq_diphthong_check.isChecked()),
        }
        self.general_quiz_rows.append(row)
        self._append_general_quiz_csv(row)
        self.gq_answer.setEnabled(False)
        self.gq_submit_btn.setEnabled(False)
        self.gq_replay_btn.setEnabled(False)
        self.gq_next_btn.setEnabled(True)
        self.gq_next_btn.setFocus()

    def _append_general_quiz_csv(self, row: Dict[str, Any]) -> None:
        if self.general_quiz_csv_path is None:
            return
        exists = self.general_quiz_csv_path.exists()
        with open(self.general_quiz_csv_path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            if not exists:
                writer.writeheader()
            writer.writerow(row)

    # --------- voice backend ---------
    def choose_backend(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Voice backend 선택", str(APP_DIR), "Python (*.py)")
        if path: self.backend_edit.setText(path)

    def choose_general_quiz_syllable(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "일반 퀴즈용 syllable_top200.xlsx 선택",
            str(APP_DIR),
            "Excel (*.xlsx)",
        )
        if path:
            self.gq_syllable_edit.setText(path)

    def choose_syllable(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Syllable XLSX 선택", str(APP_DIR), "Excel (*.xlsx)")
        if path:
            self.syllable_edit.setText(path)
            self.update_duration_estimate()

    def choose_voice_profiles(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "voice_profiles 폴더 선택", str(APP_DIR))
        if path: self.voice_profiles_edit.setText(path)

    def get_voice_backend(self):
        path = Path(self.backend_edit.text().strip())
        if not path.exists():
            raise RuntimeError(f"Voice backend 파일이 없습니다: {path}")
        if self.voice_backend is not None and Path(getattr(self.voice_backend, "__file__", "")) == path:
            return self.voice_backend
        spec = importlib.util.spec_from_file_location("hangul_voice_backend_dynamic", path)
        if spec is None or spec.loader is None:
            raise RuntimeError("Voice backend를 load할 수 없습니다.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        # The legacy backend uses a module-global BASE_DIR. Redirect it to the
        # folder selected in this designer so existing profiles can be reused
        # without moving files.
        profile_dir = Path(self.voice_profiles_edit.text().strip()) if hasattr(self, "voice_profiles_edit") else APP_DIR / "voice_profiles"
        profile_dir.mkdir(parents=True, exist_ok=True)
        if hasattr(module, "BASE_DIR"):
            module.BASE_DIR = profile_dir
        self.voice_backend = module
        self.voice_backend_path = path
        return module

    def check_voice_model(self) -> None:
        subject = self.subject_edit.text().strip()
        if not subject:
            QMessageBox.warning(self, "Subject", "Subject ID를 입력하세요."); return
        try:
            b = self.get_voice_backend()
            model = b.load_hangul_jamo_model(subject)
            profile_root = Path(self.voice_profiles_edit.text().strip()) if self.voice_profiles_edit.text().strip() else APP_DIR / "voice_profiles"
            possible = [
                profile_root / subject / "jamo_voice_model.joblib",
                profile_root / subject / "hangul_jamo_model.joblib",
                profile_root / subject / "learned_voice_model.joblib",
            ]
            msg = "사용 가능한 모델이 있습니다." if model is not None else "모델이 없습니다. 자모 퀴즈 전 calibration 또는 pooled 모델 생성이 필요합니다."
            msg += "\n\n" + "\n".join(str(p) for p in possible)
            QMessageBox.information(self, "Voice model", msg)
        except Exception as exc:
            QMessageBox.critical(self, "Voice backend 오류", str(exc))

    def build_pooled_model(self) -> None:
        subject = self.subject_edit.text().strip()
        if not subject:
            QMessageBox.warning(self, "Subject", "Subject ID를 입력하세요."); return
        try:
            b = self.get_voice_backend()
            bundle, path = b.train_hangul_jamo_model(subject, labels=b.ALL_JAMO_LABELS, include_pooled=True)
            QMessageBox.information(self, "Pooled model", f"생성 완료\n{path}\nSamples: {bundle.get('n_samples')}")
        except Exception as exc:
            QMessageBox.critical(self, "Pooled model 실패", str(exc))

    # --------- quiz ---------
    def build_quiz_labels(self) -> List[str]:
        qtype = self.quiz_type.currentData()
        compiler = HangulCompiler(self.setup)
        if qtype == "consonant": source = CONSONANTS
        elif qtype == "vowel": source = VOWELS
        elif qtype == "jamo_all": source = CONSONANTS + VOWELS
        else:
            source = load_syllables(Path(self.syllable_edit.text().strip()), self.syllable_limit.value())
        valid: List[str] = []
        for lab in source:
            try:
                if qtype == "syllable": compiler.compile_syllable(lab)
                else: compiler.compile_jamo(lab)
                valid.append(lab)
            except Exception:
                pass
        return valid

    def make_balanced_plan(self, labels: Sequence[str], n: int) -> List[str]:
        if not labels: return []
        plan: List[str] = []
        rng = random.Random(time.time_ns())
        while len(plan) < n:
            batch = list(labels); rng.shuffle(batch); plan.extend(batch)
        return plan[:n]

    def start_quiz(self) -> None:
        subject = self.subject_edit.text().strip()
        if not subject:
            QMessageBox.warning(self, "Subject", "Subject ID를 입력하세요."); return
        try:
            backend = self.get_voice_backend()
            labels = self.build_quiz_labels()
            if len(labels) < 2:
                raise RuntimeError("퀴즈에 사용할 수 있는 label이 2개 미만입니다. 디자인 매핑과 syllable 파일을 확인하세요.")
            if self.quiz_type.currentData() != "syllable" and backend.load_hangul_jamo_model(subject) is None:
                raise RuntimeError("이 subject의 자모 음성 모델이 없습니다. 모델 확인/pooled 모델 생성을 먼저 하세요.")
            self.quiz_labels = labels
            self.quiz_plan = self.make_balanced_plan(labels, self.quiz_trials.value())
            self.quiz_index = -1
            self.quiz_rows = []
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.quiz_csv_path = DATA_DIR / f"quiz_{subject}_{stamp}.csv"
            self.quiz_action.setEnabled(True)
            self.quiz_action.setText("첫 자극 시작 + 음성 듣기")
            self.quiz_instruction.setText("준비되었습니다. 버튼을 누르면 tactile stimulus와 음성 녹음이 시작됩니다.")
            self.quiz_feedback.setText(f"Labels: {len(labels)} · Trials: {len(self.quiz_plan)}")
            self.clear_candidates()
            self.next_quiz_trial()
        except Exception as exc:
            QMessageBox.critical(self, "퀴즈 시작 실패", str(exc))

    def next_quiz_trial(self) -> None:
        self.quiz_index += 1
        if self.quiz_index >= len(self.quiz_plan):
            self.quiz_action.setEnabled(False)
            acc = sum(int(r["correct"]) for r in self.quiz_rows) / max(1, len(self.quiz_rows))
            self.quiz_instruction.setText("퀴즈 완료")
            self.quiz_feedback.setText(f"Accuracy {acc*100:.1f}% · {self.quiz_csv_path}")
            return
        self.quiz_target = self.quiz_plan[self.quiz_index]
        compiler = HangulCompiler(self.setup)
        self.quiz_command = compiler.compile_syllable(self.quiz_target) if self.quiz_type.currentData() == "syllable" else compiler.compile_jamo(self.quiz_target)
        self.quiz_progress.setText(f"TRIAL {self.quiz_index + 1} / {len(self.quiz_plan)}")
        self.quiz_instruction.setText("자극을 느끼고 소리 내어 답해주세요.")
        self.quiz_feedback.setText("")
        self.quiz_action.setText("자극 시작 + 음성 듣기")
        self.quiz_action.setEnabled(True)
        self.quiz_manual.clear(); self.clear_candidates()

    def quiz_action_clicked(self) -> None:
        if self.quiz_index < 0 or self.quiz_index >= len(self.quiz_plan): return
        self.quiz_action.setEnabled(False)
        self.quiz_feedback.setText("자극 전송 중…")
        try:
            self.send_command(self.quiz_command)
            self.quiz_started_perf = time.perf_counter()
            self.start_voice_worker()
        except Exception as exc:
            self.quiz_feedback.setText(f"자극 실패: {exc}")
            self.quiz_action.setEnabled(True)

    def start_voice_worker(self) -> None:
        try:
            b = self.get_voice_backend()
            session_type = "syllable" if self.quiz_type.currentData() == "syllable" else "jamo"
            recorder_settings = dict(
                vad_level=2,
                speech_frames_required=2,
                pre_buffer_frames=20,
                max_wait_sec=8.0,
                max_record_sec=3.0,
                end_silence_sec=0.45,
                noise_calibration_sec=0.25,
                energy_multiplier=2.0,
                min_energy_rms=100.0,
                min_record_sec=0.15,
            )
            self.quiz_worker = b.TrialVoiceWorker(
                self.subject_edit.text().strip(),
                session_type,
                self.quiz_labels,
                recorder_settings,
                syllable_xlsx=self.syllable_edit.text().strip(),
                stt_model="small",
                wav_dir=DATA_DIR / "voice_wav" / self.subject_edit.text().strip(),
            )
            self.quiz_worker.status.connect(self.quiz_feedback.setText)
            self.quiz_worker.result_ready.connect(self.handle_voice_result)
            self.quiz_worker.start()
        except Exception as exc:
            self.quiz_feedback.setText(f"음성 시작 실패: {exc}")
            self.quiz_action.setEnabled(True)

    def extract_candidates(self, result: Dict[str, Any]) -> List[str]:
        out: List[str] = []
        def add(x):
            if x is None: return
            if isinstance(x, dict): lab = x.get("label") or x.get("predicted_label")
            else: lab = x
            if lab is not None and str(lab).strip() and str(lab).strip() not in out: out.append(str(lab).strip())
        add(result.get("predicted_label")); add(result.get("second_label"))
        for key in ("top", "top_candidates", "candidates", "scores"):
            vals = result.get(key) or []
            for x in vals: add(x)
        if self.quiz_type.currentData() == "syllable":
            allowed = set(self.quiz_labels); out = [x for x in out if x in allowed]
        return out[:5]

    def handle_voice_result(self, result: Dict[str, Any]) -> None:
        if "error" in result:
            self.quiz_feedback.setText(f"음성 인식 실패: {result.get('message', result['error'])}")
            self.quiz_action.setText("다시 듣기")
            self.quiz_action.setEnabled(True)
            return
        self.current_voice_result = result
        candidates = self.extract_candidates(result)
        self.quiz_feedback.setText("음성 후보를 확인하고 실제 응답을 선택하세요.")
        self.show_candidates(candidates)

    def clear_candidates(self) -> None:
        clear_layout(self.candidate_layout)

    def show_candidates(self, labels: Sequence[str]) -> None:
        self.clear_candidates()
        for i, lab in enumerate(labels):
            b = QPushButton(f"{i+1}.  {lab}")
            b.setObjectName("HardwareButton")
            b.setMinimumHeight(70)
            b.clicked.connect(lambda _=False, x=lab: self.finalize_quiz_answer(x, manual=False))
            self.candidate_layout.addWidget(b, 0, i)

    def submit_manual_quiz(self) -> None:
        ans = self.quiz_manual.text().strip()
        if ans: self.finalize_quiz_answer(ans, manual=True)

    def finalize_quiz_answer(self, answer: str, manual: bool) -> None:
        correct = int(answer == self.quiz_target)
        elapsed_ms = (time.perf_counter() - self.quiz_started_perf) * 1000 if self.quiz_started_perf else ""
        result = getattr(self, "current_voice_result", {}) or {}
        row = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "app_version": APP_VERSION,
            "subject": self.subject_edit.text().strip(),
            "setup_name": self.setup.setup_name,
            "setup_hash": self.setup.stable_hash(),
            "setup_path": str(self.setup_path or ""),
            "quiz_type": self.quiz_type.currentData(),
            "trial_index": self.quiz_index + 1,
            "target": self.quiz_target,
            "response": answer,
            "manual_response": int(manual),
            "correct": correct,
            "command": self.quiz_command,
            "elapsed_ms": elapsed_ms,
            "voice_engine": result.get("engine", ""),
            "voice_wav_path": result.get("wav_path", ""),
            "voice_stt_raw": result.get("stt_raw", ""),
            "voice_stt_norm": result.get("stt_norm", ""),
            "voice_onset_rt_sec": result.get("voice_onset_rt_sec", ""),
            "top_candidates_json": json.dumps(result.get("top", result.get("top_candidates", [])), ensure_ascii=False),
        }
        self.quiz_rows.append(row)
        self.append_quiz_csv(row)
        self.clear_candidates()
        if correct:
            self.quiz_feedback.setText(f"정답  {self.quiz_target}")
            self.quiz_feedback.setStyleSheet("background:#EAF8EF;color:#147A3D;border-radius:16px;padding:12px;font-weight:700;")
        else:
            self.quiz_feedback.setText(f"오답 · 응답 {answer}  →  정답 {self.quiz_target}")
            self.quiz_feedback.setStyleSheet("background:#FFF0F0;color:#B3261E;border-radius:16px;padding:12px;font-weight:700;")
        self.quiz_action.setText("다음")
        self.quiz_action.setEnabled(True)
        try: self.quiz_action.clicked.disconnect()
        except Exception: pass
        self.quiz_action.clicked.connect(self.quiz_next_clicked)

    def quiz_next_clicked(self) -> None:
        try: self.quiz_action.clicked.disconnect()
        except Exception: pass
        self.quiz_action.clicked.connect(self.quiz_action_clicked)
        self.quiz_feedback.setStyleSheet("")
        self.next_quiz_trial()

    def append_quiz_csv(self, row: Dict[str, Any]) -> None:
        assert self.quiz_csv_path is not None
        exists = self.quiz_csv_path.exists()
        with open(self.quiz_csv_path, "a", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(row.keys()))
            if not exists: w.writeheader()
            w.writerow(row)

    def closeEvent(self, event) -> None:  # noqa: N802
        self.serial_ctl.emergency_off(); self.serial_ctl.close()
        if self.dirty:
            reply = QMessageBox.question(self, "종료", "저장하지 않은 디자인 변경사항이 있습니다. 종료할까요?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                event.ignore(); return
        event.accept()


def main() -> None:
    app = QApplication(sys.argv)
    app.setApplicationName("Hangul Tactile Designer")
    win = MainWindow()
    win.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
